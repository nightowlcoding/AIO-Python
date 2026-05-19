import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
# from openpyxl.chart import BarChart, Reference
from datetime import datetime

LOCATIONS = ["Kingsville", "Alice"]
WEEK_SHEETS = ["Q1W1", "Q1W2", "Q1W3", "Q1W4", "Q1W5"]


def _to_number(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.strip().replace("$", "").replace(",", "")
        if cleaned in {"", "-", "--"}:
            return 0.0
        try:
            return float(cleaned)
        except ValueError:
            return 0.0
    return 0.0

def _find_location_blocks(data_ws):
    forecast_rows = []
    for row in data_ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip() == "Total Week Forecast":
                forecast_rows.append(cell.row)

    blocks = []
    for idx, forecast_row in enumerate(sorted(forecast_rows)):
        location = None
        for r in range(max(1, forecast_row - 3), forecast_row + 1):
            for c in range(1, data_ws.max_column + 1):
                value = data_ws.cell(row=r, column=c).value
                if isinstance(value, str) and value.strip() in LOCATIONS:
                    location = value.strip()
                    break
            if location:
                break

        end_row = (forecast_rows[idx + 1] - 1) if idx + 1 < len(forecast_rows) else data_ws.max_row
        blocks.append((location, forecast_row, end_row))

    return blocks

def _find_shift_totals(data_ws, start_row, end_row):
    header_row = None
    header_cols = {}
    for r in range(start_row, min(end_row, start_row + 12) + 1):
        labels = [(c, data_ws.cell(row=r, column=c).value) for c in range(1, data_ws.max_column + 1)]
        label_set = {v.strip() for _, v in labels if isinstance(v, str)}
        if {"Beer", "Liquor", "Wine", "Total"}.issubset(label_set) and any(v in ["Food/NA", "Food / NA", "Food"] for v in label_set):
            header_row = r
            for c, v in labels:
                if isinstance(v, str):
                    header_cols[v.strip()] = c
            break

    if header_row is None:
        return None, None

    total_col = header_cols.get("Total")
    morning_total = None
    night_total = None
    for r in range(header_row + 1, min(end_row, header_row + 6) + 1):
        row_values = [data_ws.cell(row=r, column=c).value for c in range(1, data_ws.max_column + 1)]
        labels = [v.strip() for v in row_values if isinstance(v, str)]
        if "Morning" in labels:
            morning_total = data_ws.cell(row=r, column=total_col).value
        if "Night" in labels:
            night_total = data_ws.cell(row=r, column=total_col).value

    return morning_total, night_total

def _find_shift_by_day(data_ws, start_row, end_row):
    day_names = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
    day_cols = {}
    morning = {}
    night = {}

    sales_row = None
    sales_col = None
    for r in range(start_row, min(end_row, start_row + 12) + 1):
        for c in range(1, data_ws.max_column + 1):
            value = data_ws.cell(row=r, column=c).value
            if isinstance(value, str) and value.strip() == "Sales by Shift":
                sales_row = r
                sales_col = c
                break
        if sales_row:
            break

    if sales_row is None:
        return morning, night, None, None

    header_row = None
    search_end_row = min(end_row, sales_row + 6)
    for r in range(sales_row + 1, search_end_row + 1):
        for start_col in range((sales_col or 1) + 1, data_ws.max_column - 5):
            values = []
            for offset in range(7):
                raw = data_ws.cell(row=r, column=start_col + offset).value
                values.append(raw.strip() if isinstance(raw, str) else raw)
            if values == day_names:
                header_row = r
                day_cols = {day: start_col + idx for idx, day in enumerate(day_names)}
                break
        if header_row:
            break

    if header_row is None:
        return morning, night, None, None

    morning_row = None
    night_row = None
    for r in range(header_row + 1, min(end_row, header_row + 5) + 1):
        value = data_ws.cell(row=r, column=sales_col).value if sales_col else None
        if isinstance(value, str):
            label = value.strip().lower()
            if label == "morning" and morning_row is None:
                morning_row = r
            elif label == "night" and night_row is None:
                night_row = r

    if morning_row:
        for day in day_names:
            day_col = day_cols.get(day)
            if day_col:
                morning[day] = _to_number(data_ws.cell(row=morning_row, column=day_col).value)

    if night_row:
        for day in day_names:
            day_col = day_cols.get(day)
            if day_col:
                night[day] = _to_number(data_ws.cell(row=night_row, column=day_col).value)

    return morning, night, morning_row, night_row

def _find_categories_by_shift(data_ws, morning_row, night_row):
    if not morning_row or not night_row:
        return {"Morning": {}, "Night": {}}

    col_map = {
        "Beer": 34,    # AH
        "Liquor": 35,  # AI
        "Wine": 36,    # AJ
        "Food/NA": 37, # AK
    }

    categories = {"Morning": {}, "Night": {}}
    for label, row in [("Morning", morning_row), ("Night", night_row)]:
        for key, col in col_map.items():
            value = data_ws.cell(row=row, column=col).value
            categories[label][key] = _to_number(value)

    return categories

def read_weekly_locations(data_ws, sheet_name):
    """Extract totals for each location within a weekly sheet."""
    results = []
    for location, forecast_row, end_row in _find_location_blocks(data_ws):
        if not location:
            continue

        forecast = None
        actual = None
        for row in data_ws.iter_rows(min_row=forecast_row, max_row=end_row):
            for cell in row:
                if isinstance(cell.value, str):
                    label = cell.value.strip()
                    if label == "Total Week Forecast":
                        value = data_ws.cell(row=cell.row + 1, column=cell.column).value
                        if value is not None:
                            forecast = _to_number(value)
                    elif label == "Total Week Actual":
                        value = data_ws.cell(row=cell.row + 1, column=cell.column).value
                        if value is not None:
                            actual = _to_number(value)

        morning_total, night_total = _find_shift_totals(data_ws, forecast_row, end_row)
        morning_by_day, night_by_day, morning_row, night_row = _find_shift_by_day(data_ws, forecast_row, end_row)
        categories_by_shift = _find_categories_by_shift(data_ws, morning_row, night_row)

        if forecast is not None or actual is not None:
            difference = (actual or 0) - (forecast or 0)
            variance_pct = (difference / forecast) if forecast else 0
            actual_value = actual or 0
            
            # Calculate morning and night totals from shift by day data
            morning_value = sum(morning_by_day.values()) if morning_by_day else 0
            night_value = sum(night_by_day.values()) if night_by_day else 0
            
            results.append({
                "location": location,
                "forecast": forecast or 0,
                "actual": actual_value,
                "difference": difference,
                "variance_pct": variance_pct,
                "morning_total": morning_value,
                "night_total": night_value,
                "morning_pct": (morning_value / actual_value) if actual_value else 0,
                "night_pct": (night_value / actual_value) if actual_value else 0,
                "morning_by_day": morning_by_day,
                "night_by_day": night_by_day,
                "categories_by_shift": categories_by_shift,
                "morning_range": f"{sheet_name}!R{morning_row}:X{morning_row}" if morning_row else "",
                "night_range": f"{sheet_name}!R{night_row}:X{night_row}" if night_row else "",
            })

    return results

def create_overview_sheet(source_file, output_file):
    """Create overview in a new file without touching weekly sheets."""
    data_wb = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    write_wb = openpyxl.load_workbook(source_file)

    if "Overview" in write_wb.sheetnames:
        old_index = write_wb.sheetnames.index("Overview")
        write_wb.remove(write_wb["Overview"])
        ws = write_wb.create_sheet("Overview", old_index)
    else:
        ws = write_wb.create_sheet("Overview", 0)

    weekly_data = []
    for sheet_name in WEEK_SHEETS:
        if sheet_name not in data_wb.sheetnames:
            continue
        data_ws = data_wb[sheet_name]
        for row in read_weekly_locations(data_ws, sheet_name):
            row["week"] = sheet_name
            weekly_data.append(row)

    category_index = {}
    for row in weekly_data:
        week = row["week"]
        location = row["location"]
        for shift_label in ["Morning", "Night"]:
            category_index[(location, week, shift_label)] = row["categories_by_shift"].get(shift_label, {})

    ws["A1"] = "Q1 2026 QUARTERLY REPORT OVERVIEW"
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A1:G1")

    ws["A2"] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:G2")

    current_row = 4
    week_order = {name: idx for idx, name in enumerate(WEEK_SHEETS)}
    weekly_data.sort(key=lambda x: (x["location"], week_order.get(x["week"], 9999), x["week"]))
    for location in LOCATIONS:
        rows = [d for d in weekly_data if d["location"] == location]
        if not rows:
            continue

        ws[f"A{current_row}"] = f"{location.upper()} SUMMARY"
        ws[f"A{current_row}"].font = Font(size=14, bold=True)
        ws.merge_cells(f"A{current_row}:I{current_row}")

        current_row += 2
        headers = [
            "Week",
            "Forecast ($)",
            "Actual ($)",
            "Difference ($)",
            "Variance (%)",
            "Morning (%)",
            "Night (%)",
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        table_start = current_row + 1
        current_row += 1
        for data in rows:
            ws.cell(row=current_row, column=1, value=data["week"])
            ws.cell(row=current_row, column=2, value=data["forecast"]).number_format = "$#,##0.00"
            ws.cell(row=current_row, column=3, value=data["actual"]).number_format = "$#,##0.00"
            ws.cell(row=current_row, column=4, value=data["difference"]).number_format = "$#,##0.00"
            ws.cell(row=current_row, column=5, value=data["variance_pct"]).number_format = "0.00%"
            ws.cell(row=current_row, column=6, value=data["morning_pct"]).number_format = "0.00%"
            ws.cell(row=current_row, column=7, value=data["night_pct"]).number_format = "0.00%"
            current_row += 1

        total_forecast = sum(d["forecast"] for d in rows)
        total_actual = sum(d["actual"] for d in rows)
        total_diff = total_actual - total_forecast
        total_variance = (total_diff / total_forecast) if total_forecast else 0
        total_morning = sum(d["morning_total"] for d in rows)
        total_night = sum(d["night_total"] for d in rows)
        total_morning_pct = (total_morning / total_actual) if total_actual else 0
        total_night_pct = (total_night / total_actual) if total_actual else 0

        ws.cell(row=current_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=current_row, column=2, value=total_forecast).number_format = "$#,##0.00"
        ws.cell(row=current_row, column=2).font = Font(bold=True)
        ws.cell(row=current_row, column=3, value=total_actual).number_format = "$#,##0.00"
        ws.cell(row=current_row, column=3).font = Font(bold=True)
        ws.cell(row=current_row, column=4, value=total_diff).number_format = "$#,##0.00"
        ws.cell(row=current_row, column=4).font = Font(bold=True)
        ws.cell(row=current_row, column=5, value=total_variance).number_format = "0.00%"
        ws.cell(row=current_row, column=5).font = Font(bold=True)
        ws.cell(row=current_row, column=6, value=total_morning_pct).number_format = "0.00%"
        ws.cell(row=current_row, column=6).font = Font(bold=True)
        ws.cell(row=current_row, column=7, value=total_night_pct).number_format = "0.00%"
        ws.cell(row=current_row, column=7).font = Font(bold=True)

        current_row += 2
        # CHART DISABLED - Morning vs Night by Week
        # chart = BarChart()
        # chart.title = f"{location}: Morning vs Night by Week"
        # chart.style = 10
        # chart.y_axis.title = "Sales ($)"
        # chart.x_axis.title = "Week"
        # data = Reference(ws, min_col=6, min_row=table_start - 1, max_row=table_start + len(rows) - 1, max_col=7)
        # cats = Reference(ws, min_col=1, min_row=table_start, max_row=table_start + len(rows) - 1)
        # chart.add_data(data, titles_from_data=True)
        # chart.set_categories(cats)
        # chart.height = 10
        # chart.width = 20
        # ws.add_chart(chart, f"A{current_row}")
        # current_row += 14

        current_row += 2  # spacing instead of chart space

        # Morning Shift by Day Section
        ws[f"A{current_row}"] = f"{location} SHIFT BY DAY - MORNING"
        ws[f"A{current_row}"].font = Font(size=12, bold=True)
        ws.merge_cells(f"A{current_row}:I{current_row}")
        current_row += 1

        day_headers = ["Week", "Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Total"]
        for col_idx, header in enumerate(day_headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="6D9EEB", end_color="6D9EEB", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        current_row += 1
        # All Morning shifts
        day_totals = {day: 0 for day in day_headers[1:-1]}  # Exclude Week and Total
        row_count = 0
        for data in rows:
            ws.cell(row=current_row, column=1, value=data["week"])
            week_total = 0
            for idx, day in enumerate(day_headers[1:-1], start=2):  # Exclude 'Week' and 'Total'
                value = data["morning_by_day"].get(day, 0)
                ws.cell(row=current_row, column=idx, value=value).number_format = "$#,##0.00"
                week_total += value
                day_totals[day] += value
            # Add total column
            ws.cell(row=current_row, column=len(day_headers), value=week_total).number_format = "$#,##0.00"
            ws.cell(row=current_row, column=len(day_headers)).font = Font(bold=True)
            current_row += 1
            row_count += 1

        # Add Total row
        ws.cell(row=current_row, column=1, value="Total").font = Font(bold=True)
        grand_total = 0
        for idx, day in enumerate(day_headers[1:-1], start=2):
            ws.cell(row=current_row, column=idx, value=day_totals[day]).number_format = "$#,##0.00"
            ws.cell(row=current_row, column=idx).font = Font(bold=True)
            grand_total += day_totals[day]
        ws.cell(row=current_row, column=len(day_headers), value=grand_total).number_format = "$#,##0.00"
        ws.cell(row=current_row, column=len(day_headers)).font = Font(bold=True)
        current_row += 1

        # Add Average row
        ws.cell(row=current_row, column=1, value="Average").font = Font(bold=True, italic=True)
        for idx, day in enumerate(day_headers[1:-1], start=2):
            avg_value = day_totals[day] / row_count if row_count > 0 else 0
            ws.cell(row=current_row, column=idx, value=avg_value).number_format = "$#,##0.00"
            ws.cell(row=current_row, column=idx).font = Font(bold=True, italic=True)
        avg_total = grand_total / row_count if row_count > 0 else 0
        ws.cell(row=current_row, column=len(day_headers), value=avg_total).number_format = "$#,##0.00"
        ws.cell(row=current_row, column=len(day_headers)).font = Font(bold=True, italic=True)
        current_row += 1

        current_row += 2

        # Night Shift by Day Section
        ws[f"A{current_row}"] = f"{location} SHIFT BY DAY - NIGHT"
        ws[f"A{current_row}"].font = Font(size=12, bold=True)
        ws.merge_cells(f"A{current_row}:I{current_row}")
        current_row += 1

        for col_idx, header in enumerate(day_headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="6D9EEB", end_color="6D9EEB", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        current_row += 1
        # All Night shifts
        day_totals = {day: 0 for day in day_headers[1:-1]}  # Exclude Week and Total
        row_count = 0
        for data in rows:
            ws.cell(row=current_row, column=1, value=data["week"])
            week_total = 0
            for idx, day in enumerate(day_headers[1:-1], start=2):  # Exclude 'Week' and 'Total'
                value = data["night_by_day"].get(day, 0)
                ws.cell(row=current_row, column=idx, value=value).number_format = "$#,##0.00"
                week_total += value
                day_totals[day] += value
            # Add total column
            ws.cell(row=current_row, column=len(day_headers), value=week_total).number_format = "$#,##0.00"
            ws.cell(row=current_row, column=len(day_headers)).font = Font(bold=True)
            current_row += 1
            row_count += 1

        # Add Total row
        ws.cell(row=current_row, column=1, value="Total").font = Font(bold=True)
        grand_total = 0
        for idx, day in enumerate(day_headers[1:-1], start=2):
            ws.cell(row=current_row, column=idx, value=day_totals[day]).number_format = "$#,##0.00"
            ws.cell(row=current_row, column=idx).font = Font(bold=True)
            grand_total += day_totals[day]
        ws.cell(row=current_row, column=len(day_headers), value=grand_total).number_format = "$#,##0.00"
        ws.cell(row=current_row, column=len(day_headers)).font = Font(bold=True)
        current_row += 1

        # Add Average row
        ws.cell(row=current_row, column=1, value="Average").font = Font(bold=True, italic=True)
        for idx, day in enumerate(day_headers[1:-1], start=2):
            avg_value = day_totals[day] / row_count if row_count > 0 else 0
            ws.cell(row=current_row, column=idx, value=avg_value).number_format = "$#,##0.00"
            ws.cell(row=current_row, column=idx).font = Font(bold=True, italic=True)
        avg_total = grand_total / row_count if row_count > 0 else 0
        ws.cell(row=current_row, column=len(day_headers), value=avg_total).number_format = "$#,##0.00"
        ws.cell(row=current_row, column=len(day_headers)).font = Font(bold=True, italic=True)
        current_row += 1

        current_row += 2

        # Combined Shift by Day Section
        ws[f"A{current_row}"] = f"{location} SHIFT BY DAY - COMBINED"
        ws[f"A{current_row}"].font = Font(size=12, bold=True)
        ws.merge_cells(f"A{current_row}:I{current_row}")
        current_row += 1

        for col_idx, header in enumerate(day_headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="6D9EEB", end_color="6D9EEB", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        current_row += 1
        # Combined Morning + Night shifts
        day_totals = {day: 0 for day in day_headers[1:-1]}  # Exclude Week and Total
        row_count = 0
        for data in rows:
            ws.cell(row=current_row, column=1, value=data["week"])
            week_total = 0
            for idx, day in enumerate(day_headers[1:-1], start=2):
                morning_value = data["morning_by_day"].get(day, 0)
                night_value = data["night_by_day"].get(day, 0)
                combined_value = morning_value + night_value
                ws.cell(row=current_row, column=idx, value=combined_value).number_format = "$#,##0.00"
                week_total += combined_value
                day_totals[day] += combined_value
            # Add total column
            ws.cell(row=current_row, column=len(day_headers), value=week_total).number_format = "$#,##0.00"
            ws.cell(row=current_row, column=len(day_headers)).font = Font(bold=True)
            current_row += 1
            row_count += 1

        # Add Total row
        ws.cell(row=current_row, column=1, value="Total").font = Font(bold=True)
        grand_total = 0
        for idx, day in enumerate(day_headers[1:-1], start=2):
            ws.cell(row=current_row, column=idx, value=day_totals[day]).number_format = "$#,##0.00"
            ws.cell(row=current_row, column=idx).font = Font(bold=True)
            grand_total += day_totals[day]
        ws.cell(row=current_row, column=len(day_headers), value=grand_total).number_format = "$#,##0.00"
        ws.cell(row=current_row, column=len(day_headers)).font = Font(bold=True)
        current_row += 1

        # Add Average row
        ws.cell(row=current_row, column=1, value="Average").font = Font(bold=True, italic=True)
        for idx, day in enumerate(day_headers[1:-1], start=2):
            avg_value = day_totals[day] / row_count if row_count > 0 else 0
            ws.cell(row=current_row, column=idx, value=avg_value).number_format = "$#,##0.00"
            ws.cell(row=current_row, column=idx).font = Font(bold=True, italic=True)
        avg_total = grand_total / row_count if row_count > 0 else 0
        ws.cell(row=current_row, column=len(day_headers), value=avg_total).number_format = "$#,##0.00"
        ws.cell(row=current_row, column=len(day_headers)).font = Font(bold=True, italic=True)
        current_row += 1

        current_row += 2

        # Morning Category Section
        ws[f"A{current_row}"] = f"{location} CATEGORY BY SHIFT - MORNING"
        ws[f"A{current_row}"].font = Font(size=12, bold=True)
        ws.merge_cells(f"A{current_row}:I{current_row}")
        current_row += 1

        cat_headers = [
            "Week",
            "Beer",
            "Liquor",
            "Wine",
            "Food/NA",
        ]
        cat_table_start = current_row
        for col_idx, header in enumerate(cat_headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="93C47D", end_color="93C47D", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        current_row += 1
        # All Morning shifts
        cat_totals = {cat: 0 for cat in cat_headers[1:]}  # Exclude Week
        row_count = 0
        for data in rows:
            ws.cell(row=current_row, column=1, value=data["week"])
            beer_val = data["categories_by_shift"].get("Morning", {}).get("Beer", 0)
            liquor_val = data["categories_by_shift"].get("Morning", {}).get("Liquor", 0)
            wine_val = data["categories_by_shift"].get("Morning", {}).get("Wine", 0)
            food_val = data["categories_by_shift"].get("Morning", {}).get("Food/NA", 0)
            
            ws.cell(row=current_row, column=2, value=beer_val).number_format = "$#,##0.00"
            ws.cell(row=current_row, column=3, value=liquor_val).number_format = "$#,##0.00"
            ws.cell(row=current_row, column=4, value=wine_val).number_format = "$#,##0.00"
            ws.cell(row=current_row, column=5, value=food_val).number_format = "$#,##0.00"
            
            cat_totals["Beer"] += beer_val
            cat_totals["Liquor"] += liquor_val
            cat_totals["Wine"] += wine_val
            cat_totals["Food/NA"] += food_val
            current_row += 1
            row_count += 1

        # Add Total row
        ws.cell(row=current_row, column=1, value="Total").font = Font(bold=True)
        for idx, cat in enumerate(cat_headers[1:], start=2):
            ws.cell(row=current_row, column=idx, value=cat_totals[cat]).number_format = "$#,##0.00"
            ws.cell(row=current_row, column=idx).font = Font(bold=True)
        current_row += 1

        # Add Average row
        ws.cell(row=current_row, column=1, value="Average").font = Font(bold=True, italic=True)
        for idx, cat in enumerate(cat_headers[1:], start=2):
            avg_value = cat_totals[cat] / row_count if row_count > 0 else 0
            ws.cell(row=current_row, column=idx, value=avg_value).number_format = "$#,##0.00"
            ws.cell(row=current_row, column=idx).font = Font(bold=True, italic=True)
        current_row += 1

        current_row += 2

        # Night Category Section
        ws[f"A{current_row}"] = f"{location} CATEGORY BY SHIFT - NIGHT"
        ws[f"A{current_row}"].font = Font(size=12, bold=True)
        ws.merge_cells(f"A{current_row}:I{current_row}")
        current_row += 1

        for col_idx, header in enumerate(cat_headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="93C47D", end_color="93C47D", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        current_row += 1
        # All Night shifts
        cat_totals = {cat: 0 for cat in cat_headers[1:]}  # Exclude Week
        row_count = 0
        for data in rows:
            ws.cell(row=current_row, column=1, value=data["week"])
            beer_val = data["categories_by_shift"].get("Night", {}).get("Beer", 0)
            liquor_val = data["categories_by_shift"].get("Night", {}).get("Liquor", 0)
            wine_val = data["categories_by_shift"].get("Night", {}).get("Wine", 0)
            food_val = data["categories_by_shift"].get("Night", {}).get("Food/NA", 0)
            
            ws.cell(row=current_row, column=2, value=beer_val).number_format = "$#,##0.00"
            ws.cell(row=current_row, column=3, value=liquor_val).number_format = "$#,##0.00"
            ws.cell(row=current_row, column=4, value=wine_val).number_format = "$#,##0.00"
            ws.cell(row=current_row, column=5, value=food_val).number_format = "$#,##0.00"
            
            cat_totals["Beer"] += beer_val
            cat_totals["Liquor"] += liquor_val
            cat_totals["Wine"] += wine_val
            cat_totals["Food/NA"] += food_val
            current_row += 1
            row_count += 1

        # Add Total row
        ws.cell(row=current_row, column=1, value="Total").font = Font(bold=True)
        for idx, cat in enumerate(cat_headers[1:], start=2):
            ws.cell(row=current_row, column=idx, value=cat_totals[cat]).number_format = "$#,##0.00"
            ws.cell(row=current_row, column=idx).font = Font(bold=True)
        current_row += 1

        # Add Average row
        ws.cell(row=current_row, column=1, value="Average").font = Font(bold=True, italic=True)
        for idx, cat in enumerate(cat_headers[1:], start=2):
            avg_value = cat_totals[cat] / row_count if row_count > 0 else 0
            ws.cell(row=current_row, column=idx, value=avg_value).number_format = "$#,##0.00"
            ws.cell(row=current_row, column=idx).font = Font(bold=True, italic=True)
        current_row += 1

        current_row += 2

        # Combined Category Section
        ws[f"A{current_row}"] = f"{location} CATEGORY BY SHIFT - COMBINED"
        ws[f"A{current_row}"].font = Font(size=12, bold=True)
        ws.merge_cells(f"A{current_row}:I{current_row}")
        current_row += 1

        for col_idx, header in enumerate(cat_headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="93C47D", end_color="93C47D", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        current_row += 1
        # Combined Morning + Night categories
        cat_totals = {cat: 0 for cat in cat_headers[1:]}  # Exclude Week
        row_count = 0
        for data in rows:
            ws.cell(row=current_row, column=1, value=data["week"])
            morning_beer = data["categories_by_shift"].get("Morning", {}).get("Beer", 0)
            night_beer = data["categories_by_shift"].get("Night", {}).get("Beer", 0)
            beer_val = morning_beer + night_beer
            ws.cell(row=current_row, column=2, value=beer_val).number_format = "$#,##0.00"
            
            morning_liquor = data["categories_by_shift"].get("Morning", {}).get("Liquor", 0)
            night_liquor = data["categories_by_shift"].get("Night", {}).get("Liquor", 0)
            liquor_val = morning_liquor + night_liquor
            ws.cell(row=current_row, column=3, value=liquor_val).number_format = "$#,##0.00"
            
            morning_wine = data["categories_by_shift"].get("Morning", {}).get("Wine", 0)
            night_wine = data["categories_by_shift"].get("Night", {}).get("Wine", 0)
            wine_val = morning_wine + night_wine
            ws.cell(row=current_row, column=4, value=wine_val).number_format = "$#,##0.00"
            
            morning_food = data["categories_by_shift"].get("Morning", {}).get("Food/NA", 0)
            night_food = data["categories_by_shift"].get("Night", {}).get("Food/NA", 0)
            food_val = morning_food + night_food
            ws.cell(row=current_row, column=5, value=food_val).number_format = "$#,##0.00"
            
            cat_totals["Beer"] += beer_val
            cat_totals["Liquor"] += liquor_val
            cat_totals["Wine"] += wine_val
            cat_totals["Food/NA"] += food_val
            current_row += 1
            row_count += 1

        # Add Total row
        ws.cell(row=current_row, column=1, value="Total").font = Font(bold=True)
        for idx, cat in enumerate(cat_headers[1:], start=2):
            ws.cell(row=current_row, column=idx, value=cat_totals[cat]).number_format = "$#,##0.00"
            ws.cell(row=current_row, column=idx).font = Font(bold=True)
        current_row += 1

        # Add Average row
        ws.cell(row=current_row, column=1, value="Average").font = Font(bold=True, italic=True)
        for idx, cat in enumerate(cat_headers[1:], start=2):
            avg_value = cat_totals[cat] / row_count if row_count > 0 else 0
            ws.cell(row=current_row, column=idx, value=avg_value).number_format = "$#,##0.00"
            ws.cell(row=current_row, column=idx).font = Font(bold=True, italic=True)
        current_row += 1

        cat_table_end = current_row - 1
        current_row += 1

        ws[f"A{current_row}"] = f"{location} BEST/WORST DAY BY SHIFT"
        ws[f"A{current_row}"].font = Font(size=12, bold=True)
        ws.merge_cells(f"A{current_row}:I{current_row}")
        current_row += 1

        summary_headers = ["Shift", "Best Day", "Best Avg ($)", "Worst Day", "Worst Avg ($)"]
        for col_idx, header in enumerate(summary_headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="B4A7D6", end_color="B4A7D6", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        current_row += 1
        day_names = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
        for shift_label, key in [("Morning", "morning_by_day"), ("Night", "night_by_day")]:
            totals = {day: 0 for day in day_names}
            counts = {day: 0 for day in day_names}
            for data in rows:
                for day in day_names:
                    value = data[key].get(day, 0)
                    totals[day] += value
                    counts[day] += 1

            averages = {day: (totals[day] / counts[day]) if counts[day] else 0 for day in day_names}
            best_day = max(averages, key=averages.get)
            worst_day = min(averages, key=averages.get)

            ws.cell(row=current_row, column=1, value=shift_label)
            ws.cell(row=current_row, column=2, value=best_day)
            ws.cell(row=current_row, column=3, value=averages[best_day]).number_format = "$#,##0.00"
            ws.cell(row=current_row, column=4, value=worst_day)
            ws.cell(row=current_row, column=5, value=averages[worst_day]).number_format = "$#,##0.00"
            current_row += 1

        current_row += 2

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["J"].width = 18

    # Apply center alignment to all cells in the worksheet
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    write_wb.save(output_file)

if __name__ == "__main__":
    SOURCE_FILE = r"C:\Users\arnol\OneDrive\Desktop\Quarterly Report 2026 Copy.xlsx"
    OUTPUT_FILE = r"C:\Users\arnol\OneDrive\Desktop\Quarterly Report 2026 - Updated.xlsx"

    print("=" * 60)
    print("QUARTERLY REPORT AUTOMATION (OVERVIEW ONLY)")
    print("=" * 60)
    print("Source:", SOURCE_FILE)
    print("Output:", OUTPUT_FILE)

    create_overview_sheet(SOURCE_FILE, OUTPUT_FILE)

    print("\n" + "=" * 60)
    print("Automation complete.")
    print("Overview sheet updated in the new file; weekly sheets preserved.")
