from flask import Flask, render_template, request, send_file, redirect, url_for, flash
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from werkzeug.utils import secure_filename



app = Flask(__name__)
app.secret_key = 'payroll_secret_key'
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
RESULTS_FOLDER = os.path.join(BASE_DIR, 'results')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(RESULTS_FOLDER, exist_ok=True)

def process_payroll(csv_path):
    df = pd.read_csv(csv_path)
    columns_to_extract = [
        'Employee', 'Job Title', 'Regular Hours', 'Overtime Hours',
        'Declared Tips', 'Non-Cash Tips', 'Total Tips'
    ]
    extracted_data = df[columns_to_extract]

    def reformat_employee_name(name):
        parts = [part.strip() for part in str(name).split(',')]
        if len(parts) == 2:
            return f"{parts[1]} {parts[0]}"
        return name

    extracted_data.loc[:, 'Employee'] = extracted_data['Employee'].apply(reformat_employee_name)

    # Set all columns to zero for MANAGER job titles, except for 'April Salinas' as MANAGER
    manager_mask = (
        (extracted_data['Job Title'].str.strip().str.upper() == 'MANAGER') &
        ~((extracted_data['Employee'].str.strip().str.upper() == 'APRIL SALINAS') & (extracted_data['Job Title'].str.strip().str.upper() == 'MANAGER'))
    )
    for col in ['Regular Hours', 'Overtime Hours', 'Declared Tips', 'Non-Cash Tips', 'Total Tips']:
        extracted_data.loc[manager_mask, col] = 0

    # Add custom employees
    new_row = {
        'Employee': 'Arnold Ramirez',
        'Job Title': 'General Manager',
        'Regular Hours': 0,
        'Overtime Hours': 0,
        'Declared Tips': 0,
        'Non-Cash Tips': 0,
        'Total Tips': 0
    }
    new_row2 = {
        'Employee': 'Victor Moreno',
        'Job Title': 'Director of Catering',
        'Regular Hours': 0,
        'Overtime Hours': 0,
        'Declared Tips': 0,
        'Non-Cash Tips': 0,
        'Total Tips': 0
    }
    extracted_data = pd.concat([extracted_data, pd.DataFrame([new_row, new_row2])], ignore_index=True)

    # Calculate totals
    totals = extracted_data.iloc[:, 2:].sum()
    total_combined_hours = float(totals['Regular Hours']) + float(totals['Overtime Hours'])
    totals_row = pd.DataFrame([{
        'Employee': 'Total',
        'Job Title': total_combined_hours,
        'Regular Hours': totals['Regular Hours'],
        'Overtime Hours': totals['Overtime Hours'],
        'Declared Tips': totals['Declared Tips'],
        'Non-Cash Tips': totals['Non-Cash Tips'],
        'Total Tips': totals['Total Tips']
    }])

    sorted_data = extracted_data.sort_values(by='Employee')
    sorted_data = pd.concat([
        sorted_data, totals_row
    ], ignore_index=True)

    reg_hours = totals['Regular Hours']
    ot_hours = totals['Overtime Hours']
    total_hours = total_combined_hours
    ot_pct = (ot_hours / total_hours) if total_hours > 0 else 0
    # Replace zeros with empty string for display
    sorted_data = sorted_data.replace(0, "")
    return reg_hours, ot_hours, total_hours, ot_pct, sorted_data

def save_to_excel(df, out_path):
    from openpyxl.styles import PatternFill, Font
    df.to_excel(out_path, index=False)
    # Apply borders and fit to one page
    wb = load_workbook(out_path)
    ws = wb.active
    max_row = ws.max_row
    max_col = ws.max_column
    thick = Side(border_style="thick", color="000000")
    thin = Side(border_style="thin", color="000000")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    bold_font = Font(bold=True)

    # Find duplicate employee names (consecutive rows)
    employee_col = 1  # 1-based index for openpyxl
    for row in range(2, max_row):  # skip header, 1-based
        emp1 = ws.cell(row=row, column=employee_col).value
        emp2 = ws.cell(row=row+1, column=employee_col).value
        if emp1 and emp2 and emp1 == emp2:
            # Highlight both rows
            for r in [row, row+1]:
                for col in range(1, max_col+1):
                    cell = ws.cell(row=r, column=col)
                    cell.fill = yellow_fill
                    cell.font = Font(bold=True)

    # Add thick border around all data
    from openpyxl.styles import Alignment
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            # Only style cells with text
            if cell.value not in (None, ""):
                # All borders (thick for outside, thin for inside)
                border = Border(
                    left=thick if col == 1 else thin,
                    right=thick if col == max_col else thin,
                    top=thick if row == 1 else thin,
                    bottom=thick if row == max_row else thin
                )
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
    # Fit to one page when printing
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    wb.save(out_path)

@app.route('/', methods=['GET'])
def index():
    return render_template('index.html')

@app.route('/process', methods=['POST'])
def process():
    file1 = request.files.get('file1')
    file2 = request.files.get('file2')
    results = []
    summary = ""
    out1 = out2 = None

    preview1 = preview2 = None
    if file1 and file1.filename:
        filename1 = secure_filename(file1.filename)
        path1 = os.path.join(UPLOAD_FOLDER, filename1)
        file1.save(path1)
        reg1, ot1, total1, pct1, df1 = process_payroll(path1)
        summary1 = f"Total hours for the BHB Kingsville for the week is {total1:.2f} of which {ot1:.2f} or {pct1:.2%} is considered overtime."
        summary_row1 = pd.DataFrame([{col: "" for col in df1.columns}])
        summary_row1.iloc[0, 0] = summary1
        df1_with_summary = pd.concat([df1, summary_row1], ignore_index=True)
        out1 = f"BHB_Kingsville_sorted_{filename1.replace('.csv', '')}.xlsx"
        out1_path = os.path.join(RESULTS_FOLDER, out1)
        save_to_excel(df1_with_summary, out1_path)
        summary += summary1 + "\n"
        preview1 = df1_with_summary.to_html(classes="table table-bordered table-sm", index=False, border=0, justify="center")

    if file2 and file2.filename:
        filename2 = secure_filename(file2.filename)
        path2 = os.path.join(UPLOAD_FOLDER, filename2)
        file2.save(path2)
        reg2, ot2, total2, pct2, df2 = process_payroll(path2)
        summary2 = f"Total hours for the BHB Alice for the week is {total2:.2f} of which {ot2:.2f} or {pct2:.2%} is considered overtime."
        summary_row2 = pd.DataFrame([{col: "" for col in df2.columns}])
        summary_row2.iloc[0, 0] = summary2
        df2_with_summary = pd.concat([df2, summary_row2], ignore_index=True)
        out2 = f"BHB_Alice_sorted_{filename2.replace('.csv', '')}.xlsx"
        out2_path = os.path.join(RESULTS_FOLDER, out2)
        save_to_excel(df2_with_summary, out2_path)
        summary += summary2 + "\n"
        preview2 = df2_with_summary.to_html(classes="table table-bordered table-sm", index=False, border=0, justify="center")

    if not summary:
        flash('Please upload at least one CSV file.')
        return redirect(url_for('index'))

    return render_template('index.html', summary=summary.strip(), out1=out1, out2=out2, preview1=preview1, preview2=preview2)

@app.route('/download/<filename>')
def download(filename):
    file_path = os.path.join(RESULTS_FOLDER, filename)
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    flash('File not found.')
    return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=True)
