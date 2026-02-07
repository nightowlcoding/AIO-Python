import os
import csv
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import subprocess
import sys
from tkcalendar import DateEntry
from utils import validate_number, format_currency, safe_file_read, safe_file_write
from auto_save import AutoSaveManager, get_backup_manager
from app_config import create_button, create_header, COLORS, FONTS

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

OUT_DIR = os.path.expanduser("~/Documents/AIO Python/daily_logs")
os.makedirs(OUT_DIR, exist_ok=True)

EMPLOYEE_LIST_FILE = os.path.expanduser("~/Documents/AIO Python/employee_list.csv")

# Initialize backup manager
backup_manager = get_backup_manager(OUT_DIR)

class DailyLogApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Daily Log")
        self.geometry("360x800")
        self.configure(bg="#f0f0f0")
        self.resizable(False, False)
        
        # Load employee names list
        self.employee_names = self._load_employee_names()
        
        # Initialize auto-save manager
        self.auto_save = AutoSaveManager(
            self, 
            self._auto_save_data,
            interval_seconds=300  # 5 minutes
        )
        self.auto_save_label = None
        
        # Create canvas with scrollbar for scrollable content
        canvas = tk.Canvas(self, bg="#f0f0f0", highlightthickness=0)
        scrollbar = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        self.scrollable_frame = tk.Frame(canvas, bg="#f0f0f0")
        
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        self.canvas = canvas
        self.main_frame = self.scrollable_frame
        
        self._build_header()
        self._build_employee_section()
        self._build_sales_summary_section()
        self._build_summary_section()
        self._build_actions()
        self._build_status_bar()
        
        self._update_totals()
        
        # Check for imported data from DLimport
        self._load_imported_data()
        
        # Setup keyboard shortcuts
        self._setup_keyboard_shortcuts()
        
        # Enable mousewheel scrolling
        self.bind_all("<MouseWheel>", self._on_mousewheel)
        self.bind_all("<Button-4>", self._on_mousewheel)
        self.bind_all("<Button-5>", self._on_mousewheel)
        
        # Start auto-save
        self.auto_save.start()
    
    def _on_mousewheel(self, event):
        """Handle mousewheel scrolling"""
        if event.num == 5 or event.delta < 0:
            self.canvas.yview_scroll(1, "units")
        elif event.num == 4 or event.delta > 0:
            self.canvas.yview_scroll(-1, "units")
    
    def _update_credit_total(self, event=None):
        """Calculate and update credit card total from individual card inputs"""
        try:
            visa = validate_number(self.visa_var.get(), default=0.0)
            mastercard = validate_number(self.mastercard_var.get(), default=0.0)
            amex = validate_number(self.amex_var.get(), default=0.0)
            discover = validate_number(self.discover_var.get(), default=0.0)
            total = visa + mastercard + amex + discover
            self.credit_var.set(format_currency(total))
            self.auto_save.mark_dirty()
        except ValueError:
            pass
    
    def _update_sales_total(self, event=None):
        """Calculate and update sales total from individual sales inputs"""
        try:
            beer = validate_number(self.beer_var.get(), default=0.0)
            liquor = validate_number(self.liquor_var.get(), default=0.0)
            wine = validate_number(self.wine_var.get(), default=0.0)
            food = validate_number(self.food_var.get(), default=0.0)
            total = beer + liquor + wine + food
            self.sales_total_var.set(format_currency(total))
            self.auto_save.mark_dirty()
        except ValueError:
            pass
    
    def _on_date_change(self, event=None):
        """Called when date selection changes - auto-load data"""
        # Use after() to avoid multiple rapid calls
        if hasattr(self, '_date_change_job'):
            self.after_cancel(self._date_change_job)
        self._date_change_job = self.after(500, self._auto_load_for_date)
    
    def _auto_load_for_date(self):
        """Automatically load data for the selected date"""
        date_str = f"{self.year_var.get()}-{self.month_var.get():02d}-{self.day_var.get():02d}"
        
        # Try to load both day and night shift files
        day_file = os.path.join(OUT_DIR, f"{date_str}_Day.csv")
        night_file = os.path.join(OUT_DIR, f"{date_str}_Night.csv")
        
        files_to_load = []
        if os.path.exists(day_file):
            files_to_load.append(("Day", day_file))
        if os.path.exists(night_file):
            files_to_load.append(("Night", night_file))
        
        if not files_to_load:
            # No saved data for this date - just update totals (including cash deductions)
            self._update_totals()
            return
        
        try:
            # Clear current data
            self.employees = []
            all_notes = []
            
            # Load from all available shift files
            for shift_name, filename in files_to_load:
                with open(filename, "r", newline="") as f:
                    reader = csv.reader(f)
                    rows = list(reader)
                    
                    # Parse the CSV
                    section = None
                    for i, row in enumerate(rows):
                        if not row:
                            continue
                        
                        if row[0] == "Notes" and len(row) > 1 and row[1]:
                            all_notes.append(f"{shift_name}: {row[1]}")
                        elif row[0] == "Employee Entries":
                            section = "employees"
                            continue
                        elif section == "employees" and row[0] != "Name":  # Skip header row
                            if len(row) >= 10:
                                # Handle both old and new CSV formats
                                employee_data = {
                                    "name": row[0],
                                    "area": row[1],
                                    "cash": row[2],
                                    "cc_tips": row[3] if len(row) > 3 else "0.00",
                                    "shift": shift_name
                                }
                                
                                # New format with individual credit cards (13 columns)
                                if len(row) >= 13:
                                    employee_data["visa"] = row[4]
                                    employee_data["mastercard"] = row[5]
                                    employee_data["amex"] = row[6]
                                    employee_data["discover"] = row[7]
                                    employee_data["credit"] = row[8]
                                    employee_data["beer"] = row[9]
                                    employee_data["liquor"] = row[10]
                                    employee_data["wine"] = row[11]
                                    employee_data["food"] = row[12]
                                # Old format (11 columns)
                                else:
                                    employee_data["visa"] = "0.00"
                                    employee_data["mastercard"] = "0.00"
                                    employee_data["amex"] = "0.00"
                                    employee_data["discover"] = "0.00"
                                    employee_data["credit"] = row[4] if len(row) > 4 else "0.00"
                                    employee_data["beer"] = row[7] if len(row) > 7 else "0.00"
                                    employee_data["liquor"] = row[8] if len(row) > 8 else "0.00"
                                    employee_data["wine"] = row[9] if len(row) > 9 else "0.00"
                                    employee_data["food"] = row[10] if len(row) > 10 else "0.00"
                                
                                self.employees.append(employee_data)
            
            # Set combined notes
            if all_notes:
                self.notes_var.set(" | ".join(all_notes))
            else:
                self.notes_var.set("")
            
            # Sort and display loaded employees
            self._sort_employees()
            self._update_display()
            self._update_totals()
        
        except Exception as e:
            print(f"Error auto-loading data: {e}")
            self._update_totals()
    
    def _build_header(self):
        """Build compact mobile header"""
        header_frame = tk.Frame(self.main_frame, bg="#2c2c2c", height=100)
        header_frame.pack(fill="x", pady=(0, 5))
        header_frame.pack_propagate(False)
        
        # Title
        title_label = tk.Label(
            header_frame,
            text="Daily Log",
            font=("Arial", 18, "bold"),
            bg="#2c2c2c",
            fg="black"
        )
        title_label.pack(pady=(5, 5))
        
        # Date selectors - Spinbox style (Rolodex)
        date_frame = tk.Frame(header_frame, bg="#2c2c2c")
        date_frame.pack(pady=2)
        
        tk.Label(date_frame, text="Date:", bg="#2c2c2c", fg="black", font=("Arial", 9)).pack(side="left", padx=2)
        
        # Month spinbox
        self.month_var = tk.IntVar(value=datetime.now().month)
        month_spin = tk.Spinbox(
            date_frame,
            from_=1,
            to=12,
            textvariable=self.month_var,
            width=3,
            font=("Arial", 10, "bold"),
            bg="white",
            fg="black",
            buttonbackground="#3a3a3a",
            command=self._on_date_change,
            wrap=True
        )
        month_spin.pack(side="left", padx=1)
        
        tk.Label(date_frame, text="/", bg="#2c2c2c", fg="black", font=("Arial", 10)).pack(side="left")
        
        # Day spinbox
        self.day_var = tk.IntVar(value=datetime.now().day)
        day_spin = tk.Spinbox(
            date_frame,
            from_=1,
            to=31,
            textvariable=self.day_var,
            width=3,
            font=("Arial", 10, "bold"),
            bg="white",
            fg="black",
            buttonbackground="#3a3a3a",
            command=self._on_date_change,
            wrap=True
        )
        day_spin.pack(side="left", padx=1)
        
        tk.Label(date_frame, text="/", bg="#2c2c2c", fg="black", font=("Arial", 10)).pack(side="left")
        
        # Year spinbox
        current_year = datetime.now().year
        self.year_var = tk.IntVar(value=current_year)
        year_spin = tk.Spinbox(
            date_frame,
            from_=current_year-5,
            to=current_year+2,
            textvariable=self.year_var,
            width=5,
            font=("Arial", 10, "bold"),
            bg="white",
            fg="black",
            buttonbackground="#3a3a3a",
            command=self._on_date_change,
            wrap=True
        )
        year_spin.pack(side="left", padx=1)
        
        # Load button
        load_btn = tk.Button(date_frame, text="Load", command=self._load_log,
                           bg="#3a3a3a", fg="black", font=("Arial", 7, "bold"),
                           relief="flat", padx=5, pady=2, cursor="hand2",
                           activebackground="#1a1a1a", activeforeground="black")
        load_btn.pack(side="left", padx=5)
        
        # Shift selector only
        controls_frame = tk.Frame(header_frame, bg="#2c2c2c")
        controls_frame.pack(pady=2)
        
        tk.Label(controls_frame, text="Shift:", bg="#2c2c2c", fg="black", font=("Arial", 9)).pack(side="left", padx=2)
        self.shift_var = tk.StringVar(value="Day")
        shift_menu = tk.OptionMenu(controls_frame, self.shift_var, "Day", "Night")
        shift_menu.config(width=5, font=("Arial", 8), bg="white", fg="black", activebackground="#e0e0e0", activeforeground="black")
        shift_menu["menu"].config(bg="white", fg="black")
        shift_menu.pack(side="left", padx=2)
        
        # Initialize notes_var for compatibility with existing code
        self.notes_var = tk.StringVar()
    
    def _build_employee_section(self):
        """Build mobile-optimized employee section with tabbed interface"""
        emp_frame = ttk.LabelFrame(self.main_frame, text="Employees", padding="5")
        emp_frame.pack(fill="x", padx=5, pady=3)
        
        # Create notebook for tabs
        notebook = ttk.Notebook(emp_frame)
        notebook.pack(fill="x", pady=2)
        
        # Tab 1: Basic Info
        basic_tab = tk.Frame(notebook, bg="white")
        notebook.add(basic_tab, text="Basic")
        
        tk.Label(basic_tab, text="Name:", font=("Arial", 9, "bold"), bg="white", fg="black").grid(row=0, column=0, sticky="w", padx=2, pady=2)
        self.name_var = tk.StringVar()
        # Use Combobox with autocomplete for employee names
        self.name_entry = ttk.Combobox(basic_tab, textvariable=self.name_var, values=self.employee_names, 
                    width=23, font=("Arial", 9))
        self.name_entry.grid(row=0, column=1, padx=2, pady=2)
        # Refresh employee list when clicking on the name field
        self.name_entry.bind("<Button-1>", lambda e: self._refresh_employee_list())
        self.name_entry.bind("<FocusIn>", lambda e: self._refresh_employee_list())
        
        tk.Label(basic_tab, text="Area:", font=("Arial", 9, "bold"), bg="white", fg="black").grid(row=1, column=0, sticky="w", padx=2, pady=2)
        self.area_var = tk.StringVar(value="Dining")
        ttk.Combobox(basic_tab, textvariable=self.area_var, values=["Dining", "Bar", "To Out"], 
                    width=23, state="readonly", font=("Arial", 9)).grid(row=1, column=1, padx=2, pady=2)
        
        tk.Label(basic_tab, text="Cash:", font=("Arial", 9, "bold"), bg="white", fg="black").grid(row=2, column=0, sticky="w", padx=2, pady=2)
        self.cash_var = tk.StringVar(value="0.00")
        tk.Entry(basic_tab, textvariable=self.cash_var, width=25, font=("Arial", 9), fg="black", bg="white").grid(row=2, column=1, padx=2, pady=2)
        
        tk.Label(basic_tab, text="C.C. Tips:", font=("Arial", 9, "bold"), bg="white", fg="black").grid(row=3, column=0, sticky="w", padx=2, pady=2)
        self.cc_tips_var = tk.StringVar(value="0.00")
        tk.Entry(basic_tab, textvariable=self.cc_tips_var, width=25, font=("Arial", 9), fg="black", bg="white").grid(row=3, column=1, padx=2, pady=2)
        
        # Import Data button
        import_btn = tk.Button(basic_tab, text="📁 Import Data", command=self._import_data,
                             bg="#2196F3", fg="black", font=("Arial", 9, "bold"),
                             relief="flat", padx=10, pady=5, cursor="hand2",
                             activebackground="#1976D2", activeforeground="black")
        import_btn.grid(row=4, column=0, columnspan=2, padx=2, pady=5, sticky="ew")
        
        # Tab 2: Credit/CC
        credit_tab = tk.Frame(notebook, bg="white")
        notebook.add(credit_tab, text="Credit")
        
        tk.Label(credit_tab, text="Visa:", font=("Arial", 9, "bold"), bg="white", fg="black").grid(row=0, column=0, sticky="w", padx=2, pady=2)
        self.visa_var = tk.StringVar(value="0.00")
        visa_entry = tk.Entry(credit_tab, textvariable=self.visa_var, width=25, font=("Arial", 9), fg="black", bg="white")
        visa_entry.grid(row=0, column=1, padx=2, pady=2)
        visa_entry.bind("<KeyRelease>", self._update_credit_total)
        
        tk.Label(credit_tab, text="Mastercard:", font=("Arial", 9, "bold"), bg="white", fg="black").grid(row=1, column=0, sticky="w", padx=2, pady=2)
        self.mastercard_var = tk.StringVar(value="0.00")
        mastercard_entry = tk.Entry(credit_tab, textvariable=self.mastercard_var, width=25, font=("Arial", 9), fg="black", bg="white")
        mastercard_entry.grid(row=1, column=1, padx=2, pady=2)
        mastercard_entry.bind("<KeyRelease>", self._update_credit_total)
        
        tk.Label(credit_tab, text="Amex:", font=("Arial", 9, "bold"), bg="white", fg="black").grid(row=2, column=0, sticky="w", padx=2, pady=2)
        self.amex_var = tk.StringVar(value="0.00")
        amex_entry = tk.Entry(credit_tab, textvariable=self.amex_var, width=25, font=("Arial", 9), fg="black", bg="white")
        amex_entry.grid(row=2, column=1, padx=2, pady=2)
        amex_entry.bind("<KeyRelease>", self._update_credit_total)
        
        tk.Label(credit_tab, text="Discover:", font=("Arial", 9, "bold"), bg="white", fg="black").grid(row=3, column=0, sticky="w", padx=2, pady=2)
        self.discover_var = tk.StringVar(value="0.00")
        discover_entry = tk.Entry(credit_tab, textvariable=self.discover_var, width=25, font=("Arial", 9), fg="black", bg="white")
        discover_entry.grid(row=3, column=1, padx=2, pady=2)
        discover_entry.bind("<KeyRelease>", self._update_credit_total)
        
        # Credit Card Total (calculated field)
        tk.Label(credit_tab, text="Credit Card Total:", font=("Arial", 9, "bold"), bg="#2c2c2c", fg="black").grid(row=4, column=0, sticky="w", padx=2, pady=5)
        self.credit_var = tk.StringVar(value="0.00")
        tk.Label(credit_tab, textvariable=self.credit_var, font=("Arial", 9, "bold"), bg="#2c2c2c", fg="black", anchor="e", width=23).grid(row=4, column=1, padx=2, pady=5, sticky="e")
        
        # Tab 3: Sales
        sales_tab = tk.Frame(notebook, bg="white")
        notebook.add(sales_tab, text="Sales")
        
        tk.Label(sales_tab, text="Beer:", font=("Arial", 9, "bold"), bg="white", fg="black").grid(row=0, column=0, sticky="w", padx=2, pady=2)
        self.beer_var = tk.StringVar(value="0.00")
        beer_entry = tk.Entry(sales_tab, textvariable=self.beer_var, width=25, font=("Arial", 9), fg="black", bg="white")
        beer_entry.grid(row=0, column=1, padx=2, pady=2)
        beer_entry.bind("<KeyRelease>", self._update_sales_total)
        
        tk.Label(sales_tab, text="Liquor:", font=("Arial", 9, "bold"), bg="white", fg="black").grid(row=1, column=0, sticky="w", padx=2, pady=2)
        self.liquor_var = tk.StringVar(value="0.00")
        liquor_entry = tk.Entry(sales_tab, textvariable=self.liquor_var, width=25, font=("Arial", 9), fg="black", bg="white")
        liquor_entry.grid(row=1, column=1, padx=2, pady=2)
        liquor_entry.bind("<KeyRelease>", self._update_sales_total)
        
        tk.Label(sales_tab, text="Wine:", font=("Arial", 9, "bold"), bg="white", fg="black").grid(row=2, column=0, sticky="w", padx=2, pady=2)
        self.wine_var = tk.StringVar(value="0.00")
        wine_entry = tk.Entry(sales_tab, textvariable=self.wine_var, width=25, font=("Arial", 9), fg="black", bg="white")
        wine_entry.grid(row=2, column=1, padx=2, pady=2)
        wine_entry.bind("<KeyRelease>", self._update_sales_total)
        
        tk.Label(sales_tab, text="Food:", font=("Arial", 9, "bold"), bg="white", fg="black").grid(row=3, column=0, sticky="w", padx=2, pady=2)
        self.food_var = tk.StringVar(value="0.00")
        food_entry = tk.Entry(sales_tab, textvariable=self.food_var, width=25, font=("Arial", 9), fg="black", bg="white")
        food_entry.grid(row=3, column=1, padx=2, pady=2)
        food_entry.bind("<KeyRelease>", self._update_sales_total)
        
        # Sales Total (calculated field)
        tk.Label(sales_tab, text="Sales Total:", font=("Arial", 9, "bold"), bg="#2c2c2c", fg="black").grid(row=4, column=0, sticky="w", padx=2, pady=5)
        self.sales_total_var = tk.StringVar(value="0.00")
        tk.Label(sales_tab, textvariable=self.sales_total_var, font=("Arial", 9, "bold"), bg="#2c2c2c", fg="black", anchor="e", width=23).grid(row=4, column=1, padx=2, pady=5, sticky="e")
        
        # Buttons
        btn_frame = tk.Frame(emp_frame, bg="white")
        btn_frame.pack(fill="x", pady=5)
        
        add_btn = tk.Button(btn_frame, text="Add", command=self._add_employee,
                          bg="white", fg="black", font=("Arial", 9, "bold"),
                          relief="solid", bd=1, padx=10, pady=3, cursor="hand2",
                          activebackground="#e0e0e0", activeforeground="black", width=8)
        add_btn.pack(side="left", padx=2)
        
        remove_btn = tk.Button(btn_frame, text="Remove", command=self._remove_selected,
                             bg="white", fg="black", font=("Arial", 9),
                             relief="solid", bd=1, padx=10, pady=3, cursor="hand2",
                             activebackground="#e0e0e0", activeforeground="black", width=8)
        remove_btn.pack(side="left", padx=2)
        
        clear_btn = tk.Button(btn_frame, text="Clear", command=self._clear_all,
                            bg="white", fg="black", font=("Arial", 9),
                            relief="solid", bd=1, padx=10, pady=3, cursor="hand2",
                            activebackground="#e0e0e0", activeforeground="black", width=8)
        clear_btn.pack(side="left", padx=2)
        
        # Employee list (compact display)
        list_frame = tk.Frame(emp_frame, bg="white")
        list_frame.pack(fill="x", pady=5)
        
        tk.Label(list_frame, text="Entries:", font=("Arial", 9, "bold"), bg="white", fg="black").pack(anchor="w", padx=2)
        
        # Store employee data
        self.employees = []
        
        # Listbox for compact view
        self.employee_listbox = tk.Listbox(list_frame, height=6, font=("Courier", 8), bg="white", fg="black")
        self.employee_listbox.pack(fill="x", padx=2, pady=2)
        self.employee_listbox.bind("<Double-Button-1>", self._edit_employee)
        
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=self.employee_listbox.yview)
        self.employee_listbox.configure(yscrollcommand=scrollbar.set)
    
    def _build_sales_summary_section(self):
        """Build sales summary section"""
        sales_frame = ttk.LabelFrame(self.main_frame, text="Sales Summary", padding="5")
        sales_frame.pack(fill="x", padx=5, pady=3)
        
        # Sales breakdown
        sales_breakdown_frame = tk.Frame(sales_frame, bg="white")
        sales_breakdown_frame.pack(fill="x", pady=2)
        
        tk.Label(sales_breakdown_frame, text="Beer Sales:", font=("Arial", 9, "bold"), bg="white", fg="black", width=12, anchor="w").grid(row=0, column=0, padx=2, pady=1, sticky="w")
        self.total_beer_var = tk.StringVar(value="$0.00")
        tk.Label(sales_breakdown_frame, textvariable=self.total_beer_var, font=("Arial", 9), bg="white", fg="black", width=12, anchor="e").grid(row=0, column=1, padx=2, pady=1, sticky="e")
        
        tk.Label(sales_breakdown_frame, text="Liquor Sales:", font=("Arial", 9, "bold"), bg="white", fg="black", width=12, anchor="w").grid(row=1, column=0, padx=2, pady=1, sticky="w")
        self.total_liquor_var = tk.StringVar(value="$0.00")
        tk.Label(sales_breakdown_frame, textvariable=self.total_liquor_var, font=("Arial", 9), bg="white", fg="black", width=12, anchor="e").grid(row=1, column=1, padx=2, pady=1, sticky="e")
        
        tk.Label(sales_breakdown_frame, text="Wine Sales:", font=("Arial", 9, "bold"), bg="white", fg="black", width=12, anchor="w").grid(row=2, column=0, padx=2, pady=1, sticky="w")
        self.total_wine_var = tk.StringVar(value="$0.00")
        tk.Label(sales_breakdown_frame, textvariable=self.total_wine_var, font=("Arial", 9), bg="white", fg="black", width=12, anchor="e").grid(row=2, column=1, padx=2, pady=1, sticky="e")
        
        tk.Label(sales_breakdown_frame, text="Food Sales:", font=("Arial", 9, "bold"), bg="white", fg="black", width=12, anchor="w").grid(row=3, column=0, padx=2, pady=1, sticky="w")
        self.total_food_var = tk.StringVar(value="$0.00")
        tk.Label(sales_breakdown_frame, textvariable=self.total_food_var, font=("Arial", 9), bg="white", fg="black", width=12, anchor="e").grid(row=3, column=1, padx=2, pady=1, sticky="e")
        
        # Separator
        ttk.Separator(sales_frame, orient="horizontal").pack(fill="x", pady=5)
        
        # Total Net Sales (highlighted)
        net_sales_frame = tk.Frame(sales_frame, bg="#2c2c2c", relief="solid", bd=2)
        net_sales_frame.pack(fill="x", pady=5)
        
        tk.Label(net_sales_frame, text="TOTAL NET SALES:", font=("Arial", 11, "bold"), bg="#2c2c2c", fg="black", anchor="w").pack(side="left", padx=10, pady=5)
        self.net_sales_var = tk.StringVar(value="$0.00")
        tk.Label(net_sales_frame, textvariable=self.net_sales_var, font=("Arial", 11, "bold"), bg="#2c2c2c", fg="black", anchor="e").pack(side="right", padx=10, pady=5)
        
        # Total Cash
        total_cash_frame = tk.Frame(sales_frame, bg="white")
        total_cash_frame.pack(fill="x", pady=2)
        
        tk.Label(total_cash_frame, text="Total Cash:", font=("Arial", 9, "bold"), bg="white", fg="black", width=12, anchor="w").pack(side="left", padx=2)
        self.total_cash_var = tk.StringVar(value="$0.00")
        tk.Label(total_cash_frame, textvariable=self.total_cash_var, font=("Arial", 9), bg="white", fg="black", width=12, anchor="e").pack(side="right", padx=2)
        
        # Total C.C. Tips
        total_cc_tips_frame = tk.Frame(sales_frame, bg="white")
        total_cc_tips_frame.pack(fill="x", pady=2)
        
        tk.Label(total_cc_tips_frame, text="Total C.C. Tips:", font=("Arial", 9, "bold"), bg="white", fg="black", width=12, anchor="w").pack(side="left", padx=2)
        self.total_cc_tips_var = tk.StringVar(value="$0.00")
        tk.Label(total_cc_tips_frame, textvariable=self.total_cc_tips_var, font=("Arial", 9), bg="white", fg="black", width=12, anchor="e").pack(side="right", padx=2)
        
        # Total Cash Deductions
        total_cash_deductions_frame = tk.Frame(sales_frame, bg="white")
        total_cash_deductions_frame.pack(fill="x", pady=2)
        
        tk.Label(total_cash_deductions_frame, text="Total Cash Deductions:", font=("Arial", 9, "bold"), bg="white", fg="black", width=18, anchor="w").pack(side="left", padx=2)
        self.total_cash_deductions_var = tk.StringVar(value="$0.00")
        tk.Label(total_cash_deductions_frame, textvariable=self.total_cash_deductions_var, font=("Arial", 9), bg="white", fg="black", width=12, anchor="e").pack(side="right", padx=2)
        
        # Expected Deposit (highlighted)
        expected_deposit_frame = tk.Frame(sales_frame, bg="#2c2c2c", relief="solid", bd=2)
        expected_deposit_frame.pack(fill="x", pady=5)
        
        tk.Label(expected_deposit_frame, text="EXPECTED DEPOSIT:", font=("Arial", 11, "bold"), bg="#2c2c2c", fg="black", anchor="w").pack(side="left", padx=10, pady=5)
        self.expected_deposit_var = tk.StringVar(value="$0.00")
        tk.Label(expected_deposit_frame, textvariable=self.expected_deposit_var, font=("Arial", 11, "bold"), bg="#2c2c2c", fg="black", anchor="e").pack(side="right", padx=10, pady=5)
    
    def _build_summary_section(self):
        """Build compact summary section"""
        summary_frame = ttk.LabelFrame(self.main_frame, text="Totals", padding="5")
        summary_frame.pack(fill="x", padx=5, pady=3)
        
        self.totals_text = tk.Text(summary_frame, height=8, width=40, font=("Courier", 8), 
                                   bg="white", fg="black", relief="flat", state="disabled", wrap="none")
        self.totals_text.pack(fill="x")
        
        # Add horizontal scrollbar for totals
        h_scroll = ttk.Scrollbar(summary_frame, orient="horizontal", command=self.totals_text.xview)
        self.totals_text.configure(xscrollcommand=h_scroll.set)
        h_scroll.pack(fill="x")
    
    def _build_actions(self):
        """Build compact action buttons"""
        action_frame = tk.Frame(self.main_frame, bg="#f0f0f0")
        action_frame.pack(fill="x", padx=5, pady=10)
        
        # Three-button row for Cash Deductions, Cash Drawer, Tax Exempt Tickets
        special_buttons_frame = tk.Frame(action_frame, bg="#f0f0f0")
        special_buttons_frame.pack(fill="x", pady=(0, 5))
        
        cash_ded_btn = tk.Button(special_buttons_frame, text="Cash Deductions", command=self._cash_deductions,
                                bg="#4a4a4a", fg="black", font=("Arial", 8),
                                relief="flat", padx=5, pady=6, cursor="hand2",
                                activebackground="#2a2a2a", activeforeground="black")
        cash_ded_btn.pack(side="left", fill="x", expand=True, padx=1)
        
        cash_drawer_btn = tk.Button(special_buttons_frame, text="Cash Drawer", command=self._cash_drawer,
                                   bg="#4a4a4a", fg="black", font=("Arial", 8),
                                   relief="flat", padx=5, pady=6, cursor="hand2",
                                   activebackground="#2a2a2a", activeforeground="black")
        cash_drawer_btn.pack(side="left", fill="x", expand=True, padx=1)
        
        tax_exempt_btn = tk.Button(special_buttons_frame, text="Tax Exempt Tickets", command=self._tax_exempt,
                                  bg="#4a4a4a", fg="black", font=("Arial", 8),
                                  relief="flat", padx=5, pady=6, cursor="hand2",
                                  activebackground="#2a2a2a", activeforeground="black")
        tax_exempt_btn.pack(side="left", fill="x", expand=True, padx=1)
        
        save_btn = tk.Button(action_frame, text="Save Log", command=self._save_log,
                           bg="#3a3a3a", fg="black", font=("Arial", 10, "bold"),
                           relief="flat", padx=15, pady=8, cursor="hand2",
                           activebackground="#1a1a1a", activeforeground="black")
        save_btn.pack(fill="x", pady=2)
        
        export_btn = tk.Button(action_frame, text="Export CSV", command=self._export_csv,
                             bg="#4a4a4a", fg="black", font=("Arial", 10),
                             relief="flat", padx=15, pady=8, cursor="hand2",
                             activebackground="#2a2a2a", activeforeground="black")
        export_btn.pack(fill="x", pady=2)
        
        clear_day_btn = tk.Button(action_frame, text="Clear Day", command=self._clear_day,
                                bg="#4a4a4a", fg="black", font=("Arial", 10),
                                relief="flat", padx=15, pady=8, cursor="hand2",
                                activebackground="#2a2a2a", activeforeground="black")
        clear_day_btn.pack(fill="x", pady=2)
    
    def _cash_deductions(self):
        """Handle cash deductions using CashManager"""
        date_str = f"{self.year_var.get()}-{self.month_var.get():02d}-{self.day_var.get():02d}"
        
        try:
            cashmanager_path = os.path.join(os.path.dirname(__file__), "CashManager.py")
            subprocess.Popen([sys.executable, cashmanager_path, date_str])
        except Exception as e:
            messagebox.showerror("Error", f"Could not open Cash Manager:\n{str(e)}")
    
    def _cash_drawer(self):
        """Handle cash drawer using CashManager"""
        date_str = f"{self.year_var.get()}-{self.month_var.get():02d}-{self.day_var.get():02d}"
        
        try:
            cashmanager_path = os.path.join(os.path.dirname(__file__), "CashManager.py")
            subprocess.Popen([sys.executable, cashmanager_path, date_str])
        except Exception as e:
            messagebox.showerror("Error", f"Could not open Cash Manager:\n{str(e)}")
    
    def _tax_exempt(self):
        """Handle tax exempt tickets"""
        messagebox.showinfo("Tax Exempt Tickets", "Tax Exempt Tickets feature coming soon!")
    
    def _add_employee(self):
        """Add employee entry"""
        name = self.name_var.get().strip()
        if not name:
            messagebox.showwarning("Input Error", "Please enter employee name")
            return
        
        employee_data = {
            "name": name,
            "area": self.area_var.get(),
            "cash": self.cash_var.get(),
            "cc_tips": self.cc_tips_var.get(),
            "visa": self.visa_var.get(),
            "mastercard": self.mastercard_var.get(),
            "amex": self.amex_var.get(),
            "discover": self.discover_var.get(),
            "credit": self.credit_var.get(),
            "beer": self.beer_var.get(),
            "liquor": self.liquor_var.get(),
            "wine": self.wine_var.get(),
            "food": self.food_var.get(),
            "shift": self.shift_var.get()  # Track which shift
        }
        
        self.employees.append(employee_data)
        
        # Sort employees by name, then by actual cash (cash - cc tips)
        self._sort_employees()
        
        # Clear fields
        self.name_var.set("")
        self.cash_var.set("0.00")
        self.cc_tips_var.set("0.00")
        self.visa_var.set("0.00")
        self.mastercard_var.set("0.00")
        self.amex_var.set("0.00")
        self.discover_var.set("0.00")
        self.credit_var.set("0.00")
        self.beer_var.set("0.00")
        self.liquor_var.set("0.00")
        self.wine_var.set("0.00")
        self.food_var.set("0.00")
        
        self._update_display()
        self._update_totals()
    
    def _import_data(self):
        """Import employee data from Excel file"""
        if not EXCEL_AVAILABLE:
            messagebox.showerror("Error", "openpyxl library is not installed.\n\nPlease install it using:\npip install openpyxl")
            return
        
        filename = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        
        if not filename:
            return
        
        try:
            workbook = openpyxl.load_workbook(filename, data_only=True)
            sheet = workbook.active
            
            # Find header row and column mappings
            header_row = None
            column_map = {}
            
            # Common column name variations
            name_variants = ['name', 'employee', 'employee name', 'server', 'bartender']
            area_variants = ['area', 'position', 'section', 'location']
            cash_variants = ['cash', 'cash tips', 'cash sales']
            cc_tips_variants = ['cc tips', 'credit tips', 'cc received', 'credit card tips']
            visa_variants = ['visa']
            mastercard_variants = ['mastercard', 'mc', 'master card']
            amex_variants = ['amex', 'american express', 'amx']
            discover_variants = ['discover', 'disc']
            credit_variants = ['credit', 'credit total', 'credit card total', 'cc total']
            beer_variants = ['beer', 'beer sales']
            liquor_variants = ['liquor', 'liquor sales', 'spirits']
            wine_variants = ['wine', 'wine sales']
            food_variants = ['food', 'food sales']
            
            # Search for header row (first 10 rows)
            for row_idx in range(1, min(11, sheet.max_row + 1)):
                row = sheet[row_idx]
                cells = [str(cell.value).lower().strip() if cell.value else '' for cell in row]
                
                # Check if this looks like a header row
                if any(variant in cells for variant in name_variants):
                    header_row = row_idx
                    
                    # Map columns
                    for idx, cell_value in enumerate(cells):
                        if any(v == cell_value for v in name_variants):
                            column_map['name'] = idx
                        elif any(v == cell_value for v in area_variants):
                            column_map['area'] = idx
                        elif any(v == cell_value for v in cash_variants):
                            column_map['cash'] = idx
                        elif any(v == cell_value for v in cc_tips_variants):
                            column_map['cc_tips'] = idx
                        elif any(v == cell_value for v in visa_variants):
                            column_map['visa'] = idx
                        elif any(v == cell_value for v in mastercard_variants):
                            column_map['mastercard'] = idx
                        elif any(v == cell_value for v in amex_variants):
                            column_map['amex'] = idx
                        elif any(v == cell_value for v in discover_variants):
                            column_map['discover'] = idx
                        elif any(v == cell_value for v in credit_variants):
                            column_map['credit'] = idx
                        elif any(v == cell_value for v in beer_variants):
                            column_map['beer'] = idx
                        elif any(v == cell_value for v in liquor_variants):
                            column_map['liquor'] = idx
                        elif any(v == cell_value for v in wine_variants):
                            column_map['wine'] = idx
                        elif any(v == cell_value for v in food_variants):
                            column_map['food'] = idx
                    break
            
            if not header_row or 'name' not in column_map:
                messagebox.showerror("Error", "Could not find employee name column in Excel file.\n\nPlease ensure the file has a header row with at least a 'Name' or 'Employee' column.")
                return
            
            # Import data
            imported_count = 0
            skipped_count = 0
            
            for row_idx in range(header_row + 1, sheet.max_row + 1):
                row = sheet[row_idx]
                
                # Get name
                name_val = row[column_map['name']].value
                if not name_val:
                    continue
                
                name = str(name_val).strip()
                if not name:
                    continue
                
                # Helper function to get cell value
                def get_value(key, default="0.00"):
                    if key not in column_map:
                        return default
                    cell_val = row[column_map[key]].value
                    if cell_val is None or str(cell_val).strip() == '':
                        return default
                    # Handle numeric values
                    try:
                        return f"{float(cell_val):.2f}"
                    except:
                        return str(cell_val).strip()
                
                # Create employee data
                employee_data = {
                    "name": name,
                    "area": get_value('area', 'Dining'),
                    "cash": get_value('cash'),
                    "cc_tips": get_value('cc_tips'),
                    "visa": get_value('visa'),
                    "mastercard": get_value('mastercard'),
                    "amex": get_value('amex'),
                    "discover": get_value('discover'),
                    "credit": get_value('credit'),
                    "beer": get_value('beer'),
                    "liquor": get_value('liquor'),
                    "wine": get_value('wine'),
                    "food": get_value('food'),
                    "shift": self.shift_var.get()
                }
                
                # Calculate credit total if individual cards are provided
                try:
                    visa = float(employee_data['visa'])
                    mc = float(employee_data['mastercard'])
                    amex = float(employee_data['amex'])
                    disc = float(employee_data['discover'])
                    card_total = visa + mc + amex + disc
                    if card_total > 0:
                        employee_data['credit'] = f"{card_total:.2f}"
                except:
                    pass
                
                self.employees.append(employee_data)
                imported_count += 1
            
            workbook.close()
            
            if imported_count > 0:
                self._sort_employees()
                self._update_display()
                self._update_totals()
                messagebox.showinfo("Success", f"Successfully imported {imported_count} employee record(s)!")
            else:
                messagebox.showwarning("No Data", "No employee data found in the Excel file.")
        
        except Exception as e:
            messagebox.showerror("Error", f"Failed to import Excel file:\n{str(e)}")
    
    def _sort_employees(self):
        """Sort employees by shift (Day first, Night second), then by position (Dining, Bar, To Out), then by actual cash"""
        # Define position order
        position_order = {"Dining": 1, "Bar": 2, "To Out": 3}
        # Define shift order
        shift_order = {"Day": 1, "Night": 2}
        
        self.employees.sort(key=lambda emp: (
            shift_order.get(emp.get("shift", "Day"), 1),  # Sort by shift: Day first, Night second
            position_order.get(emp["area"], 4),  # Sort by position: Dining, Bar, To Out
            -(float(emp["cash"]) - float(emp.get("cc_tips", "0.00")))  # Then by actual cash (descending)
        ))
    
    def _update_display(self):
        """Update the listbox display"""
        self.employee_listbox.delete(0, "end")
        for emp in self.employees:
            try:
                actual_cash = float(emp["cash"]) - float(emp.get("cc_tips", "0.00"))
                shift_indicator = emp.get("shift", "Day")[0]  # "D" or "N"
                display_text = f"{shift_indicator}|{emp['name'][:13]:13} {emp['area'][:6]:6} ${actual_cash:6.2f}"
                self.employee_listbox.insert("end", display_text)
            except ValueError:
                shift_indicator = emp.get("shift", "Day")[0]
                display_text = f"{shift_indicator}|{emp['name'][:13]:13} {emp['area'][:6]:6} ${float(emp['cash']):6.2f}"
                self.employee_listbox.insert("end", display_text)
    
    def _edit_employee(self, event):
        """Edit selected employee by double-clicking"""
        selection = self.employee_listbox.curselection()
        if not selection:
            return
        
        index = selection[0]
        emp = self.employees[index]
        
        # Load employee data into form fields
        self.name_var.set(emp["name"])
        self.area_var.set(emp["area"])
        self.cash_var.set(emp["cash"])
        self.cc_tips_var.set(emp.get("cc_tips", "0.00"))
        self.visa_var.set(emp.get("visa", "0.00"))
        self.mastercard_var.set(emp.get("mastercard", "0.00"))
        self.amex_var.set(emp.get("amex", "0.00"))
        self.discover_var.set(emp.get("discover", "0.00"))
        self.credit_var.set(emp.get("credit", "0.00"))
        self.beer_var.set(emp["beer"])
        self.liquor_var.set(emp["liquor"])
        self.wine_var.set(emp["wine"])
        self.food_var.set(emp["food"])
        
        # Remove the old entry
        del self.employees[index]
        self._update_display()
        self._update_totals()
    
    def _remove_selected(self):
        """Remove selected entry"""
        selection = self.employee_listbox.curselection()
        if not selection:
            messagebox.showwarning("Selection Error", "Please select an entry to remove")
            return
        
        index = selection[0]
        del self.employees[index]
        self._update_display()
        self._update_totals()
    
    def _clear_all(self):
        """Clear all entries"""
        if messagebox.askyesno("Confirm", "Clear all employee entries?"):
            self.employees = []
            self._update_display()
            self._update_totals()
    
    def _update_totals(self):
        """Update running totals display"""
        areas = {}
        total_cash = 0.0
        total_credit = 0.0
        total_cc_tips = 0.0
        total_beer = 0.0
        total_liquor = 0.0
        total_wine = 0.0
        total_food = 0.0
        
        for emp in self.employees:
            area = emp["area"]
            
            if area not in areas:
                areas[area] = {
                    "cash": 0.0, "credit": 0.0, "cc_tips": 0.0,
                    "beer": 0.0, "liquor": 0.0, "wine": 0.0, "food": 0.0
                }
            
            try:
                cash_val = float(emp["cash"])
                credit_val = float(emp.get("credit", "0.00"))
                cc_tips_val = float(emp.get("cc_tips", "0.00"))
                beer_val = float(emp["beer"])
                liquor_val = float(emp["liquor"])
                wine_val = float(emp["wine"])
                food_val = float(emp["food"])
                
                areas[area]["cash"] += cash_val
                areas[area]["credit"] += credit_val
                areas[area]["cc_tips"] += cc_tips_val
                areas[area]["beer"] += beer_val
                areas[area]["liquor"] += liquor_val
                areas[area]["wine"] += wine_val
                areas[area]["food"] += food_val
                
                # Track grand totals across all employees
                total_cash += cash_val
                total_credit += credit_val
                total_cc_tips += cc_tips_val
                total_beer += beer_val
                total_liquor += liquor_val
                total_wine += wine_val
                total_food += food_val
            except ValueError:
                pass
        
        # Calculate and update deposit (Cash - CC Tips)
        calculated_deposit = total_cash - total_cc_tips
        
        # Update sales summary section
        self.total_beer_var.set(f"${total_beer:.2f}")
        self.total_liquor_var.set(f"${total_liquor:.2f}")
        self.total_wine_var.set(f"${total_wine:.2f}")
        self.total_food_var.set(f"${total_food:.2f}")
        
        # Calculate and update total net sales
        total_net_sales = total_beer + total_liquor + total_wine + total_food
        self.net_sales_var.set(f"${total_net_sales:.2f}")
        
        # Update total cash
        self.total_cash_var.set(f"${total_cash:.2f}")
        
        # Update total C.C. tips
        self.total_cc_tips_var.set(f"${total_cc_tips:.2f}")
        
        # Update total cash deductions (load from file)
        cash_deductions_total = self._load_cash_deductions_total()
        self.total_cash_deductions_var.set(f"${cash_deductions_total:.2f}")
        
        # Calculate and update expected deposit (Total Cash - Total CC Tips - Cash Deductions)
        expected_deposit = total_cash - total_cc_tips - cash_deductions_total
        self.expected_deposit_var.set(f"${expected_deposit:.2f}")
        
        # Display totals
        self.totals_text.config(state="normal")
        self.totals_text.delete("1.0", "end")
        
        if not areas:
            self.totals_text.insert("end", "No entries yet")
        else:
            for area, totals in areas.items():
                self.totals_text.insert("end", f"{area}:\n")
                self.totals_text.insert("end", f"  Cash:     ${totals['cash']:8.2f}\n")
                self.totals_text.insert("end", f"  CC Tips:  ${totals['cc_tips']:8.2f}\n")
                self.totals_text.insert("end", f"  Credit:   ${totals['credit']:8.2f}\n")
                self.totals_text.insert("end", f"  Beer:     ${totals['beer']:8.2f}\n")
                self.totals_text.insert("end", f"  Liquor:   ${totals['liquor']:8.2f}\n")
                self.totals_text.insert("end", f"  Wine:     ${totals['wine']:8.2f}\n")
                self.totals_text.insert("end", f"  Food:     ${totals['food']:8.2f}\n")
                self.totals_text.insert("end", "\n")
            
            # Show deposit calculation
            self.totals_text.insert("end", "=" * 35 + "\n")
            self.totals_text.insert("end", f"DEPOSIT CALCULATION:\n")
            self.totals_text.insert("end", f"Total Cash:    ${total_cash:8.2f}\n")
            self.totals_text.insert("end", f"Total CC Tips: ${total_cc_tips:8.2f}\n")
            self.totals_text.insert("end", f"Deposit:       ${calculated_deposit:8.2f}\n")
            
            # Calculate and show Total Net Sales (highlighted section)
            total_net_sales = total_beer + total_liquor + total_wine + total_food
            self.totals_text.insert("end", "\n" + "=" * 35 + "\n")
            self.totals_text.insert("end", f"*** TOTAL NET SALES ***\n")
            self.totals_text.insert("end", f"${total_net_sales:8.2f}\n")
            self.totals_text.insert("end", "=" * 35 + "\n")
            
            # Show grand totals across all employees
            self.totals_text.insert("end", "\n" + "-" * 35 + "\n")
            self.totals_text.insert("end", f"GRAND TOTALS (ALL EMPLOYEES):\n")
            self.totals_text.insert("end", f"Total Cash:    ${total_cash:8.2f}\n")
            self.totals_text.insert("end", f"Total CC Tips: ${total_cc_tips:8.2f}\n")
            self.totals_text.insert("end", f"Total Credit:  ${total_credit:8.2f}\n")
            self.totals_text.insert("end", f"Total Beer:    ${total_beer:8.2f}\n")
            self.totals_text.insert("end", f"Total Liquor:  ${total_liquor:8.2f}\n")
            self.totals_text.insert("end", f"Total Wine:    ${total_wine:8.2f}\n")
            self.totals_text.insert("end", f"Total Food:    ${total_food:8.2f}\n")
        
        self.totals_text.config(state="disabled")
    
    def _save_log(self):
        """Save daily log to files (separate by shift)"""
        date_str = f"{self.year_var.get()}-{self.month_var.get():02d}-{self.day_var.get():02d}"
        
        # Group employees by shift
        day_employees = [emp for emp in self.employees if emp.get("shift", "Day") == "Day"]
        night_employees = [emp for emp in self.employees if emp.get("shift", "Night") == "Night"]
        
        files_saved = []
        
        # Save Day shift if there are day employees
        if day_employees:
            filename = os.path.join(OUT_DIR, f"{date_str}_Day.csv")
            # Create backup before saving
            if os.path.exists(filename):
                backup_manager.create_backup(f"{date_str}_Day.csv")
            self._save_shift_file(filename, date_str, "Day", day_employees)
            files_saved.append("Day")
        
        # Save Night shift if there are night employees
        if night_employees:
            filename = os.path.join(OUT_DIR, f"{date_str}_Night.csv")
            # Create backup before saving
            if os.path.exists(filename):
                backup_manager.create_backup(f"{date_str}_Night.csv")
            self._save_shift_file(filename, date_str, "Night", night_employees)
            files_saved.append("Night")
        
        if files_saved:
            self.auto_save.mark_clean()
            messagebox.showinfo("Success", f"Logs saved for {date_str}\nShifts: {', '.join(files_saved)}")
        else:
            messagebox.showwarning("No Data", "No employees to save!")
    
    def _save_shift_file(self, filename, date_str, shift, employees):
        """Save a single shift file"""
        with open(filename, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["Date", date_str])
            writer.writerow(["Shift", shift])
            writer.writerow(["Notes", self.notes_var.get()])
            writer.writerow([])
            
            writer.writerow(["Employee Entries"])
            writer.writerow(["Name", "Area", "Cash", "C.C. Tips", "Visa", "Mastercard", "Amex", "Discover",
                           "Credit Total", "Beer", "Liquor", "Wine", "Food"])
            
            for emp in employees:
                writer.writerow([
                    emp["name"], emp["area"], emp["cash"], emp.get("cc_tips", "0.00"),
                    emp.get("visa", "0.00"), emp.get("mastercard", "0.00"), emp.get("amex", "0.00"), 
                    emp.get("discover", "0.00"), emp.get("credit", "0.00"),
                    emp["beer"], emp["liquor"], emp["wine"], emp["food"]
                ])
    
    def _load_log(self):
        """Load saved log from selected date (both shifts)"""
        date_str = f"{self.year_var.get()}-{self.month_var.get():02d}-{self.day_var.get():02d}"
        
        # Try to load both day and night shift files
        day_file = os.path.join(OUT_DIR, f"{date_str}_Day.csv")
        night_file = os.path.join(OUT_DIR, f"{date_str}_Night.csv")
        
        files_to_load = []
        if os.path.exists(day_file):
            files_to_load.append(("Day", day_file))
        if os.path.exists(night_file):
            files_to_load.append(("Night", night_file))
        
        if not files_to_load:
            messagebox.showwarning("Not Found", f"No saved logs found for:\n{date_str}")
            return
        
        try:
            # Clear current data
            self.employees = []
            all_notes = []
            
            # Load from all available shift files
            for shift_name, filename in files_to_load:
                with open(filename, "r", newline="") as f:
                    reader = csv.reader(f)
                    rows = list(reader)
                    
                    # Parse the CSV
                    section = None
                    for i, row in enumerate(rows):
                        if not row:
                            continue
                        
                        if row[0] == "Notes" and len(row) > 1 and row[1]:
                            all_notes.append(f"{shift_name}: {row[1]}")
                        elif row[0] == "Employee Entries":
                            section = "employees"
                            continue
                        elif section == "employees" and row[0] != "Name":  # Skip header row
                            if len(row) >= 10:
                                # Handle both old and new CSV formats
                                employee_data = {
                                    "name": row[0],
                                    "area": row[1],
                                    "cash": row[2],
                                    "cc_tips": row[3] if len(row) > 3 else "0.00",
                                    "shift": shift_name
                                }
                                
                                # New format with individual credit cards (13 columns)
                                if len(row) >= 13:
                                    employee_data["visa"] = row[4]
                                    employee_data["mastercard"] = row[5]
                                    employee_data["amex"] = row[6]
                                    employee_data["discover"] = row[7]
                                    employee_data["credit"] = row[8]
                                    employee_data["beer"] = row[9]
                                    employee_data["liquor"] = row[10]
                                    employee_data["wine"] = row[11]
                                    employee_data["food"] = row[12]
                                # Old format (11 columns)
                                else:
                                    employee_data["visa"] = "0.00"
                                    employee_data["mastercard"] = "0.00"
                                    employee_data["amex"] = "0.00"
                                    employee_data["discover"] = "0.00"
                                    employee_data["credit"] = row[4] if len(row) > 4 else "0.00"
                                    employee_data["beer"] = row[7] if len(row) > 7 else "0.00"
                                    employee_data["liquor"] = row[8] if len(row) > 8 else "0.00"
                                    employee_data["wine"] = row[9] if len(row) > 9 else "0.00"
                                    employee_data["food"] = row[10] if len(row) > 10 else "0.00"
                                
                                self.employees.append(employee_data)
            
            # Set combined notes
            if all_notes:
                self.notes_var.set(" | ".join(all_notes))
            
            # Sort and display loaded employees
            self._sort_employees()
            self._update_display()
            self._update_totals()
            
            shifts_loaded = ", ".join([s for s, f in files_to_load])
            messagebox.showinfo("Success", f"Loaded logs from:\n{date_str}\nShifts: {shifts_loaded}\n{len(self.employees)} employees")
        
        except Exception as e:
            messagebox.showerror("Error", f"Error loading log:\n{str(e)}")
    
    def _export_csv(self):
        """Export current data to CSV"""
        # Generate default filename from current date (MM-DD-YR format)
        month = self.month_var.get()
        day = self.day_var.get()
        year = str(self.year_var.get())[-2:]  # Get last 2 digits of year
        default_filename = f"{month:02d}-{day:02d}-{year}.csv"
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".csv",
            initialfile=default_filename,
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        
        if filename:
            with open(filename, "w", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["Name", "Area", "Cash", "C.C. Tips", "Visa", "Mastercard", "Amex", "Discover",
                               "Credit Total", "Beer", "Liquor", "Wine", "Food"])
                
                for emp in self.employees:
                    writer.writerow([
                        emp["name"], emp["area"], emp["cash"], emp.get("cc_tips", "0.00"),
                        emp.get("visa", "0.00"), emp.get("mastercard", "0.00"), emp.get("amex", "0.00"),
                        emp.get("discover", "0.00"), emp.get("credit", "0.00"),
                        emp["beer"], emp["liquor"], emp["wine"], emp["food"]
                    ])
            
            messagebox.showinfo("Success", f"Data exported to:\n{filename}")
    
    def _clear_day(self):
        """Clear all day data"""
        if messagebox.askyesno("Confirm", "Clear all data for this day?"):
            self.employees = []
            self.notes_var.set("")
            self._update_display()
            self._update_totals()
    
    def _load_employee_names(self):
        """Load employee names from CSV file"""
        if os.path.exists(EMPLOYEE_LIST_FILE):
            try:
                with open(EMPLOYEE_LIST_FILE, "r", encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    # Filter only active employees and get their names
                    names = [row['name'] for row in reader if row.get('name') and row.get('status', 'Active') == 'Active']
                    return sorted(names)
            except Exception as e:
                print(f"Error loading employee list: {e}")
                return []
        return []
    
    def _refresh_employee_list(self):
        """Refresh the employee names list in the dropdown"""
        self.employee_names = self._load_employee_names()
        if hasattr(self, 'name_entry'):
            self.name_entry['values'] = self.employee_names
    
    def _load_imported_data(self):
        """Load imported data from DLimport application"""
        import json
        temp_file = os.path.expanduser("~/Documents/AIO Python/dailylog_import_data.json")
        
        if not os.path.exists(temp_file):
            return
        
        try:
            with open(temp_file, 'r') as f:
                data = json.load(f)
            
            # Fill in the data
            if 'cash' in data:
                self.cash_var.set(f"{data['cash']:.2f}")
            if 'cc_tips' in data:
                self.cc_tips_var.set(f"{data['cc_tips']:.2f}")
            if 'visa' in data:
                self.visa_var.set(f"{data['visa']:.2f}")
            if 'mastercard' in data:
                self.mastercard_var.set(f"{data['mastercard']:.2f}")
            if 'amex' in data:
                self.amex_var.set(f"{data['amex']:.2f}")
            if 'discover' in data:
                self.discover_var.set(f"{data['discover']:.2f}")
            if 'liquor' in data:
                self.liquor_var.set(f"{data['liquor']:.2f}")
            if 'beer' in data:
                self.beer_var.set(f"{data['beer']:.2f}")
            if 'wine' in data:
                self.wine_var.set(f"{data['wine']:.2f}")
            if 'food' in data:
                self.food_var.set(f"{data['food']:.2f}")
            
            # Update credit total and sales total
            self._update_credit_total()
            self._update_sales_total()
            
            # Delete the temp file after loading
            os.remove(temp_file)
            
            # Show confirmation
            messagebox.showinfo("Data Loaded", "Imported data has been loaded into the form!")
        
        except Exception as e:
            print(f"Error loading imported data: {e}")
            # Don't show error to user, just fail silently
    
    def _load_cash_deductions_total(self):
        """Load cash deductions total from saved file"""
        date_str = f"{self.year_var.get()}-{self.month_var.get():02d}-{self.day_var.get():02d}"
        filename = os.path.join(OUT_DIR, f"{date_str}_CashDeductions.csv")
        
        if not os.path.exists(filename):
            return 0.0
        
        try:
            total = 0.0
            with open(filename, "r", newline="") as f:
                reader = csv.reader(f)
                rows = list(reader)
                
                section = None
                for row in rows:
                    if not row:
                        continue
                    
                    if row[0] == "Cash Deductions":
                        section = "deductions"
                        continue
                    elif section == "deductions" and row[0] != "Items Purchased" and row[0] != "Total":
                        if len(row) >= 3:
                            try:
                                total += float(row[2])
                            except ValueError:
                                pass
            
            return total
        except Exception as e:
            print(f"Error loading cash deductions: {e}")
            return 0.0
    
    def _import_employee_list(self):
        """Import employee names from a CSV file"""
        filename = filedialog.askopenfilename(
            title="Select Employee List CSV",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        
        if filename:
            try:
                with open(filename, "r", newline="") as f:
                    reader = csv.reader(f)
                    names = [row[0].strip() for row in reader if row and row[0].strip()]
                
                # Save to employee list file
                with open(EMPLOYEE_LIST_FILE, "w", newline="") as f:
                    writer = csv.writer(f)
                    for name in names:
                        writer.writerow([name])
                
                # Update the list and combobox
                self.employee_names = sorted(names)
                self.name_entry['values'] = self.employee_names
                
                messagebox.showinfo("Success", f"Imported {len(names)} employee names!\n\nList saved to:\n{EMPLOYEE_LIST_FILE}")
            
            except Exception as e:
                messagebox.showerror("Error", f"Error importing employee list:\n{str(e)}")


    def _setup_keyboard_shortcuts(self):
        """Setup keyboard shortcuts"""
        self.bind('<Control-s>', lambda e: self._save_csv())
        self.bind('<Control-n>', lambda e: self._add_employee())
        self.bind('<Control-e>', lambda e: self._export_to_excel())
        self.bind('<Escape>', lambda e: self.destroy())
    
    def _build_status_bar(self):
        """Build status bar showing auto-save status"""
        status_frame = tk.Frame(self.main_frame, bg="#f0f0f0")
        status_frame.pack(fill="x", padx=10, pady=5)
        
        self.auto_save_label = tk.Label(
            status_frame,
            text="Auto-save: Ready",
            font=("Arial", 9),
            fg="#666666",
            bg="#f0f0f0"
        )
        self.auto_save_label.pack(side="left")
        
        # Shortcuts info
        shortcuts_label = tk.Label(
            status_frame,
            text="Ctrl+S: Save | Ctrl+N: Add | Ctrl+E: Export",
            font=("Arial", 9),
            fg="#666666",
            bg="#f0f0f0"
        )
        shortcuts_label.pack(side="right")
    
    def _auto_save_data(self):
        """Auto-save callback for AutoSaveManager"""
        try:
            self._save_csv(show_message=False)
            if self.auto_save_label:
                now = datetime.now().strftime("%H:%M:%S")
                self.auto_save_label.config(text=f"Auto-saved: {now}")
        except Exception as e:
            print(f"Auto-save failed: {e}")


if __name__ == "__main__":
    app = DailyLogApp()
    app.mainloop()
