import os
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

class DLImportApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Daily Log Importer")
        self.geometry("360x800")
        self.configure(bg="#f0f0f0")
        self.resizable(False, False)
        
        # Data storage
        self.imported_data = {}
        self.excel_file = None
        
        # Create canvas with scrollbar for scrollable content
        canvas = tk.Canvas(self, bg="#f0f0f0", highlightthickness=0)
        scrollbar = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        self.scrollable_frame = tk.Frame(canvas, bg="#f0f0f0")
        
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        self.canvas = canvas
        self.main_frame = self.scrollable_frame
        
        self._build_header()
        self._build_import_section()
        self._build_data_display()
        self._build_actions()
        
        # Enable mousewheel scrolling
        self.bind_all("<MouseWheel>", self._on_mousewheel)
        self.bind_all("<Button-4>", self._on_mousewheel)
        self.bind_all("<Button-5>", self._on_mousewheel)
    
    def _on_mousewheel(self, event):
        """Handle mousewheel scrolling"""
        if event.num == 5 or event.delta < 0:
            self.canvas.yview_scroll(1, "units")
        elif event.num == 4 or event.delta > 0:
            self.canvas.yview_scroll(-1, "units")
    
    def _build_header(self):
        """Build compact mobile header"""
        header_frame = tk.Frame(self.main_frame, bg="#2c2c2c", height=100)
        header_frame.pack(fill="x", pady=(0, 5))
        header_frame.pack_propagate(False)
        
        # Title
        title_label = tk.Label(
            header_frame,
            text="Daily Log Importer",
            font=("Arial", 18, "bold"),
            bg="#2c2c2c",
            fg="black"
        )
        title_label.pack(pady=(5, 5))
        
        # Subtitle
        subtitle_label = tk.Label(
            header_frame,
            text="Import Excel data automatically",
            font=("Arial", 10),
            bg="#2c2c2c",
            fg="black"
        )
        subtitle_label.pack(pady=(0, 5))
    
    def _build_import_section(self):
        """Build import file section"""
        import_frame = ttk.LabelFrame(self.main_frame, text="Import File", padding="10")
        import_frame.pack(fill="x", padx=5, pady=5)
        
        # File info
        self.file_label = tk.Label(import_frame, text="No file selected", 
                                   font=("Arial", 9), bg="white", fg="black",
                                   wraplength=300, justify="left")
        self.file_label.pack(fill="x", pady=5)
        
        # Import button
        import_btn = tk.Button(import_frame, text="📁 Select Excel File", 
                              command=self._import_excel,
                              bg="#2196F3", fg="black", font=("Arial", 11, "bold"),
                              relief="flat", padx=15, pady=10, cursor="hand2",
                              activebackground="#1976D2", activeforeground="black")
        import_btn.pack(fill="x", pady=5)
    
    def _build_data_display(self):
        """Build data display section"""
        display_frame = ttk.LabelFrame(self.main_frame, text="Extracted Data", padding="10")
        display_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Cash
        cash_container = tk.Frame(display_frame, bg="white", relief="solid", bd=1)
        cash_container.pack(fill="x", pady=3)
        
        tk.Label(cash_container, text="Cash:", font=("Arial", 10, "bold"),
                bg="white", fg="black").pack(side="left", padx=10, pady=8)
        
        self.cash_var = tk.StringVar(value="Not imported")
        tk.Label(cash_container, textvariable=self.cash_var, font=("Arial", 10),
                bg="white", fg="black").pack(side="right", padx=10, pady=8)
        
        # C.C. Tips
        cc_tips_container = tk.Frame(display_frame, bg="white", relief="solid", bd=1)
        cc_tips_container.pack(fill="x", pady=3)
        
        tk.Label(cc_tips_container, text="C.C. Tips:", font=("Arial", 10, "bold"),
                bg="white", fg="black").pack(side="left", padx=10, pady=8)
        
        self.cc_tips_var = tk.StringVar(value="Not imported")
        tk.Label(cc_tips_container, textvariable=self.cc_tips_var, font=("Arial", 10),
                bg="white", fg="black").pack(side="right", padx=10, pady=8)
        
        # Credit Cards Section
        cc_frame = tk.Frame(display_frame, bg="#f5f5f5", relief="solid", bd=1)
        cc_frame.pack(fill="x", pady=5)
        
        tk.Label(cc_frame, text="Credit Card Breakdown", font=("Arial", 10, "bold"),
                bg="#f5f5f5", fg="black").pack(pady=(8, 5))
        
        # Visa
        visa_row = tk.Frame(cc_frame, bg="#f5f5f5")
        visa_row.pack(fill="x", padx=10, pady=2)
        tk.Label(visa_row, text="Visa:", font=("Arial", 9, "bold"),
                bg="#f5f5f5", fg="black", width=15, anchor="w").pack(side="left")
        self.visa_var = tk.StringVar(value="$0.00")
        tk.Label(visa_row, textvariable=self.visa_var, font=("Arial", 9),
                bg="#f5f5f5", fg="black", anchor="e").pack(side="right")
        
        # Mastercard
        mc_row = tk.Frame(cc_frame, bg="#f5f5f5")
        mc_row.pack(fill="x", padx=10, pady=2)
        tk.Label(mc_row, text="Mastercard:", font=("Arial", 9, "bold"),
                bg="#f5f5f5", fg="black", width=15, anchor="w").pack(side="left")
        self.mastercard_var = tk.StringVar(value="$0.00")
        tk.Label(mc_row, textvariable=self.mastercard_var, font=("Arial", 9),
                bg="#f5f5f5", fg="black", anchor="e").pack(side="right")
        
        # Amex
        amex_row = tk.Frame(cc_frame, bg="#f5f5f5")
        amex_row.pack(fill="x", padx=10, pady=2)
        tk.Label(amex_row, text="Amex:", font=("Arial", 9, "bold"),
                bg="#f5f5f5", fg="black", width=15, anchor="w").pack(side="left")
        self.amex_var = tk.StringVar(value="$0.00")
        tk.Label(amex_row, textvariable=self.amex_var, font=("Arial", 9),
                bg="#f5f5f5", fg="black", anchor="e").pack(side="right")
        
        # Discover
        discover_row = tk.Frame(cc_frame, bg="#f5f5f5")
        discover_row.pack(fill="x", padx=10, pady=2)
        tk.Label(discover_row, text="Discover:", font=("Arial", 9, "bold"),
                bg="#f5f5f5", fg="black", width=15, anchor="w").pack(side="left")
        self.discover_var = tk.StringVar(value="$0.00")
        tk.Label(discover_row, textvariable=self.discover_var, font=("Arial", 9),
                bg="#f5f5f5", fg="black", anchor="e").pack(side="right", padx=0, pady=(0, 8))
        
        # Credit Total
        total_row = tk.Frame(cc_frame, bg="#2c2c2c")
        total_row.pack(fill="x", pady=0)
        tk.Label(total_row, text="Credit Total:", font=("Arial", 10, "bold"),
                bg="#2c2c2c", fg="black").pack(side="left", padx=10, pady=8)
        self.credit_total_var = tk.StringVar(value="$0.00")
        tk.Label(total_row, textvariable=self.credit_total_var, font=("Arial", 10, "bold"),
                bg="#2c2c2c", fg="black").pack(side="right", padx=10, pady=8)
        
        # Sales Category Section
        sales_frame = tk.Frame(display_frame, bg="#f5f5f5", relief="solid", bd=1)
        sales_frame.pack(fill="x", pady=5)
        
        tk.Label(sales_frame, text="Sales Category Summary", font=("Arial", 10, "bold"),
                bg="#f5f5f5", fg="black").pack(pady=(8, 5))
        
        # Liquor
        liquor_row = tk.Frame(sales_frame, bg="#f5f5f5")
        liquor_row.pack(fill="x", padx=10, pady=2)
        tk.Label(liquor_row, text="Liquor:", font=("Arial", 9, "bold"),
                bg="#f5f5f5", fg="black", width=15, anchor="w").pack(side="left")
        self.liquor_var = tk.StringVar(value="$0.00")
        tk.Label(liquor_row, textvariable=self.liquor_var, font=("Arial", 9),
                bg="#f5f5f5", fg="black", anchor="e").pack(side="right")
        
        # Beer
        beer_row = tk.Frame(sales_frame, bg="#f5f5f5")
        beer_row.pack(fill="x", padx=10, pady=2)
        tk.Label(beer_row, text="Beer:", font=("Arial", 9, "bold"),
                bg="#f5f5f5", fg="black", width=15, anchor="w").pack(side="left")
        self.beer_var = tk.StringVar(value="$0.00")
        tk.Label(beer_row, textvariable=self.beer_var, font=("Arial", 9),
                bg="#f5f5f5", fg="black", anchor="e").pack(side="right")
        
        # Wine
        wine_row = tk.Frame(sales_frame, bg="#f5f5f5")
        wine_row.pack(fill="x", padx=10, pady=2)
        tk.Label(wine_row, text="Wine:", font=("Arial", 9, "bold"),
                bg="#f5f5f5", fg="black", width=15, anchor="w").pack(side="left")
        self.wine_var = tk.StringVar(value="$0.00")
        tk.Label(wine_row, textvariable=self.wine_var, font=("Arial", 9),
                bg="#f5f5f5", fg="black", anchor="e").pack(side="right")
        
        # Food
        food_row = tk.Frame(sales_frame, bg="#f5f5f5")
        food_row.pack(fill="x", padx=10, pady=2)
        tk.Label(food_row, text="Food:", font=("Arial", 9, "bold"),
                bg="#f5f5f5", fg="black", width=15, anchor="w").pack(side="left")
        self.food_var = tk.StringVar(value="$0.00")
        tk.Label(food_row, textvariable=self.food_var, font=("Arial", 9),
                bg="#f5f5f5", fg="black", anchor="e").pack(side="right", padx=0, pady=(0, 8))
    
    def _build_actions(self):
        """Build action buttons"""
        action_frame = tk.Frame(self.main_frame, bg="#f0f0f0")
        action_frame.pack(fill="x", padx=5, pady=10)
        
        # Send to Daily Log button
        send_btn = tk.Button(action_frame, text="📤 Send to Daily Log", 
                            command=self._send_to_dailylog,
                            bg="#4CAF50", fg="black", font=("Arial", 10, "bold"),
                            relief="flat", padx=15, pady=8, cursor="hand2",
                            activebackground="#388E3C", activeforeground="black")
        send_btn.pack(fill="x", pady=2)
        
        # Copy to Clipboard button
        copy_btn = tk.Button(action_frame, text="📋 Copy to Clipboard", 
                            command=self._copy_to_clipboard,
                            bg="#2196F3", fg="black", font=("Arial", 10, "bold"),
                            relief="flat", padx=15, pady=8, cursor="hand2",
                            activebackground="#1976D2", activeforeground="black")
        copy_btn.pack(fill="x", pady=2)
        
        # Export CSV button
        export_btn = tk.Button(action_frame, text="💾 Export to CSV", 
                              command=self._export_csv,
                              bg="#3a3a3a", fg="black", font=("Arial", 10, "bold"),
                              relief="flat", padx=15, pady=8, cursor="hand2",
                              activebackground="#1a1a1a", activeforeground="black")
        export_btn.pack(fill="x", pady=2)
        
        # Clear button
        clear_btn = tk.Button(action_frame, text="🗑️ Clear Data", 
                             command=self._clear_data,
                             bg="#4a4a4a", fg="black", font=("Arial", 10),
                             relief="flat", padx=15, pady=8, cursor="hand2",
                             activebackground="#2a2a2a", activeforeground="black")
        clear_btn.pack(fill="x", pady=2)
    
    def _import_excel(self):
        """Import and parse Excel file"""
        if not EXCEL_AVAILABLE:
            messagebox.showerror("Error", "openpyxl library is not installed.\n\nPlease install it using:\npip install openpyxl")
            return
        
        filename = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        
        if not filename:
            return
        
        try:
            self.excel_file = filename
            self.file_label.config(text=f"File: {os.path.basename(filename)}", fg="black")
            
            workbook = openpyxl.load_workbook(filename, data_only=True)
            
            # Initialize values
            cash = 0.0
            cc_tips = 0.0
            visa = 0.0
            mastercard = 0.0
            amex = 0.0
            discover = 0.0
            liquor = 0.0
            beer = 0.0
            wine = 0.0
            food = 0.0
            
            # Search all sheets for data
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Look for "All Data" sheet or similar
                if "all data" in sheet_name.lower() or "data" in sheet_name.lower():
                    cash_found, cc_tips_found = self._parse_all_data_sheet(sheet)
                    if cash_found is not None:
                        cash = cash_found
                    if cc_tips_found is not None:
                        cc_tips = cc_tips_found
                
                # Look for "Payment Summary" sheet
                if "payment" in sheet_name.lower() and "summary" in sheet_name.lower():
                    visa_val, mc_val, amex_val, disc_val = self._parse_payment_summary_sheet(sheet)
                    if visa_val is not None:
                        visa = visa_val
                    if mc_val is not None:
                        mastercard = mc_val
                    if amex_val is not None:
                        amex = amex_val
                    if disc_val is not None:
                        discover = disc_val
                
                # Look for "Sales Category Summary" sheet
                if "sales" in sheet_name.lower() and "category" in sheet_name.lower():
                    liquor_val, beer_val, wine_val, food_val = self._parse_sales_category_sheet(sheet)
                    if liquor_val is not None:
                        liquor = liquor_val
                    if beer_val is not None:
                        beer = beer_val
                    if wine_val is not None:
                        wine = wine_val
                    if food_val is not None:
                        food = food_val
            
            workbook.close()
            
            # Update display
            self.cash_var.set(f"${cash:,.2f}")
            self.cc_tips_var.set(f"${cc_tips:,.2f}")
            self.visa_var.set(f"${visa:,.2f}")
            self.mastercard_var.set(f"${mastercard:,.2f}")
            self.amex_var.set(f"${amex:,.2f}")
            self.discover_var.set(f"${discover:,.2f}")
            self.liquor_var.set(f"${liquor:,.2f}")
            self.beer_var.set(f"${beer:,.2f}")
            self.wine_var.set(f"${wine:,.2f}")
            self.food_var.set(f"${food:,.2f}")
            
            credit_total = visa + mastercard + amex + discover
            self.credit_total_var.set(f"${credit_total:,.2f}")
            
            # Store data
            self.imported_data = {
                'cash': cash,
                'cc_tips': cc_tips,
                'visa': visa,
                'mastercard': mastercard,
                'amex': amex,
                'discover': discover,
                'credit_total': credit_total,
                'liquor': liquor,
                'beer': beer,
                'wine': wine,
                'food': food
            }
            
            messagebox.showinfo("Success", "Data imported successfully!")
        
        except Exception as e:
            messagebox.showerror("Error", f"Failed to import Excel file:\n{str(e)}")
    
    def _parse_all_data_sheet(self, sheet):
        """Parse All Data sheet for cash and CC tips"""
        cash = None
        cc_tips = None
        
        try:
            # Search for "Total Cash Payments" and "Credit/non-cash tips"
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value:
                        cell_str = str(cell.value).lower().strip()
                        
                        # Look for cash payments
                        if "total cash payment" in cell_str:
                            # Look for value in adjacent cells (right)
                            for offset in range(1, 10):
                                if cell.column + offset <= sheet.max_column:
                                    value_cell = sheet.cell(row=cell.row, column=cell.column + offset)
                                    if value_cell.value is not None:
                                        try:
                                            cash = float(value_cell.value)
                                            break
                                        except:
                                            pass
                        
                        # Look for CC tips
                        if "credit" in cell_str and "non-cash" in cell_str and "tip" in cell_str:
                            # Look for value in adjacent cells (right)
                            for offset in range(1, 10):
                                if cell.column + offset <= sheet.max_column:
                                    value_cell = sheet.cell(row=cell.row, column=cell.column + offset)
                                    if value_cell.value is not None:
                                        try:
                                            cc_tips = float(value_cell.value)
                                            break
                                        except:
                                            pass
        except Exception as e:
            print(f"Error parsing All Data sheet: {e}")
        
        return cash, cc_tips
    
    def _parse_payment_summary_sheet(self, sheet):
        """Parse Payment Summary sheet for credit card breakdown"""
        visa = None
        mastercard = None
        amex = None
        discover = None
        
        try:
            # Find "Payment Sub Type" column
            subtype_col = None
            amount_col = None
            header_row = None
            
            # Search for header row
            for row_idx, row in enumerate(sheet.iter_rows(max_row=20), start=1):
                for cell in row:
                    if cell.value:
                        cell_str = str(cell.value).lower().strip()
                        if "payment sub type" in cell_str or "payment subtype" in cell_str:
                            subtype_col = cell.column
                            header_row = row_idx
                            
                            # Look for amount column (typically to the right)
                            for offset in range(1, 10):
                                header_cell = sheet.cell(row=row_idx, column=cell.column + offset)
                                if header_cell.value:
                                    header_str = str(header_cell.value).lower().strip()
                                    if "amount" in header_str or "total" in header_str:
                                        amount_col = header_cell.column
                                        break
                            break
                if subtype_col:
                    break
            
            if not subtype_col or not amount_col:
                return None, None, None, None
            
            # Parse data rows
            for row_idx in range(header_row + 1, sheet.max_row + 1):
                subtype_cell = sheet.cell(row=row_idx, column=subtype_col)
                amount_cell = sheet.cell(row=row_idx, column=amount_col)
                
                if subtype_cell.value and amount_cell.value:
                    subtype_str = str(subtype_cell.value).lower().strip()
                    
                    try:
                        amount = float(amount_cell.value)
                        
                        if "visa" in subtype_str:
                            if visa is None:
                                visa = 0
                            visa += amount
                        elif "mastercard" in subtype_str or "master card" in subtype_str:
                            if mastercard is None:
                                mastercard = 0
                            mastercard += amount
                        elif "amex" in subtype_str or "american express" in subtype_str:
                            if amex is None:
                                amex = 0
                            amex += amount
                        elif "discover" in subtype_str:
                            if discover is None:
                                discover = 0
                            discover += amount
                    except:
                        pass
        
        except Exception as e:
            print(f"Error parsing Payment Summary sheet: {e}")
        
        return visa, mastercard, amex, discover
    
    def _parse_sales_category_sheet(self, sheet):
        """Parse Sales Category Summary sheet for liquor, beer, wine, food"""
        liquor = None
        beer = None
        wine = None
        food = None
        
        try:
            # Find "Category" or "Sales Category" column
            category_col = None
            amount_col = None
            header_row = None
            
            # Search for header row
            for row_idx, row in enumerate(sheet.iter_rows(max_row=20), start=1):
                for cell in row:
                    if cell.value:
                        cell_str = str(cell.value).lower().strip()
                        if "category" in cell_str or "sales category" in cell_str:
                            category_col = cell.column
                            header_row = row_idx
                            
                            # Look for amount/total column (typically to the right)
                            for offset in range(1, 10):
                                header_cell = sheet.cell(row=row_idx, column=cell.column + offset)
                                if header_cell.value:
                                    header_str = str(header_cell.value).lower().strip()
                                    if "amount" in header_str or "total" in header_str or "sales" in header_str:
                                        amount_col = header_cell.column
                                        break
                            break
                if category_col:
                    break
            
            if not category_col or not amount_col:
                return None, None, None, None
            
            # Parse data rows
            for row_idx in range(header_row + 1, sheet.max_row + 1):
                category_cell = sheet.cell(row=row_idx, column=category_col)
                amount_cell = sheet.cell(row=row_idx, column=amount_col)
                
                if category_cell.value and amount_cell.value:
                    category_str = str(category_cell.value).lower().strip()
                    
                    try:
                        amount = float(amount_cell.value)
                        
                        if "liquor" in category_str:
                            if liquor is None:
                                liquor = 0
                            liquor += amount
                        elif "beer" in category_str:
                            if beer is None:
                                beer = 0
                            beer += amount
                        elif "wine" in category_str:
                            if wine is None:
                                wine = 0
                            wine += amount
                        elif "food" in category_str:
                            if food is None:
                                food = 0
                            food += amount
                    except:
                        pass
        
        except Exception as e:
            print(f"Error parsing Sales Category Summary sheet: {e}")
        
        return liquor, beer, wine, food
    
    def _send_to_dailylog(self):
        """Send data to Daily Log application"""
        if not self.imported_data:
            messagebox.showwarning("Warning", "No data to send. Please import an Excel file first.")
            return
        
        try:
            # Create a temporary file to pass data
            import json
            temp_file = os.path.expanduser("~/Documents/AIO Python/dailylog_import_data.json")
            
            with open(temp_file, 'w') as f:
                json.dump(self.imported_data, f)
            
            # Launch dailylog.py
            dailylog_path = os.path.join(os.path.dirname(__file__), "dailylog.py")
            
            import subprocess
            import sys
            subprocess.Popen([sys.executable, dailylog_path])
            
            messagebox.showinfo("Success", "Data sent to Daily Log!\n\nThe Daily Log application will open and load the imported data.")
        
        except Exception as e:
            messagebox.showerror("Error", f"Failed to send data to Daily Log:\n{str(e)}")
    
    def _copy_to_clipboard(self):
        """Copy data to clipboard"""
        if not self.imported_data:
            messagebox.showwarning("Warning", "No data to copy. Please import an Excel file first.")
            return
        
        # Format data for clipboard
        data = self.imported_data
        text = f"""Cash: {data['cash']:.2f}
CC Tips: {data['cc_tips']:.2f}
Visa: {data['visa']:.2f}
Mastercard: {data['mastercard']:.2f}
Amex: {data['amex']:.2f}
Discover: {data['discover']:.2f}
Credit Total: {data['credit_total']:.2f}
Liquor: {data['liquor']:.2f}
Beer: {data['beer']:.2f}
Wine: {data['wine']:.2f}
Food: {data['food']:.2f}"""
        
        self.clipboard_clear()
        self.clipboard_append(text)
        messagebox.showinfo("Success", "Data copied to clipboard!")
    
    def _export_csv(self):
        """Export data to CSV"""
        if not self.imported_data:
            messagebox.showwarning("Warning", "No data to export. Please import an Excel file first.")
            return
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            initialfile=f"imported_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        )
        
        if filename:
            try:
                import csv
                with open(filename, 'w', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerow(['Field', 'Amount'])
                    for key, value in self.imported_data.items():
                        writer.writerow([key.replace('_', ' ').title(), f"{value:.2f}"])
                
                messagebox.showinfo("Success", f"Data exported to:\n{filename}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to export:\n{str(e)}")
    
    def _clear_data(self):
        """Clear all data"""
        self.cash_var.set("Not imported")
        self.cc_tips_var.set("Not imported")
        self.visa_var.set("$0.00")
        self.mastercard_var.set("$0.00")
        self.amex_var.set("$0.00")
        self.discover_var.set("$0.00")
        self.credit_total_var.set("$0.00")
        self.liquor_var.set("$0.00")
        self.beer_var.set("$0.00")
        self.wine_var.set("$0.00")
        self.food_var.set("$0.00")
        self.file_label.config(text="No file selected", fg="black")
        self.imported_data = {}
        self.excel_file = None


if __name__ == "__main__":
    app = DLImportApp()
    app.mainloop()
