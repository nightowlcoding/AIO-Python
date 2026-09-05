from flask import Flask, render_template, request, send_file, redirect, url_for, flash
import os
import re
import pandas as pd
from datetime import date, datetime
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from werkzeug.utils import secure_filename



app = Flask(__name__)
app.secret_key = 'payroll_secret_key'
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
RESULTS_FOLDER = os.path.join(BASE_DIR, 'results')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(RESULTS_FOLDER, exist_ok=True)

def get_last_csv_date(df, original_filename=""):
    # Prefer explicit date-like columns from the CSV.
    candidate_columns = [
        col for col in df.columns
        if any(keyword in str(col).strip().lower() for keyword in ["date", "day", "week ending"])
    ]

    for col in candidate_columns:
        parsed_dates = pd.to_datetime(df[col], errors='coerce')
        parsed_dates = parsed_dates.dropna()
        if not parsed_dates.empty:
            return parsed_dates.max().date()

    # Fallback: scan the filename for date-like tokens and use the latest value.
    # Use digit-based lookarounds instead of \b so underscore-glued suffixes
    # (e.g. a "_1" appended to avoid overwriting a duplicate upload) don't
    # swallow the boundary and cause the date to go undetected.
    patterns = [
        r"(?<!\d)\d{1,2}[/-]\d{1,2}[/-]\d{2,4}(?!\d)",
        r"(?<!\d)\d{4}[_/-]\d{1,2}[_/-]\d{1,2}(?!\d)",
        r"(?<!\d)\d{1,2}[_/-]\d{1,2}[_/-]\d{2,4}(?!\d)",
    ]
    found_dates = []
    for pattern in patterns:
        for token in re.findall(pattern, original_filename):
            normalized_token = token.replace('_', '-')
            parsed = pd.to_datetime(normalized_token, errors='coerce')
            if not pd.isna(parsed):
                found_dates.append(parsed)

    if found_dates:
        return max(found_dates).date()

    # Final fallback so filename generation always succeeds.
    return datetime.today().date()


def build_report_filename(report_date, location):
    date_str = report_date.strftime("%m-%d-%y")
    safe_location = re.sub(r'[<>:"/\\|?*]', '-', str(location)).strip()
    safe_location = re.sub(r'\s+', ' ', safe_location).rstrip('. ')
    if not safe_location:
        safe_location = "Unknown Location"
    return f"LaborReport {date_str} - {safe_location}.xlsx"


def get_restaurant_location(df, fallback_location):
    location_columns = [
        col for col in df.columns
        if any(keyword in str(col).strip().lower() for keyword in ["location", "restaurant", "store"])
    ]

    for col in location_columns:
        non_empty = df[col].dropna().astype(str).str.strip()
        non_empty = non_empty[non_empty != ""]
        if not non_empty.empty:
            return non_empty.iloc[-1]

    return fallback_location

def process_payroll(csv_path, include_kingsville_only_row=False, fallback_location=""):
    df = pd.read_csv(csv_path)
    report_date = get_last_csv_date(df, os.path.basename(csv_path))
    report_location = get_restaurant_location(df, fallback_location)
    columns_to_extract = [
        'Employee', 'Job Title', 'Regular Hours', 'Overtime Hours',
        'Declared Tips', 'Non-Cash Tips', 'Total Tips'
    ]
    extracted_data = df[columns_to_extract]

    def reformat_employee_name(name):
        parts = [part.strip() for part in str(name).split(',')]
        if len(parts) == 2:
            return f"{parts[1]} {parts[0]}"
        return name

    extracted_data.loc[:, 'Employee'] = extracted_data['Employee'].apply(reformat_employee_name)

    # Set all columns to zero for MANAGER job titles, except for 'April Salinas' as MANAGER
    manager_mask = (
        (extracted_data['Job Title'].str.strip().str.upper() == 'MANAGER') &
        ~((extracted_data['Employee'].str.strip().str.upper() == 'APRIL SALINAS') & (extracted_data['Job Title'].str.strip().str.upper() == 'MANAGER'))
    )
    for col in ['Regular Hours', 'Overtime Hours', 'Declared Tips', 'Non-Cash Tips', 'Total Tips']:
        extracted_data.loc[manager_mask, col] = 0

    # Add custom employees
    new_row = {
        'Employee': 'Arnold Ramirez',
        'Job Title': 'General Manager',
        'Regular Hours': 0,
        'Overtime Hours': 0,
        'Declared Tips': 0,
        'Non-Cash Tips': 0,
        'Total Tips': 0
    }
    new_row2 = {
        'Employee': 'Lesli Rodriguez',
        'Job Title': 'Media Admin',
        'Regular Hours': 20,
        'Overtime Hours': 0,
        'Declared Tips': 0,
        'Non-Cash Tips': 0,
        'Total Tips': 0
    }
    new_row3 = {
        'Employee': 'Maria Julia Martinez Villalta',
        'Job Title': '',
        'Regular Hours': 0,
        'Overtime Hours': 0,
        'Declared Tips': 0,
        'Non-Cash Tips': 0,
        'Total Tips': 0
    }
    new_row4 = {
        'Employee': 'Ysidro T Villarreal',
        'Job Title': '',
        'Regular Hours': 40,
        'Overtime Hours': 0,
        'Declared Tips': 0,
        'Non-Cash Tips': 0,
        'Total Tips': 0
    }
    custom_rows = [new_row, new_row2]
    if include_kingsville_only_row:
        custom_rows.extend([new_row3, new_row4])
    extracted_data = pd.concat([extracted_data, pd.DataFrame(custom_rows)], ignore_index=True)

    # Calculate totals
    totals = extracted_data.iloc[:, 2:].sum()
    total_combined_hours = float(totals['Regular Hours']) + float(totals['Overtime Hours'])
    totals_row = pd.DataFrame([{
        'Employee': 'Total',
        'Job Title': total_combined_hours,
        'Regular Hours': totals['Regular Hours'],
        'Overtime Hours': totals['Overtime Hours'],
        'Declared Tips': totals['Declared Tips'],
        'Non-Cash Tips': totals['Non-Cash Tips'],
        'Total Tips': totals['Total Tips']
    }])

    sorted_data = extracted_data.sort_values(by='Employee')
    sorted_data = pd.concat([
        sorted_data, totals_row
    ], ignore_index=True)

    reg_hours = totals['Regular Hours']
    ot_hours = totals['Overtime Hours']
    total_hours = total_combined_hours
    ot_pct = (ot_hours / total_hours) if total_hours > 0 else 0
    # Replace zeros with empty string for display
    sorted_data = sorted_data.replace(0, "")
    return reg_hours, ot_hours, total_hours, ot_pct, sorted_data, report_date, report_location

def save_to_excel(df, out_path):
    from openpyxl.styles import PatternFill, Font
    df.to_excel(out_path, index=False)
    # Apply borders and fit to one page
    wb = load_workbook(out_path)
    ws = wb.active
    max_row = ws.max_row
    max_col = ws.max_column
    thin = Side(border_style="thin", color="000000")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    bold_font = Font(bold=True)

    # Find duplicate employee names (consecutive rows)
    employee_col = 1  # 1-based index for openpyxl
    for row in range(2, max_row):  # skip header, 1-based
        emp1 = ws.cell(row=row, column=employee_col).value
        emp2 = ws.cell(row=row+1, column=employee_col).value
        if emp1 and emp2 and emp1 == emp2:
            # Highlight both rows
            for r in [row, row+1]:
                for col in range(1, max_col+1):
                    cell = ws.cell(row=r, column=col)
                    cell.fill = yellow_fill
                    cell.font = Font(bold=True)

    # Add thick border around all data
    from openpyxl.styles import Alignment
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            value = cell.value
            is_text = isinstance(value, str) and value.strip() != ""
            is_date = isinstance(value, (date, datetime))
            if is_text or is_date:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")
    # Fit to one page when printing
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    wb.save(out_path)

@app.route('/', methods=['GET'])
def index():
    return render_template('index.html')

@app.route('/process', methods=['POST'])
def process():
    file1 = request.files.get('file1')
    file2 = request.files.get('file2')
    summary = ""
    out1 = out2 = None
    preview1 = preview2 = None

    if file1 and file1.filename:
        filename1 = secure_filename(file1.filename)
        path1 = os.path.join(UPLOAD_FOLDER, filename1)
        file1.save(path1)
        reg1, ot1, total1, pct1, df1, report_date1, report_location1 = process_payroll(
            path1,
            include_kingsville_only_row=True,
            fallback_location="Kingsville"
        )
        summary1 = f"Total hours for the BHB Kingsville for the week is {total1:.2f} of which {ot1:.2f} or {pct1:.2%} is considered overtime."
        summary_row1 = pd.DataFrame([{col: "" for col in df1.columns}])
        summary_row1.iloc[0, 0] = summary1
        df1_with_summary = pd.concat([df1, summary_row1], ignore_index=True)
        out1 = build_report_filename(report_date1, report_location1)
        out1_path = os.path.join(RESULTS_FOLDER, out1)
        save_to_excel(df1_with_summary, out1_path)
        summary += summary1 + "\n"
        preview1 = df1_with_summary.to_html(classes="table table-bordered table-sm", index=False, border=0, justify="center")

    if file2 and file2.filename:
        filename2 = secure_filename(file2.filename)
        path2 = os.path.join(UPLOAD_FOLDER, filename2)
        file2.save(path2)
        reg2, ot2, total2, pct2, df2, report_date2, report_location2 = process_payroll(
            path2,
            include_kingsville_only_row=False,
            fallback_location="Alice"
        )
        summary2 = f"Total hours for the BHB Alice for the week is {total2:.2f} of which {ot2:.2f} or {pct2:.2%} is considered overtime."
        summary_row2 = pd.DataFrame([{col: "" for col in df2.columns}])
        summary_row2.iloc[0, 0] = summary2
        df2_with_summary = pd.concat([df2, summary_row2], ignore_index=True)
        out2 = build_report_filename(report_date2, report_location2)
        out2_path = os.path.join(RESULTS_FOLDER, out2)
        save_to_excel(df2_with_summary, out2_path)
        summary += summary2 + "\n"
        preview2 = df2_with_summary.to_html(classes="table table-bordered table-sm", index=False, border=0, justify="center")

    if not summary:
        flash('Please upload at least one CSV file.')
        return redirect(url_for('index'))

    return render_template('index.html', summary=summary.strip(), out1=out1, out2=out2, preview1=preview1, preview2=preview2)

@app.route('/download/<filename>')
def download(filename):
    file_path = os.path.join(RESULTS_FOLDER, filename)
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    flash('File not found.')
    return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=True)
