from flask import send_from_directory
"""
Manager App - Flask Web Application
Multi-tenant restaurant management system
"""
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import check_password_hash, generate_password_hash
from functools import wraps
import os
import csv
from datetime import datetime, timedelta
import secrets

# Import existing modules
import sys
sys.path.insert(0, os.path.dirname(__file__))

from database import get_db
from security import InputValidator, EncryptionHelper

app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', secrets.token_hex(32))
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=30)  # Remember me for 30 days
app.config['REMEMBER_COOKIE_DURATION'] = timedelta(days=30)  # Remember me cookie duration
app.config['SESSION_COOKIE_SECURE'] = False  # Set to True only with HTTPS
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['TEMPLATES_AUTO_RELOAD'] = True  # Auto-reload templates in development
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0  # Disable caching for static files in development

# Flask-Login setup
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Please log in to access this page.'

db = get_db()


# ==================== DATA STORAGE HELPERS ====================
def create_employee_json(company_id, employee):
    """Create a JSON file for a new employee with basic info, discipline, and certifications."""
    emp_dir = f"company_data/{company_id}/employees"
    os.makedirs(emp_dir, exist_ok=True)
    filename = f"{employee['firstName']}_{employee['lastName']}.json"
    filepath = os.path.join(emp_dir, filename)
    data = {
        "employeeInfo": {
            "firstName": employee["firstName"],
            "lastName": employee["lastName"],
            "dateOfHire": employee["dateOfHire"],
            "id": employee["id"]
        },
        "disciplinaryReports": [],
        "certifications": {
            "TABC": None,
            "FoodHandlers": None,
            "Other": []
        }
    }
    import json
    with open(filepath, "w") as f:
        json.dump(data, f, indent=2)
    return filepath

def save_daily_log(company_id, log_data):
    """Save daily log to CSV file - matching desktop dailylog.py format"""
    data_dir = f"company_data/{company_id}/daily_logs"
    os.makedirs(data_dir, exist_ok=True)
    
    date_str = log_data['date'].replace('-', '')
    shift = log_data.get('shift', 'Day')
    filepath = f"{data_dir}/{date_str}_{shift}.csv"
    
    with open(filepath, 'w', newline='') as f:
        writer = csv.writer(f)
        
        # Header matching desktop format
        writer.writerow(['Date', log_data['date']])
        writer.writerow(['Shift', shift])
        writer.writerow(['Notes', log_data.get('notes', '')])
        writer.writerow([])
        
        # Employee entries section
        writer.writerow(['Employee Entries'])
        writer.writerow(['Name', 'Shift', 'Area', 'Cash', 'C.C. Tips', 'Cash Diff', 'Visa', 'Mastercard', 'Amex', 'Discover',
                        'Credit Total', 'Beer', 'Liquor', 'Wine', 'Food', 'Voids'])
        
        employees = log_data.get('employees', [])
        for emp in employees:
            writer.writerow([
                emp.get('name', ''),
                emp.get('shift', 'Day'),
                emp.get('area', ''),
                emp.get('cash', 0),
                emp.get('cc_tips', 0),
                emp.get('cash_diff', 0),
                emp.get('visa', 0),
                emp.get('mastercard', 0),
                emp.get('amex', 0),
                emp.get('discover', 0),
                emp.get('credit', 0),
                emp.get('beer', 0),
                emp.get('liquor', 0),
                emp.get('wine', 0),
                emp.get('food', 0),
                emp.get('voids', 0)
            ])
        
        # Cash drawer count
        writer.writerow([])
        writer.writerow(['Cash Drawer Count'])
        writer.writerow(['Pennies', log_data.get('pennies', 0)])
        writer.writerow(['Nickels', log_data.get('nickels', 0)])
        writer.writerow(['Dimes', log_data.get('dimes', 0)])
        writer.writerow(['Quarters', log_data.get('quarters', 0)])
        writer.writerow(['Ones', log_data.get('ones', 0)])
        writer.writerow(['Fives', log_data.get('fives', 0)])
        writer.writerow(['Tens', log_data.get('tens', 0)])
        writer.writerow(['Twenties', log_data.get('twenties', 0)])
        writer.writerow(['Fifties', log_data.get('fifties', 0)])
        writer.writerow(['Hundreds', log_data.get('hundreds', 0)])
        writer.writerow(['Drawer Total', log_data.get('drawer_total', 0)])
        
        # Cash Deductions
        writer.writerow([])
        writer.writerow(['Cash Deductions'])
        deduction_descs = log_data.get('deduction_descs', [])
        deduction_locations = log_data.get('deduction_locations', [])
        deduction_amounts = log_data.get('deduction_amounts', [])
        for i, desc in enumerate(deduction_descs):
            if desc:
                location = deduction_locations[i] if i < len(deduction_locations) else ''
                amount = deduction_amounts[i] if i < len(deduction_amounts) else '0'
                writer.writerow([desc, location, amount])
        
        # Deposit summary
        writer.writerow([])
        writer.writerow(['Deposit Summary'])
        total_deductions = sum(float(amt) for amt in deduction_amounts if amt)
        writer.writerow(['Cash Adjustments', total_deductions])
        writer.writerow(['Cash in Drawer', log_data.get('drawer_total', 0)])
        writer.writerow(['DEPOSIT AMOUNT', log_data.get('deposit_amount', 0)])


def load_daily_log(company_id, date_str):
    """Load daily log from CSV file - matching desktop dailylog.py format"""
    data_dir = f"company_data/{company_id}/daily_logs"
    date_str_clean = date_str.replace('-', '')

    # Try both Day and Night shifts
    filepath = f"{data_dir}/{date_str_clean}_Day.csv"
    legacy_dir = os.path.expanduser("~/Documents/AIO Python/daily_logs")
    legacy_day = os.path.join(legacy_dir, f"{date_str}_Day.csv")
    legacy_night = os.path.join(legacy_dir, f"{date_str}_Night.csv")

    if not os.path.exists(filepath):
        # If not found, check legacy folder and copy if exists
        if os.path.exists(legacy_day):
            import shutil
            shutil.copy(legacy_day, filepath)
        elif os.path.exists(legacy_night):
            filepath = f"{data_dir}/{date_str_clean}_Night.csv"
            shutil.copy(legacy_night, filepath)

    if not os.path.exists(filepath):
        return None
    
    log_data = {
        'employees': [],
        'shift': 'Day',
        'notes': '',
        'pennies': 0,
        'nickels': 0,
        'dimes': 0,
        'quarters': 0,
        'ones': 0,
        'fives': 0,
        'tens': 0,
        'twenties': 0,
        'fifties': 0,
        'hundreds': 0,
        'drawer_total': 0,
        'deduction_descs': [],
        'deduction_locations': [],
        'deduction_amounts': [],
        'deposit_amount': 0
    }
    
    try:
        with open(filepath, 'r') as f:
            reader = csv.reader(f)
            rows = list(reader)
            
            section = None
            for i, row in enumerate(rows):
                if not row:
                    continue
                
                # Parse header info
                if row[0] == 'Shift' and len(row) > 1:
                    log_data['shift'] = row[1]
                elif row[0] == 'Notes' and len(row) > 1:
                    log_data['notes'] = row[1]
                
                # Detect sections
                elif row[0] == 'Employee Entries':
                    section = 'employees'
                    continue
                elif row[0] == 'Cash Drawer Count':
                    section = 'drawer'
                    continue
                elif row[0] == 'Deductions':
                    section = 'deductions'
                    continue
                elif row[0] == 'Cash Deductions':
                    section = 'deductions'
                    continue
                elif row[0] == 'Deposit Summary':
                    section = 'deposit'
                    continue
                
                # Parse employee data
                if section == 'employees' and row[0] != 'Name' and len(row) >= 13:
                    employee = {
                        'name': row[0],
                        'shift': row[1] if len(row) > 1 else 'Day',
                        'area': row[2] if len(row) > 2 else '',
                        'cash': float(row[3]) if len(row) > 3 and row[3] else 0,
                        'cc_tips': float(row[4]) if len(row) > 4 and row[4] else 0,
                        'cash_diff': float(row[5]) if len(row) > 5 and row[5] else 0,
                        'visa': float(row[6]) if len(row) > 6 and row[6] else 0,
                        'mastercard': float(row[7]) if len(row) > 7 and row[7] else 0,
                        'amex': float(row[8]) if len(row) > 8 and row[8] else 0,
                        'discover': float(row[9]) if len(row) > 9 and row[9] else 0,
                        'credit': float(row[10]) if len(row) > 10 and row[10] else 0,
                        'beer': float(row[11]) if len(row) > 11 and row[11] else 0,
                        'liquor': float(row[12]) if len(row) > 12 and row[12] else 0,
                        'wine': float(row[13]) if len(row) > 13 and row[13] else 0,
                        'food': float(row[14]) if len(row) > 14 and row[14] else 0,
                        'voids': float(row[15]) if len(row) > 15 and row[15] else 0
                    }
                    log_data['employees'].append(employee)
                
                # Parse drawer data
                elif section == 'drawer' and len(row) >= 2:
                    key = row[0].lower().replace(' ', '_')
                    if key in ['pennies', 'nickels', 'dimes', 'quarters', 'ones', 'fives', 'tens', 'twenties', 'fifties', 'hundreds']:
                        log_data[key] = float(row[1]) if row[1] else 0
                    elif key == 'drawer_total':
                        log_data['drawer_total'] = float(row[1]) if row[1] else 0
                
                # Parse deductions (support both old 2-column and new 3-column format)
                elif section == 'deductions' and len(row) >= 2 and row[0] not in ['Total Deductions']:
                    log_data['deduction_descs'].append(row[0])
                    # Check if this is the new 3-column format (desc, location, amount)
                    if len(row) >= 3:
                        log_data['deduction_locations'].append(row[1] if row[1] else '')
                        log_data['deduction_amounts'].append(float(row[2]) if row[2] else 0)
                    else:
                        # Old 2-column format (desc, amount)
                        log_data['deduction_locations'].append('')
                        log_data['deduction_amounts'].append(float(row[1]) if row[1] else 0)
                
                # Parse deposit
                elif section == 'deposit' and len(row) >= 2:
                    if row[0] == 'DEPOSIT AMOUNT':
                        log_data['deposit_amount'] = float(row[1]) if row[1] else 0
    
    except Exception as e:
        print(f"Error loading daily log: {e}")
        return None
    
    # Calculate deposit if it's missing (for backwards compatibility with old CSV files)
    if log_data['deposit_amount'] == 0 and log_data['employees']:
        total_cash = sum(emp.get('cash', 0) for emp in log_data['employees'])
        total_cc_tips = sum(emp.get('cc_tips', 0) for emp in log_data['employees'])
        total_deductions = sum(log_data['deduction_amounts'])
        log_data['deposit_amount'] = total_cash - total_cc_tips - total_deductions
    
    return log_data


def save_cash_drawer(company_id, drawer_data):
    """Save cash drawer counts to CSV file"""
    data_dir = f"company_data/{company_id}/daily_logs"
    os.makedirs(data_dir, exist_ok=True)
    
    date_str = drawer_data['date'].replace('-', '')
    shift = drawer_data.get('shift', 'Day')
    filepath = f"{data_dir}/{date_str}_{shift}.csv"
    
    cash_section = [
        ['Cash Drawer Count'],
        ['Pennies', drawer_data.get('pennies', 0)],
        ['Nickels', drawer_data.get('nickels', 0)],
        ['Dimes', drawer_data.get('dimes', 0)],
        ['Quarters', drawer_data.get('quarters', 0)],
        ['$1', drawer_data.get('ones', 0)],
        ['$5', drawer_data.get('fives', 0)],
        ['$10', drawer_data.get('tens', 0)],
        ['$20', drawer_data.get('twenties', 0)],
        ['$50', drawer_data.get('fifties', 0)],
        ['$100', drawer_data.get('hundreds', 0)],
        ['Total', drawer_data.get('total', 0)]
    ]
    
    with open(filepath, 'w', newline='') as f:
        writer = csv.writer(f)
        for row in cash_section:
            writer.writerow(row)


def load_cash_drawer(company_id, date_str):
    """Load cash drawer counts from CSV file"""
    data_dir = f"company_data/{company_id}/daily_logs"
    date_str = date_str.replace('-', '')
    
    drawer_data = {
        'pennies': 0, 'nickels': 0, 'dimes': 0, 'quarters': 0,
        'ones': 0, 'fives': 0, 'tens': 0, 'twenties': 0,
        'fifties': 0, 'hundreds': 0, 'total': 0
    }
    
    for shift in ['Day', 'Night']:
        filepath = f"{data_dir}/{date_str}_{shift}.csv"
        if os.path.exists(filepath):
            try:
                with open(filepath, 'r') as f:
                    reader = csv.reader(f)
                    for row in reader:
                        if len(row) >= 2:
                            if row[0] == 'Pennies':
                                drawer_data['pennies'] = row[1]
                            elif row[0] == 'Nickels':
                                drawer_data['nickels'] = row[1]
                            elif row[0] == 'Dimes':
                                drawer_data['dimes'] = row[1]
                            elif row[0] == 'Quarters':
                                drawer_data['quarters'] = row[1]
                            elif row[0] == '$1':
                                drawer_data['ones'] = row[1]
                            elif row[0] == '$5':
                                drawer_data['fives'] = row[1]
                            elif row[0] == '$10':
                                drawer_data['tens'] = row[1]
                            elif row[0] == '$20':
                                drawer_data['twenties'] = row[1]
                            elif row[0] == '$50':
                                drawer_data['fifties'] = row[1]
                            elif row[0] == '$100':
                                drawer_data['hundreds'] = row[1]
                            elif row[0] == 'Total':
                                drawer_data['total'] = row[1]
                break
            except:
                pass
    
    return drawer_data


def save_cash_deduction(company_id, deduction_data):
    """Save cash deduction to CSV file"""
    data_dir = f"company_data/{company_id}/daily_logs"
    os.makedirs(data_dir, exist_ok=True)
    
    date_str = deduction_data['date'].replace('-', '')
    filepath = f"{data_dir}/{date_str}_CashDeductions.csv"
    
    with open(filepath, 'a', newline='') as f:
        writer = csv.writer(f)
        writer.writerow([
            deduction_data.get('description', ''),
            deduction_data.get('amount', 0),
            datetime.now().isoformat()
        ])


def load_cash_deductions(company_id, date_str):
    """Load cash deductions from CSV file"""
    data_dir = f"company_data/{company_id}/daily_logs"
    date_str = date_str.replace('-', '')
    filepath = f"{data_dir}/{date_str}_CashDeductions.csv"
    
    deductions = []
    if os.path.exists(filepath):
        try:
            with open(filepath, 'r') as f:
                reader = csv.reader(f)
                for i, row in enumerate(reader):
                    if len(row) >= 2:
                        deductions.append({
                            'id': i,
                            'description': row[0],
                            'amount': row[1],
                            'timestamp': row[2] if len(row) > 2 else ''
                        })
        except:
            pass
    
    return deductions


def delete_cash_deduction(deduction_id):
    """Delete a cash deduction"""
    pass  # Implement if needed


class User(UserMixin):
    def __init__(self, user_data):
        self.id = user_data['id']
        self.username = user_data['username']
        self.email = user_data['email']
        self.full_name = user_data.get('full_name', user_data['username'])
        self.is_system_admin = bool(user_data.get('is_system_admin', 0))
        self.companies = []
        self.current_company_id = None
        self.current_role = None


@login_manager.user_loader
def load_user(user_id):
    """Load user from database"""
    conn = db.get_connection()
    try:
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM users WHERE id = ? AND is_active = 1', (user_id,))
        user_data = cursor.fetchone()
    finally:
        conn.close()
    
    if user_data:
        user = User(dict(user_data))
        user.companies = db.get_user_companies(user_id)
        
        # Load current company from session
        if session.get('current_company_id'):
            user.current_company_id = session['current_company_id']
            company = next((c for c in user.companies if c['id'] == user.current_company_id), None)
            if company:
                user.current_role = company['role']
        
        return user
    return None


def company_required(f):
    """Decorator to require company selection"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.current_company_id:
            flash('Please select a company first.', 'warning')
            return redirect(url_for('select_company'))
        return f(*args, **kwargs)
    return decorated_function


def role_required(*roles):
    """Decorator to require specific role"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if current_user.current_role not in roles and not current_user.is_system_admin:
                flash('You do not have permission to access this page.', 'danger')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator


# ==================== ROUTES ====================

@app.route('/')
def index():
    """Landing page"""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return render_template('index.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    """Login page"""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        remember = request.form.get('remember') == 'on'  # Checkbox returns 'on' when checked
        
        if not username or not password:
            flash('Please enter username and password.', 'warning')
            return render_template('login.html')
        
        try:
            user_data = db.authenticate_user(username, password)
            
            if user_data:
                # Check terms acceptance BEFORE creating User object
                if not user_data.get('accepted_terms_version'):
                    session['pending_user_id'] = user_data['id']
                    session['pending_username'] = username  # Store username for re-login after terms
                    session.modified = True  # Ensure session is saved
                    flash('Please accept the Terms of Service to continue.', 'info')
                    return redirect(url_for('terms_acceptance'))
                
                user = User(user_data)
                user.companies = db.get_user_companies(user_data['id'])
                
                # Set session to permanent if remember me is checked
                if remember:
                    session.permanent = True
                
                login_user(user, remember=remember)
                db.log_action(user.id, 'login', None)
                
                next_page = request.args.get('next')
                if next_page:
                    return redirect(next_page)
                
                if len(user.companies) > 1:
                    return redirect(url_for('select_company'))
                elif len(user.companies) == 1:
                    session['current_company_id'] = user.companies[0]['id']
                    return redirect(url_for('dashboard'))
                else:
                    return redirect(url_for('create_company'))
            else:
                flash('Invalid username or password.', 'danger')
        
        except PermissionError as e:
            flash(str(e), 'danger')
        except ValueError as e:
            flash(str(e), 'warning')
        except Exception as e:
            flash(f'An error occurred: {str(e)}', 'danger')
    
    return render_template('login.html')


@app.route('/register', methods=['GET', 'POST'])
def register():
    """Registration page"""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        confirm = request.form.get('confirm_password', '')
        full_name = request.form.get('full_name', '').strip()
        
        # Validation
        validator = InputValidator()
        
        if not username or not email or not password:
            flash('Please fill in all required fields.', 'warning')
            return render_template('register.html')
        
        if len(username) < 3:
            flash('Username must be at least 3 characters.', 'warning')
            return render_template('register.html')
        
        if not validator.validate_email(email):
            flash('Please enter a valid email address.', 'warning')
            return render_template('register.html')
        
        password_check = validator.validate_password_strength(password)
        if not password_check['valid']:
            flash('Password must have: ' + ', '.join(password_check['missing']), 'warning')
            return render_template('register.html')
        
        if password != confirm:
            flash('Passwords do not match.', 'warning')
            return render_template('register.html')
        
        try:
            user_id, token = db.create_user(username, email, password, full_name)
            
            if user_id:
                session['pending_user_id'] = user_id
                flash('Account created successfully! Please accept the Terms of Service.', 'success')
                return redirect(url_for('terms_acceptance'))
            else:
                flash('Username or email already exists.', 'danger')
        except Exception as e:
            flash(f'Registration failed: {str(e)}', 'danger')
    
    return render_template('register.html')


@app.route('/terms', methods=['GET', 'POST'])
def terms_acceptance():
    """Terms of Service acceptance"""
    pending_user_id = session.get('pending_user_id')
    
    # Debug logging
    print(f"DEBUG: Terms page - pending_user_id: {pending_user_id}")
    print(f"DEBUG: Session contents: {dict(session)}")
    
    if not pending_user_id:
        flash('Session expired. Please login again.', 'warning')
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        if request.form.get('accept') == 'true':
            # Record acceptance
            conn = db.get_connection()
            try:
                cursor = conn.cursor()
                cursor.execute('''
                    UPDATE users 
                    SET accepted_terms_version = ?,
                        accepted_terms_date = ?
                    WHERE id = ?
                ''', ('1.0', datetime.now().isoformat(), pending_user_id))
                conn.commit()
            finally:
                conn.close()
            
            db.log_action(pending_user_id, 'terms_accepted', None, {'version': '1.0'})
            
            # Login user
            conn = db.get_connection()
            try:
                cursor = conn.cursor()
                cursor.execute('SELECT * FROM users WHERE id = ?', (pending_user_id,))
                user_data = dict(cursor.fetchone())
            finally:
                conn.close()
            
            user = User(user_data)
            user.companies = db.get_user_companies(pending_user_id)
            login_user(user)
            
            session.pop('pending_user_id', None)
            
            if user.companies:
                return redirect(url_for('select_company'))
            else:
                return redirect(url_for('create_company'))
        else:
            session.pop('pending_user_id', None)
            flash('You must accept the Terms of Service to use this application.', 'warning')
            return redirect(url_for('login'))
    
    return render_template('terms.html')


@app.route('/logout')
@login_required
def logout():
    """Logout"""
    db.log_action(current_user.id, 'logout', current_user.current_company_id)
    logout_user()
    session.clear()
    flash('You have been logged out.', 'info')
    return redirect(url_for('index'))


@app.route('/select-company')
@login_required
def select_company():
    """Company selection page"""
    if not current_user.companies:
        return redirect(url_for('create_company'))
    
    return render_template('select_company.html', companies=current_user.companies)


@app.route('/select-company/<company_id>')
@login_required
def set_company(company_id):
    """Set active company"""
    company = next((c for c in current_user.companies if c['id'] == company_id), None)
    
    if not company:
        flash('You do not have access to this company.', 'danger')
        return redirect(url_for('select_company'))
    
    session['current_company_id'] = company_id
    current_user.current_company_id = company_id
    current_user.current_role = company['role']
    
    db.log_action(current_user.id, 'company_selected', company_id)
    
    return redirect(url_for('dashboard'))


@app.route('/create-company', methods=['GET', 'POST'])
@login_required
def create_company():
    """Create new company"""
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        address = request.form.get('address', '').strip()
        phone = request.form.get('phone', '').strip()
        email = request.form.get('email', '').strip()
        
        if not name:
            flash('Company name is required.', 'warning')
            return render_template('create_company.html')
        
        company_id = db.create_company(
            name=name,
            admin_user_id=current_user.id,
            address=address,
            phone=phone,
            email=email
        )
        
        if company_id:
            db.log_action(current_user.id, 'company_created', company_id, {'company_name': name})
            
            # Reload user companies
            current_user.companies = db.get_user_companies(current_user.id)
            session['current_company_id'] = company_id
            
            flash(f'Company "{name}" created successfully!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Company name already exists.', 'danger')
    
    return render_template('create_company.html')


@app.route('/dashboard')
@login_required
@company_required
def dashboard():
    """Main dashboard"""
    company = db.get_company(current_user.current_company_id)
    
    # Get today's sales from daily log
    today = datetime.now().strftime('%Y-%m-%d')
    today_log = load_daily_log(current_user.current_company_id, today)
    today_sales = float(today_log.get('total_sales', 0)) if today_log else 0
    
    # Get today's cash on hand from cash drawer
    today_drawer = load_cash_drawer(current_user.current_company_id, today)
    cash_on_hand = float(today_drawer.get('total', 0)) if today_drawer else 0
    
    # Get recent activity
    recent_logs = db.get_audit_log(
        company_id=current_user.current_company_id,
        limit=10
    )
    
    # Get company users
    users = db.get_company_users(current_user.current_company_id)
    
    return render_template(
        'dashboard.html',
        company=company,
        recent_logs=recent_logs,
        users=users,
        user_count=len(users),
        today_sales=today_sales,
        cash_on_hand=cash_on_hand
    )


@app.route('/daily-log', methods=['GET', 'POST'])
@login_required
@company_required
def daily_log():
    """Daily operations log"""
    selected_date = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    
    if request.method == 'POST':
        # Save daily log data
        try:
            import json
            
            # Parse employee data from JSON
            employee_data_json = request.form.get('employee_data', '[]')
            employees = json.loads(employee_data_json) if employee_data_json else []
            
            log_data = {
                'date': request.form.get('date'),
                'shift': request.form.get('shift', 'Day'),
                'notes': request.form.get('notes', ''),
                'employees': employees,
                # Cash drawer counts
                'pennies': request.form.get('pennies', 0),
                'nickels': request.form.get('nickels', 0),
                'dimes': request.form.get('dimes', 0),
                'quarters': request.form.get('quarters', 0),
                'ones': request.form.get('ones', 0),
                'fives': request.form.get('fives', 0),
                'tens': request.form.get('tens', 0),
                'twenties': request.form.get('twenties', 0),
                'fifties': request.form.get('fifties', 0),
                'hundreds': request.form.get('hundreds', 0),
                'drawer_total': request.form.get('drawer_total', 0),
                # Cash Deductions
                'deduction_descs': request.form.getlist('deduction_desc'),
                'deduction_locations': request.form.getlist('deduction_location'),
                'deduction_amounts': request.form.getlist('deduction_amount'),
                # Deposit
                'deposit_amount': request.form.get('deposit_amount', 0)
            }
            
            # Save to file
            save_daily_log(current_user.current_company_id, log_data)
            
            db.log_action(
                current_user.id,
                'daily_log_saved',
                current_user.current_company_id,
                {'date': log_data['date'], 'employees': len(employees), 'deposit': log_data['deposit_amount']}
            )
            
            flash(f'Daily log saved! {len(employees)} employees, Deposit: ${log_data["deposit_amount"]}', 'success')
            return redirect(url_for('daily_log', date=log_data['date']))
        except Exception as e:
            flash(f'Error saving daily log: {str(e)}', 'danger')
    
    # Load existing data for selected date
    log_data = load_daily_log(current_user.current_company_id, selected_date)
    
    return render_template('daily_log.html', date=selected_date, log_data=log_data)


@app.route('/cash-manager', methods=['GET', 'POST'])
@login_required
@company_required
def cash_manager():
    """Cash management"""
    selected_date = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    active_tab = request.args.get('tab', 'drawer')
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        try:
            if action == 'save_drawer':
                # Save cash drawer counts
                drawer_data = {
                    'date': request.form.get('date'),
                    'shift': request.form.get('shift'),
                    'pennies': request.form.get('pennies', 0),
                    'nickels': request.form.get('nickels', 0),
                    'dimes': request.form.get('dimes', 0),
                    'quarters': request.form.get('quarters', 0),
                    'ones': request.form.get('ones', 0),
                    'fives': request.form.get('fives', 0),
                    'tens': request.form.get('tens', 0),
                    'twenties': request.form.get('twenties', 0),
                    'fifties': request.form.get('fifties', 0),
                    'hundreds': request.form.get('hundreds', 0),
                    'total': request.form.get('total', 0)
                }
                save_cash_drawer(current_user.current_company_id, drawer_data)
                flash('Cash drawer saved successfully!', 'success')
                
            elif action == 'save_deduction':
                # Save cash deduction
                deduction_data = {
                    'date': request.form.get('date'),
                    'description': request.form.get('description'),
                    'amount': request.form.get('amount', 0)
                }
                save_cash_deduction(current_user.current_company_id, deduction_data)
                flash('Deduction saved successfully!', 'success')
                
            elif action == 'delete_deduction':
                deduction_id = request.form.get('deduction_id')
                delete_cash_deduction(deduction_id)
                flash('Deduction deleted successfully!', 'success')
            
            db.log_action(
                current_user.id,
                f'cash_manager_{action}',
                current_user.current_company_id,
                {'date': request.form.get('date')}
            )
            
            return redirect(url_for('cash_manager', date=request.form.get('date'), tab=active_tab))
        except Exception as e:
            flash(f'Error: {str(e)}', 'danger')
    
    # Load data
    drawer_data = load_cash_drawer(current_user.current_company_id, selected_date)
    deductions = load_cash_deductions(current_user.current_company_id, selected_date)
    
    return render_template(
        'cash_manager.html',
        date=selected_date,
        active_tab=active_tab,
        drawer_data=drawer_data,
        deductions=deductions
    )


@app.route('/employees')
@login_required
@company_required
def employees():
    """Employee management"""
    return render_template('employees.html')


@app.route('/reports')
@login_required
@company_required
def reports():
    """Reports and analytics"""
    # Get available date range from daily logs
    data_dir = f"company_data/{current_user.current_company_id}/daily_logs"
    available_dates = []
    
    if os.path.exists(data_dir):
        for filename in os.listdir(data_dir):
            if filename.endswith('.csv') and not filename.startswith('.'):
                # Extract date from filename (format: YYYYMMDD_Shift.csv)
                date_str = filename.split('_')[0]
                if len(date_str) == 8 and date_str.isdigit():
                    try:
                        date_obj = datetime.strptime(date_str, '%Y%m%d')
                        available_dates.append(date_obj.strftime('%Y-%m-%d'))
                    except:
                        pass
    
    available_dates = sorted(set(available_dates), reverse=True)
    
    return render_template('reports.html', available_dates=available_dates)


@app.route('/api/reports/daily-summary')
@login_required
@company_required
def api_daily_summary():
    """Get daily summary report data"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    shift_filter = request.args.get('shift_filter', 'Full')  # Default to Full (combined)
    
    data_dir = f"company_data/{current_user.current_company_id}/daily_logs"
    summary_data = []
    
    if not os.path.exists(data_dir):
        return jsonify({'success': True, 'data': []})
    
    # Parse date range
    try:
        start = datetime.strptime(start_date, '%Y-%m-%d') if start_date else None
        end = datetime.strptime(end_date, '%Y-%m-%d') if end_date else None
    except:
        return jsonify({'success': False, 'error': 'Invalid date format'})
    
    # Group files by date to combine Day and Night shifts
    date_files = {}
    import sys
    for filename in os.listdir(data_dir):
        if not filename.endswith('.csv') or filename.startswith('.'):
            print(f"Skipping file (not CSV or hidden): {filename}", file=sys.stderr)
            continue

        date_str = filename.split('_')[0]
        if len(date_str) != 8 or not date_str.isdigit():
            print(f"Skipping file (bad date format): {filename}", file=sys.stderr)
            continue

        try:
            file_date = datetime.strptime(date_str, '%Y%m%d')

            # Check if in date range
            if start and file_date < start:
                print(f"Skipping file (before start date): {filename}", file=sys.stderr)
                continue
            if end and file_date > end:
                print(f"Skipping file (after end date): {filename}", file=sys.stderr)
                continue

            date_key = file_date.strftime('%Y-%m-%d')
            if date_key not in date_files:
                date_files[date_key] = []
            date_files[date_key].append(filename)
        except Exception as e:
            print(f"Skipping file (exception): {filename} - {e}", file=sys.stderr)
            continue
    
    # Process each date (combining Day and Night shifts)
    for date_key, filenames in date_files.items():
        try:
            combined_employees = []
            combined_deposit = 0
            combined_deductions = 0
            
            # Load all shifts for this date
            for filename in filenames:
                filepath = f"{data_dir}/{filename}"
                
                log_data = {
                    'employees': [],
                    'deposit_amount': 0,
                    'deductions': 0
                }
                
                # Read CSV file directly
                try:
                    with open(filepath, 'r') as f:
                        reader = csv.reader(f)
                        rows = list(reader)
                        
                        section = None
                        print(f"\n=== PARSING FILE: {filename} ===")  # Debug
                        for row_idx, row in enumerate(rows):
                            if not row:
                                continue
                            
                            if row[0] == 'Employee Entries':
                                section = 'employees'
                                print(f"Row {row_idx}: Section changed to 'employees'")  # Debug
                                continue
                            elif row[0] == 'Deposit Summary':
                                section = 'deposit'
                                print(f"Row {row_idx}: Section changed to 'deposit'")  # Debug
                                continue
                            elif row[0] in ['Cash Drawer Count', 'Deductions', 'Cash Deductions']:
                                section = None
                                print(f"Row {row_idx}: Section set to None ({row[0]})")  # Debug
                                continue
                            
                            # Parse employee data
                            if section == 'employees' and row[0] != 'Name' and row[0] and len(row) >= 15:
                                try:
                                    # Get employee shift from row (column index 1)
                                    emp_shift = row[1] if len(row) > 1 and row[1] else 'Day'
                                    
                                    # Filter by shift if not Full
                                    if shift_filter != 'Full' and emp_shift != shift_filter:
                                        continue
                                    
                                    employee = {
                                        'shift': emp_shift,
                                        'cash': float(row[3]) if len(row) > 3 and row[3] else 0,
                                        'cc_tips': float(row[4]) if len(row) > 4 and row[4] else 0,
                                        'cash_diff': float(row[5]) if len(row) > 5 and row[5] else 0,
                                        'visa': float(row[6]) if len(row) > 6 and row[6] else 0,
                                        'mastercard': float(row[7]) if len(row) > 7 and row[7] else 0,
                                        'amex': float(row[8]) if len(row) > 8 and row[8] else 0,
                                        'discover': float(row[9]) if len(row) > 9 and row[9] else 0,
                                        'beer': float(row[11]) if len(row) > 11 and row[11] else 0,
                                        'liquor': float(row[12]) if len(row) > 12 and row[12] else 0,
                                        'wine': float(row[13]) if len(row) > 13 and row[13] else 0,
                                        'food': float(row[14]) if len(row) > 14 and row[14] else 0,
                                        'voids': float(row[15]) if len(row) > 15 and row[15] else 0,
                                    }
                                    log_data['employees'].append(employee)
                                    combined_employees.append(employee)
                                except (ValueError, IndexError) as e:
                                    print(f"Error parsing employee row: {e}")
                                    continue
                            
                            # Parse deposit
                            elif section == 'deposit' and len(row) >= 2:
                                print(f"Row {row_idx} in deposit section: {row[0]} = {row[1]}")  # Debug
                                if row[0] == 'DEPOSIT AMOUNT':
                                    try:
                                        deposit_val = float(row[1]) if row[1] else 0
                                        log_data['deposit_amount'] = deposit_val
                                        print(f"  -> Found deposit in {filename}: {deposit_val}")  # Debug
                                    except ValueError:
                                        print(f"  -> ValueError parsing deposit: {row[1]}")  # Debug
                                        pass
                                elif row[0] in ['Cash Adjustments', 'Total Deductions']:
                                    try:
                                        deductions_val = float(row[1]) if row[1] else 0
                                        log_data['deductions'] = deductions_val
                                        print(f"  -> Found deductions in {filename}: {deductions_val}")  # Debug
                                    except ValueError:
                                        print(f"  -> ValueError parsing deductions: {row[1]}")  # Debug
                                        pass
                except Exception as e:
                    print(f"Error reading file {filename}: {e}")
                    continue
                
                # Calculate deposit if it's missing (backwards compatibility)
                if log_data['deposit_amount'] == 0 and log_data['employees']:
                    total_cash = sum(emp.get('cash', 0) for emp in log_data['employees'])
                    total_cc_tips = sum(emp.get('cc_tips', 0) for emp in log_data['employees'])
                    log_data['deposit_amount'] = total_cash - total_cc_tips - log_data['deductions']
                    print(f"Calculated deposit for {filename}: {log_data['deposit_amount']}")  # Debug
                
                # Add to combined totals
                combined_deposit += log_data['deposit_amount']
                combined_deductions += log_data['deductions']
                print(f"File {filename}: deposit={log_data['deposit_amount']}, deductions={log_data['deductions']}")  # Debug
                print(f"Running totals: combined_deposit={combined_deposit}, combined_deductions={combined_deductions}")  # Debug
            
            if combined_employees:
                # Calculate totals from all shifts
                total_beer = sum(emp.get('beer', 0) for emp in combined_employees)
                total_liquor = sum(emp.get('liquor', 0) for emp in combined_employees)
                total_wine = sum(emp.get('wine', 0) for emp in combined_employees)
                total_food = sum(emp.get('food', 0) for emp in combined_employees)
                total_sales = total_beer + total_liquor + total_wine + total_food
                total_cash = sum(emp.get('cash', 0) for emp in combined_employees)
                total_credit = sum(
                    (emp.get('visa', 0) + emp.get('mastercard', 0) + 
                     emp.get('amex', 0) + emp.get('discover', 0))
                    for emp in combined_employees
                )
                total_tips = sum(emp.get('cc_tips', 0) for emp in combined_employees)
                total_cash_adjustments = combined_deductions  # Use actual deductions, not cash_diff
                employee_count = len(combined_employees)
                
                # Calculate voids
                total_voids = sum(emp.get('voids', 0) for emp in combined_employees)
                
                # Calculate percentages
                tip_percentage = (total_tips / total_sales * 100) if total_sales > 0 else 0
                void_percentage = (total_voids / total_sales * 100) if total_sales > 0 else 0
                
                print(f"DEBUG - Date {date_key}: total_cash_adjustments = {total_cash_adjustments}, combined_deductions = {combined_deductions}")  # Debug
                
                # Determine shift label based on what was filtered
                if shift_filter == 'Full':
                    # Check if we have both Day and Night employees
                    shifts_present = set(emp.get('shift', 'Day') for emp in combined_employees)
                    shift_label = 'Full' if len(shifts_present) > 1 else list(shifts_present)[0]
                else:
                    shift_label = shift_filter
                
                print(f"Date: {date_key}, Combined Deposit: {combined_deposit}")  # Debug
                
                summary_data.append({
                    'date': date_key,
                    'shift': shift_label,
                    'employees': employee_count,
                    'total_sales': round(total_sales, 2),
                    'total_cash': round(total_cash, 2),
                    'total_credit': round(total_credit, 2),
                    'total_tips': round(total_tips, 2),
                    'total_cash_adjustments': round(total_cash_adjustments, 2),
                    'deposit': round(combined_deposit, 2),
                    'beer': round(total_beer, 2),
                    'liquor': round(total_liquor, 2),
                    'wine': round(total_wine, 2),
                    'food': round(total_food, 2),
                    'voids': round(total_voids, 2),
                    'tip_percentage': round(tip_percentage, 2),
                    'void_percentage': round(void_percentage, 2)
                })
        except Exception as e:
            print(f"Error processing date {date_key}: {e}")
            import traceback
            traceback.print_exc()
            continue
    
    # Sort by date
    summary_data.sort(key=lambda x: x['date'], reverse=True)
    
    print(f"Daily summary returning {len(summary_data)} records")
    return jsonify({'success': True, 'data': summary_data})


# === NEW ENDPOINT: Cash Deductions Report ===
@app.route('/api/reports/cash-deductions')
@login_required
@company_required
def api_cash_deductions():
    """Get total cash deductions for a date range"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    data_dir = f"company_data/{current_user.current_company_id}/daily_logs"
    total_deductions = 0.0
    daily_deductions = []
    if not os.path.exists(data_dir):
        return jsonify({'success': True, 'total_deductions': 0.0, 'daily': []})
    try:
        start = datetime.strptime(start_date, '%Y-%m-%d') if start_date else None
        end = datetime.strptime(end_date, '%Y-%m-%d') if end_date else None
    except:
        return jsonify({'success': False, 'error': 'Invalid date format'})
    for filename in os.listdir(data_dir):
        if not filename.endswith('.csv') or filename.startswith('.'):
            continue
        date_str = filename.split('_')[0]
        if len(date_str) != 8 or not date_str.isdigit():
            continue
        try:
            file_date = datetime.strptime(date_str, '%Y%m%d')
            if start and file_date < start:
                continue
            if end and file_date > end:
                continue
            # Read CSV and extract deductions
            filepath = f"{data_dir}/{filename}"
            with open(filepath, 'r') as f:
                reader = csv.reader(f)
                rows = list(reader)
                found = False
                for row in rows:
                    if row and row[0] in ['Cash Adjustments', 'Total Deductions']:
                        try:
                            deduction_val = float(row[1]) if len(row) > 1 and row[1] else 0.0
                            total_deductions += deduction_val
                            daily_deductions.append({
                                'date': file_date.strftime('%Y-%m-%d'),
                                'deductions': round(deduction_val, 2)
                            })
                            found = True
                        except Exception:
                            continue
                if not found:
                    daily_deductions.append({
                        'date': file_date.strftime('%Y-%m-%d'),
                        'deductions': 0.0
                    })
        except Exception:
            continue
    # Sort by date
    daily_deductions.sort(key=lambda x: x['date'], reverse=True)
    return jsonify({'success': True, 'total_deductions': round(total_deductions, 2), 'daily': daily_deductions})


@app.route('/api/reports/employee-performance')
@login_required
@company_required
def api_employee_performance():
    """Get employee performance report"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    employee_name = request.args.get('employee_name')
    shift_filter = request.args.get('shift_filter', 'Full')  # Default to Full (combined)
    
    data_dir = f"company_data/{current_user.current_company_id}/daily_logs"
    employee_data = {}
    daily_breakdown = {}  # Store daily data when filtering by specific employee
    
    if not os.path.exists(data_dir):
        return jsonify({'success': True, 'data': []})
    
    try:
        start = datetime.strptime(start_date, '%Y-%m-%d') if start_date else None
        end = datetime.strptime(end_date, '%Y-%m-%d') if end_date else None
    except:
        return jsonify({'success': False, 'error': 'Invalid date format'})
    
    # Aggregate employee data across all days
    for filename in os.listdir(data_dir):
        if not filename.endswith('.csv') or filename.startswith('.'):
            continue
            
        date_str = filename.split('_')[0]
        if len(date_str) != 8 or not date_str.isdigit():
            continue
            
        try:
            file_date = datetime.strptime(date_str, '%Y%m%d')
            
            if start and file_date < start:
                continue
            if end and file_date > end:
                continue
            
            log_data = load_daily_log(current_user.current_company_id, file_date.strftime('%Y-%m-%d'))
            
            if log_data and log_data.get('employees'):
                for emp in log_data['employees']:
                    name = emp.get('name', 'Unknown')
                    emp_shift = emp.get('shift', 'Day')  # Get employee's shift from their data
                    
                    # Filter by shift if not Full
                    if shift_filter != 'Full' and emp_shift != shift_filter:
                        continue
                    
                    # Filter by employee name if provided
                    if employee_name and name.lower() != employee_name.lower():
                        continue
                    
                    if name not in employee_data:
                        employee_data[name] = {
                            'name': name,
                            'shifts_worked': 0,
                            'total_sales': 0,
                            'total_cash': 0,
                            'total_credit': 0,
                            'total_tips': 0,
                            'total_voids': 0,
                            'total_beer': 0,
                            'total_liquor': 0,
                            'total_wine': 0,
                            'total_food': 0,
                            'avg_sales': 0,
                            'tip_percentage': 0
                        }
                    
                    sales = (emp.get('beer', 0) + emp.get('liquor', 0) + 
                            emp.get('wine', 0) + emp.get('food', 0))
                    
                    employee_data[name]['shifts_worked'] += 1
                    employee_data[name]['total_sales'] += sales
                    employee_data[name]['total_cash'] += emp.get('cash', 0)
                    employee_data[name]['total_credit'] += emp.get('credit', 0)
                    employee_data[name]['total_tips'] += emp.get('cc_tips', 0)
                    employee_data[name]['total_voids'] += emp.get('voids', 0)
                    employee_data[name]['total_beer'] += emp.get('beer', 0)
                    employee_data[name]['total_liquor'] += emp.get('liquor', 0)
                    employee_data[name]['total_wine'] += emp.get('wine', 0)
                    employee_data[name]['total_food'] += emp.get('food', 0)
                    
                    # Store daily breakdown if filtering by specific employee
                    if employee_name and name.lower() == employee_name.lower():
                        date_key = file_date.strftime('%Y-%m-%d')
                        shift = filename.split('_')[1].replace('.csv', '') if '_' in filename else 'Day'
                        
                        if name not in daily_breakdown:
                            daily_breakdown[name] = []
                        
                        tips = emp.get('cc_tips', 0)
                        tip_pct = round((tips / sales * 100), 2) if sales > 0 else 0
                        
                        daily_breakdown[name].append({
                            'date': date_key,
                            'shift': shift,
                            'sales': round(sales, 2),
                            'cash': round(emp.get('cash', 0), 2),
                            'credit': round(emp.get('credit', 0), 2),
                            'tips': round(tips, 2),
                            'tip_percentage': tip_pct,
                            'voids': round(emp.get('voids', 0), 2),
                            'beer': round(emp.get('beer', 0), 2),
                            'liquor': round(emp.get('liquor', 0), 2),
                            'wine': round(emp.get('wine', 0), 2),
                            'food': round(emp.get('food', 0), 2)
                        })
        except Exception as e:
            print(f"Error processing {filename}: {e}")
            continue
    
    # Calculate averages
    result = []
    for name, data in employee_data.items():
        if data['shifts_worked'] > 0:
            data['avg_sales'] = round(data['total_sales'] / data['shifts_worked'], 2)
        data['total_sales'] = round(data['total_sales'], 2)
        data['total_cash'] = round(data['total_cash'], 2)
        data['total_credit'] = round(data['total_credit'], 2)
        data['total_tips'] = round(data['total_tips'], 2)
        data['total_voids'] = round(data['total_voids'], 2)
        data['total_beer'] = round(data['total_beer'], 2)
        data['total_liquor'] = round(data['total_liquor'], 2)
        data['total_wine'] = round(data['total_wine'], 2)
        data['total_food'] = round(data['total_food'], 2)
        
        # Calculate overall tip percentage
        if data['total_sales'] > 0:
            data['tip_percentage'] = round((data['total_tips'] / data['total_sales'] * 100), 2)
        else:
            data['tip_percentage'] = 0
        
        # Add daily breakdown if available
        if name in daily_breakdown:
            data['daily_breakdown'] = sorted(daily_breakdown[name], key=lambda x: x['date'])
        
        result.append(data)
    
    # Sort by total sales descending
    result.sort(key=lambda x: x['total_sales'], reverse=True)
    
    return jsonify({'success': True, 'data': result})


@app.route('/settings')
@login_required
@company_required
@role_required('business_admin')
def settings():
    """Company settings"""
    company = db.get_company(current_user.current_company_id)
    return render_template('settings.html', company=company)


@app.route('/users')
@login_required
@company_required
@role_required('business_admin')
def manage_users():
    """User management"""
    users = db.get_company_users(current_user.current_company_id)
    return render_template('users.html', users=users)


@app.route('/audit-log')
@login_required
@company_required
@role_required('business_admin', 'manager')
def audit_log():
    """Audit log viewer"""
    logs = db.get_audit_log(company_id=current_user.current_company_id, limit=100)
    return render_template('audit_log.html', logs=logs)


# ==================== API ROUTES ====================

@app.route('/api/profile')
@login_required
def api_profile():
    """Get user profile"""
    return jsonify({
        'username': current_user.username,
        'email': current_user.email,
        'full_name': current_user.full_name,
        'current_company': current_user.current_company_id,
        'role': current_user.current_role
    })


@app.route('/api/companies')
@login_required
def api_companies():
    """Get user's companies"""
    return jsonify(current_user.companies)


@app.route('/api/data-export')
@login_required
def api_data_export():
    """Export user data (GDPR)"""
    from data_export import DataExporter
    
    exporter = DataExporter()
    # This would generate and return the export
    # For now, return status
    return jsonify({'status': 'Export initiated'})


@app.route('/api/employee-names')
@login_required
@company_required
def api_employee_names():
    print("[DEBUG] api_employee_names called")
    """Get employee names for autocomplete"""
    try:
        # Read employee roster from company data
        company_id = current_user.current_company_id
        employee_file = f"company_data/{company_id}/employees.json"
        print(f"[DEBUG] employee_file path: {employee_file}")

        employees = []
        removedEmployees = []
        print(f"[DEBUG] Checking if employee_file exists: {os.path.exists(employee_file)}")
        if not os.path.exists(employee_file):
            print(f"[ERROR] employee_file does not exist at path: {employee_file}")
        else:
            raw = None
            try:
                with open(employee_file, 'r') as f:
                    raw = f.read()
            except Exception as e:
                print(f"[ERROR] Failed to read employees.json: {e}")
            print(f"[DEBUG] RAW employee_file contents: {raw}")
            try:
                if raw:
                    import json
                    data = json.loads(raw)
                    print(f"[DEBUG] Parsed employee_file contents: {data}")
                    employees = data.get('employees', [])
                    removedEmployees = data.get('removedEmployees', [])
            except Exception as e:
                print(f"[ERROR] Failed to parse employees.json: {e}")

        # Build list of full names for autocomplete
        employee_names = [f"{emp.get('firstName', '')} {emp.get('lastName', '')}".strip() for emp in employees if emp.get('firstName') and emp.get('lastName')]
        print(f"[DEBUG] employees returned: {employees}")
        print(f"[DEBUG] employee_names for autocomplete: {employee_names}")
        return jsonify({
            'success': True,
            'employees': employees,
            'employeeNames': employee_names,
            'removedEmployees': removedEmployees
        })

    except Exception as e:
        print(f"[ERROR] api_employee_names: {e}")
        return jsonify({
            'success': False, 
            'error': str(e), 
            'employees': [],
            'removedEmployees': []
        }), 200


@app.route('/api/save-employees', methods=['POST'])
@login_required
@company_required
def api_save_employees():
    """Save employee roster"""
    try:
        import json
        company_id = current_user.current_company_id
        data_dir = f"company_data/{company_id}"
        os.makedirs(data_dir, exist_ok=True)
        
        employee_file = f"{data_dir}/employees.json"
        employees = request.json.get('employees', [])
        removedEmployees = request.json.get('removedEmployees', [])
        
        with open(employee_file, 'w') as f:
            json.dump({
                'employees': employees,
                'removedEmployees': removedEmployees
            }, f, indent=2)

        # Create individual JSON file for each employee
        for emp in employees:
            # Ensure required fields exist
            if emp.get('firstName') and emp.get('lastName') and emp.get('dateOfHire'):
                # Add an 'id' field if missing
                if 'id' not in emp:
                    emp['id'] = f"{emp['firstName']}_{emp['lastName']}"
                create_employee_json(company_id, emp)

        return jsonify({'success': True})
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== FILE IMPORT API ====================

@app.route('/api/import-file', methods=['POST'])
@login_required
def import_file():
    """Import Excel or CSV file with sales data"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400
        
        filename = file.filename.lower()
        
        if filename.endswith('.xlsx'):
            data = parse_excel_file(file)
        elif filename.endswith('.csv'):
            data = parse_csv_file(file)
        else:
            return jsonify({'success': False, 'error': 'Unsupported file type. Please upload .xlsx or .csv'}), 400
        
        return jsonify({'success': True, 'data': data})
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/import-employees', methods=['POST'])
@login_required
def import_employees():
    """Import employees from Excel or CSV file"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400
        
        filename = file.filename.lower()
        employees = []
        
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            employees = parse_employee_excel(file)
        elif filename.endswith('.csv'):
            employees = parse_employee_csv(file)
        else:
            return jsonify({'success': False, 'error': 'Unsupported file type. Please upload Excel or CSV'}), 400
        
        return jsonify({'success': True, 'employees': employees})
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


def parse_employee_excel(file):
    """Parse employee data from Excel file"""
    import openpyxl
    from datetime import datetime
    
    employees = []
    workbook = openpyxl.load_workbook(file, data_only=True)
    
    # Try first sheet
    sheet = workbook.active
    
    # Find header row
    headers = []
    header_row_idx = 0
    
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if any(cell and isinstance(cell, str) and any(keyword in cell.lower() for keyword in ['first', 'last', 'name', 'hire']) for cell in row):
            headers = [str(cell).strip().lower() if cell else '' for cell in row]
            header_row_idx = row_idx
            break
    
    if not headers:
        return employees
    
    # Find column indices
    first_name_idx = None
    last_name_idx = None
    date_of_hire_idx = None
    
    for idx, header in enumerate(headers):
        if ('first' in header and 'name' in header) or header == 'firstname':
            first_name_idx = idx
        elif ('last' in header and 'name' in header) or header == 'lastname':
            last_name_idx = idx
        elif ('date' in header and 'hire' in header) or 'hire' in header or ('start' in header and 'date' in header):
            date_of_hire_idx = idx
    
    if first_name_idx is None or last_name_idx is None:
        return employees
    
    # Parse data rows
    for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if len(row) > max(first_name_idx, last_name_idx):
            first_name = str(row[first_name_idx]).strip() if row[first_name_idx] else None
            last_name = str(row[last_name_idx]).strip() if row[last_name_idx] else None
            
            if first_name and last_name and first_name.lower() not in ['none', ''] and last_name.lower() not in ['none', '']:
                date_of_hire = None
                
                if date_of_hire_idx is not None and len(row) > date_of_hire_idx and row[date_of_hire_idx]:
                    hire_value = row[date_of_hire_idx]
                    
                    if isinstance(hire_value, datetime):
                        date_of_hire = hire_value.strftime('%Y-%m-%d')
                    elif isinstance(hire_value, str):
                        try:
                            parsed_date = datetime.strptime(hire_value, '%Y-%m-%d')
                            date_of_hire = parsed_date.strftime('%Y-%m-%d')
                        except:
                            try:
                                parsed_date = datetime.strptime(hire_value, '%m/%d/%Y')
                                date_of_hire = parsed_date.strftime('%Y-%m-%d')
                            except:
                                date_of_hire = datetime.now().strftime('%Y-%m-%d')
                
                if not date_of_hire:
                    date_of_hire = datetime.now().strftime('%Y-%m-%d')
                
                employees.append({
                    'firstName': first_name,
                    'lastName': last_name,
                    'dateOfHire': date_of_hire
                })
    
    return employees


def parse_employee_csv(file):
    """Parse employee data from CSV file"""
    import csv
    from datetime import datetime
    from io import StringIO
    
    employees = []
    content = file.read().decode('utf-8')
    reader = csv.reader(StringIO(content))
    
    rows = list(reader)
    if len(rows) == 0:
        return employees
    
    # Get headers
    headers = [h.strip().lower() for h in rows[0]]
    
    # Find column indices
    first_name_idx = None
    last_name_idx = None
    date_of_hire_idx = None
    
    for idx, header in enumerate(headers):
        if ('first' in header and 'name' in header) or header == 'firstname':
            first_name_idx = idx
        elif ('last' in header and 'name' in header) or header == 'lastname':
            last_name_idx = idx
        elif ('date' in header and 'hire' in header) or 'hire' in header or ('start' in header and 'date' in header):
            date_of_hire_idx = idx
    
    if first_name_idx is None or last_name_idx is None:
        return employees
    
    # Parse data rows
    for row in rows[1:]:
        if len(row) > max(first_name_idx, last_name_idx):
            first_name = row[first_name_idx].strip() if row[first_name_idx] else None
            last_name = row[last_name_idx].strip() if row[last_name_idx] else None
            
            if first_name and last_name:
                date_of_hire = None
                
                if date_of_hire_idx is not None and len(row) > date_of_hire_idx and row[date_of_hire_idx]:
                    hire_value = row[date_of_hire_idx].strip()
                    
                    try:
                        parsed_date = datetime.strptime(hire_value, '%Y-%m-%d')
                        date_of_hire = parsed_date.strftime('%Y-%m-%d')
                    except:
                        try:
                            parsed_date = datetime.strptime(hire_value, '%m/%d/%Y')
                            date_of_hire = parsed_date.strftime('%Y-%m-%d')
                        except:
                            date_of_hire = datetime.now().strftime('%Y-%m-%d')
                
                if not date_of_hire:
                    date_of_hire = datetime.now().strftime('%Y-%m-%d')
                
                employees.append({
                    'firstName': first_name,
                    'lastName': last_name,
                    'dateOfHire': date_of_hire
                })
    
    return employees


def parse_excel_file(file):
    """Parse Excel file like DLimport.py does"""
    import openpyxl
    
    # Load workbook
    workbook = openpyxl.load_workbook(file, data_only=True)
    
    # Initialize data structure matching DLimport.py
    data = {
        'cash': 0.0,
        'cc_tips': 0.0,
        'visa': 0.0,
        'mastercard': 0.0,
        'amex': 0.0,
        'discover': 0.0,
        'liquor': 0.0,
        'beer': 0.0,
        'wine': 0.0,
        'food': 0.0,
        'voids': 0.0
    }
    
    # Search through all sheets for data
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet_name_lower = sheet_name.lower()
        
        # Parse "Cash activity" sheet - gets actual cash values
        if "cash" in sheet_name_lower and "activity" in sheet_name_lower:
            cash_payments, cc_tips = parse_cash_activity_sheet(sheet)
            if cash_payments is not None:
                data['cash'] = cash_payments
            if cc_tips is not None:
                data['cc_tips'] = abs(cc_tips)  # CC tips are usually negative
        
        # Parse "All Data" sheet (fallback if Cash activity not found)
        elif "all data" in sheet_name_lower and data['cash'] == 0:
            cash, cc_tips = parse_all_data_sheet(sheet)
            if cash is not None:
                data['cash'] = cash
            if cc_tips is not None:
                data['cc_tips'] = cc_tips
        
        # Parse "Payment Summary" sheet
        if "payment" in sheet_name_lower and "summary" in sheet_name_lower:
            visa, mc, amex, disc, tips = parse_payment_summary_sheet(sheet)
            if visa is not None:
                data['visa'] = visa
            if mc is not None:
                data['mastercard'] = mc
            if amex is not None:
                data['amex'] = amex
            if disc is not None:
                data['discover'] = disc
            # Use CC tips from payment summary if not found in cash activity
            if tips is not None and data['cc_tips'] == 0:
                data['cc_tips'] = tips
        
        # Parse "Sales Category Summary" sheet from "All data"
        if "all data" in sheet_name_lower:
            liquor, beer, wine, food = parse_sales_category_sheet(sheet)
            if liquor is not None:
                data['liquor'] = liquor
            if beer is not None:
                data['beer'] = beer
            if wine is not None:
                data['wine'] = wine
            if food is not None:
                data['food'] = food
            
            # Parse voids from "All data" sheet
            voids = parse_voids_summary_sheet(sheet)
            if voids is not None:
                data['voids'] = voids
    
    # Calculate credit total
    data['credit_total'] = data['visa'] + data['mastercard'] + data['amex'] + data['discover']
    
    return data


def parse_cash_activity_sheet(sheet):
    """Parse 'Cash activity' sheet for cash payments and CC tips"""
    cash_payments = None
    cc_tips = None
    
    try:
        # Find header row
        header_row = None
        for row_idx, row in enumerate(sheet.iter_rows()):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and "total cash payment" in cell.value.lower():
                    header_row = row_idx
                    break
            if header_row is not None:
                break
        
        if header_row is None:
            return None, None
        
        # Get column headers and data
        headers = [cell.value for cell in list(sheet.iter_rows())[header_row]]
        data_row = list(sheet.iter_rows())[header_row + 1]
        
        # Find "Total cash payments" column
        for idx, header in enumerate(headers):
            if header and isinstance(header, str):
                header_lower = header.lower()
                if "total cash payment" in header_lower:
                    if idx < len(data_row) and isinstance(data_row[idx].value, (int, float)):
                        cash_payments = float(data_row[idx].value)
                
                # Look for "Credit/non-cash tips"
                if "credit" in header_lower and "non" in header_lower and "tip" in header_lower:
                    if idx < len(data_row) and isinstance(data_row[idx].value, (int, float)):
                        cc_tips = abs(float(data_row[idx].value))  # Get absolute value
    except Exception as e:
        print(f"Error parsing Cash activity sheet: {e}")
    
    return cash_payments, cc_tips


def parse_all_data_sheet(sheet):
    """Parse 'All Data' sheet for cash and CC tips"""
    cash = None
    cc_tips = None
    
    try:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    cell_lower = cell.value.lower()
                    
                    # Look for "Cash" row
                    if "cash" in cell_lower and "total" not in cell_lower:
                        # Amount should be in next cell or same row
                        for check_cell in row:
                            if check_cell != cell and isinstance(check_cell.value, (int, float)):
                                if cash is None:
                                    cash = float(check_cell.value)
                                break
                    
                    # Look for "CC Tips" or "Credit Card Tips"
                    if ("cc" in cell_lower or "credit card" in cell_lower) and "tip" in cell_lower:
                        for check_cell in row:
                            if check_cell != cell and isinstance(check_cell.value, (int, float)):
                                if cc_tips is None:
                                    cc_tips = float(check_cell.value)
                                break
    except Exception as e:
        print(f"Error parsing All Data sheet: {e}")
    
    return cash, cc_tips


def parse_payment_summary_sheet(sheet):
    """Parse 'Payment Summary' sheet for card types and tips"""
    visa = None
    mastercard = None
    amex = None
    discover = None
    total_tips = None
    
    try:
        # Find header row
        header_row = None
        headers = []
        for row_idx, row in enumerate(sheet.iter_rows()):
            row_values = [cell.value for cell in row]
            if any(val and isinstance(val, str) and "payment sub type" in val.lower() for val in row_values):
                header_row = row_idx
                headers = row_values
                break
        
        if header_row is None:
            return None, None, None, None, None
        
        # Find column indices
        subtype_col = None
        amount_col = None
        tips_col = None
        
        for idx, header in enumerate(headers):
            if header and isinstance(header, str):
                header_lower = header.lower()
                if "payment sub type" in header_lower:
                    subtype_col = idx
                elif header_lower == "amount":
                    amount_col = idx
                elif "tips" in header_lower and "grat" not in header_lower:
                    tips_col = idx
        
        # Parse data rows
        for row in list(sheet.iter_rows())[header_row + 1:]:
            row_values = [cell.value for cell in row]
            
            # Get payment sub type and amount
            if subtype_col is not None and subtype_col < len(row_values):
                subtype = row_values[subtype_col]
                
                if subtype and isinstance(subtype, str):
                    subtype_lower = subtype.lower()
                    
                    # Get amount from Amount column
                    if amount_col is not None and amount_col < len(row_values):
                        amount = row_values[amount_col]
                        
                        if isinstance(amount, (int, float)) and amount > 0:
                            if "visa" in subtype_lower:
                                visa = float(amount)
                            elif "mastercard" in subtype_lower:
                                mastercard = float(amount)
                            elif "amex" in subtype_lower:
                                amex = float(amount)
                            elif "discover" in subtype_lower:
                                discover = float(amount)
            
            # Get total tips from Credit/debit row
            payment_type = row_values[0] if len(row_values) > 0 else None
            if payment_type and isinstance(payment_type, str) and "credit/debit" in payment_type.lower():
                if subtype_col is not None and subtype_col < len(row_values):
                    # This is the total Credit/debit row (Payment sub type is NaN)
                    subtype_val = row_values[subtype_col]
                    if subtype_val is None or (isinstance(subtype_val, float) and str(subtype_val) == 'nan'):
                        if tips_col is not None and tips_col < len(row_values):
                            tips_val = row_values[tips_col]
                            if isinstance(tips_val, (int, float)):
                                total_tips = float(tips_val)
    except Exception as e:
        print(f"Error parsing Payment Summary sheet: {e}")
    
    return visa, mastercard, amex, discover, total_tips


def parse_sales_category_sheet(sheet):
    """Parse 'All data' sheet for sales by category with Net sales column"""
    liquor = None
    beer = None
    wine = None
    food = None
    
    try:
        # Find the "Sales category summary" section header
        header_row_idx = None
        for row_idx, row in enumerate(sheet.iter_rows()):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and "sales category summary" in cell.value.lower():
                    header_row_idx = row_idx
                    break
            if header_row_idx is not None:
                break
        
        if header_row_idx is None:
            return None, None, None, None
        
        # Find the column headers row (next row after section header)
        rows_list = list(sheet.iter_rows())
        column_headers_row = rows_list[header_row_idx + 1] if header_row_idx + 1 < len(rows_list) else None
        
        if column_headers_row is None:
            return None, None, None, None
        
        # Find "Net sales" column index
        net_sales_col_idx = None
        for idx, cell in enumerate(column_headers_row):
            if cell.value and isinstance(cell.value, str) and "net sales" in cell.value.lower():
                net_sales_col_idx = idx
                break
        
        if net_sales_col_idx is None:
            return None, None, None, None
        
        # Parse data rows starting from header_row_idx + 2
        for row_idx in range(header_row_idx + 2, len(rows_list)):
            row = rows_list[row_idx]
            
            # First column has category name
            category_cell = row[0]
            if not category_cell.value or not isinstance(category_cell.value, str):
                continue
            
            category = category_cell.value.lower()
            
            # Check if we've reached the end of this section (Total row or empty)
            if "total" in category or category.strip() == "":
                break
            
            # Get Net sales value
            if net_sales_col_idx < len(row):
                net_sales_cell = row[net_sales_col_idx]
                if isinstance(net_sales_cell.value, (int, float)):
                    amount = float(net_sales_cell.value)
                    
                    # Categorize by name
                    if "beer" in category:
                        beer = amount
                    elif "liquor" in category or "spirits" in category:
                        liquor = amount
                    elif "wine" in category:
                        wine = amount
                    elif "food" in category:
                        food = amount
    except Exception as e:
        print(f"Error parsing Sales category summary section: {e}")
    
    return liquor, beer, wine, food


def parse_voids_summary_sheet(sheet):
    """Parse voids from All data sheet - void summary section"""
    voids = 0.0
    
    try:
        # Search for "void amount" anywhere in the sheet
        void_amount_row = None
        for row in sheet.iter_rows(values_only=False):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if "void amount" in cell.value.lower():
                        void_amount_row = cell.row
                        print(f"Found 'void amount' at row {cell.row}")
                        break
            if void_amount_row is not None:
                break
        
        if void_amount_row is None:
            print("'void amount' text not found in sheet")
            return voids
        
        # Get the value from column B of the same row
        void_cell = sheet.cell(row=void_amount_row, column=2)  # Column B = 2
        print(f"Void cell at row {void_amount_row}, column B: value={void_cell.value}, type={type(void_cell.value)}")
        
        if void_cell.value:
            if isinstance(void_cell.value, (int, float)):
                voids = abs(float(void_cell.value))
                print(f"Parsed void amount: {voids}")
            else:
                # Try to convert string to float
                try:
                    voids = abs(float(str(void_cell.value).replace('$', '').replace(',', '').strip()))
                    print(f"Parsed void amount from string: {voids}")
                except:
                    print(f"Could not convert void value to number: {void_cell.value}")
    
    except Exception as e:
        print(f"Error parsing Void summary section: {e}")
        import traceback
        traceback.print_exc()
    
    return voids


def parse_csv_file(file):
    """Parse CSV file with simple structure"""
    import csv
    from io import StringIO
    
    data = {
        'employees': [],
        'cash_sales': 0.0,
        'visa': 0.0,
        'mastercard': 0.0,
        'amex': 0.0,
        'other_income': 0.0
    }
    
    try:
        content = file.read().decode('utf-8')
        reader = csv.reader(StringIO(content))
        
        section = None
        for row in reader:
            if not row or all(not cell.strip() for cell in row):
                continue
            
            # Detect sections
            if row[0].strip().lower() == 'employee hours':
                section = 'employees'
                continue
            elif row[0].strip().lower() == 'sales summary':
                section = 'sales'
                continue
            
            # Parse data based on section
            if section == 'employees' and len(row) >= 2:
                try:
                    name = row[0].strip()
                    hours = float(row[1].strip())
                    data['employees'].append({'name': name, 'hours': hours})
                except ValueError:
                    pass
            
            elif section == 'sales' and len(row) >= 2:
                key = row[0].strip().lower()
                try:
                    value = float(row[1].strip())
                    if 'cash' in key:
                        data['cash_sales'] = value
                    elif 'visa' in key:
                        data['visa'] = value
                    elif 'mastercard' in key:
                        data['mastercard'] = value
                    elif 'amex' in key:
                        data['amex'] = value
                    elif 'other' in key or 'income' in key:
                        data['other_income'] = value
                except ValueError:
                    pass
    
    except Exception as e:
        raise Exception(f"CSV parsing error: {str(e)}")
    
    return data


# ==================== ERROR HANDLERS ====================

@app.errorhandler(404)
def not_found(e):
    return render_template('404.html'), 404


@app.errorhandler(500)
def server_error(e):
    return render_template('500.html'), 500


@app.errorhandler(403)
def forbidden(e):
    return render_template('403.html'), 403


@app.route('/employee-json/<company_id>/<filename>')
@login_required
@company_required
def serve_employee_json(company_id, filename):
    emp_dir = os.path.join('company_data', company_id, 'employees')
    # Security: only allow .json files, no path traversal
    if not filename.endswith('.json') or '/' in filename or '\\' in filename:
        return "Invalid file request", 400
    return send_from_directory(emp_dir, filename, mimetype='application/json')


@app.route('/employee/<company_id>/<filename>')
@login_required
@company_required
def employee_profile(company_id, filename):
    import json
    import os
    base_dir = os.path.dirname(os.path.abspath(__file__))
    emp_dir = os.path.join(base_dir, 'company_data', company_id, 'employees')
    filepath = os.path.join(emp_dir, filename)
    print(f"[DEBUG] Absolute file path: {filepath}")
    if not filename.endswith('.json') or '/' in filename or '\\' in filename:
        flash('Invalid file request.', 'danger')
        return redirect(url_for('employees'))
    if not os.path.exists(filepath):
        print(f"[ERROR] Employee file not found at: {filepath}")
        flash('Employee file not found.', 'danger')
        return redirect(url_for('employees'))
    with open(filepath, 'r') as f:
        employee = json.load(f)
    return render_template('employee_profile.html', employee=employee)


if __name__ == '__main__':
    # Development server
    app.run(debug=True, host='0.0.0.0', port=8000, use_reloader=False)
