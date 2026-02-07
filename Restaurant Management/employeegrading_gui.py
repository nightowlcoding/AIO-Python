import pandas as pd
import pdfplumber
import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from datetime import datetime

class EmployeeGradingGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Employee Grading System")
        self.root.geometry("1000x750")
        self.root.configure(bg="#e8f4f8")
        
        self.pdf_path = None
        self.grading_results = None
        self.df = None
        
        self._build_ui()
    
    def _build_ui(self):
        # Header
        header_frame = tk.Frame(self.root, bg="#1976D2", height=80)
        header_frame.pack(fill="x", pady=(0, 10))
        header_frame.pack_propagate(False)
        
        title_label = tk.Label(
            header_frame,
            text="Employee Grading System",
            font=("Arial", 24, "bold"),
            bg="#1976D2",
            fg="black"
        )
        title_label.pack(pady=20)
        
        # File selection frame
        file_frame = tk.Frame(self.root, bg="#e8f4f8")
        file_frame.pack(fill="x", padx=20, pady=10)
        
        tk.Label(file_frame, text="Selected PDF:", font=("Arial", 11, "bold"),
                bg="#e8f4f8", fg="black").pack(side="left", padx=(0, 10))
        
        self.file_label = tk.Label(file_frame, text="No file selected", 
                                   font=("Arial", 10), bg="white", fg="black",
                                   relief="solid", bd=1, padx=10, pady=5, anchor="w")
        self.file_label.pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        select_btn = tk.Button(file_frame, text="📁 Select PDF", 
                              command=self._select_pdf,
                              bg="#1976D2", fg="black", font=("Arial", 10, "bold"),
                              relief="flat", padx=20, pady=8, cursor="hand2")
        select_btn.pack(side="left")
        
        # Process button
        process_btn = tk.Button(file_frame, text="⚙️ Process & Grade", 
                               command=self._process_pdf,
                               bg="#388E3C", fg="black", font=("Arial", 10, "bold"),
                               relief="flat", padx=20, pady=8, cursor="hand2")
        process_btn.pack(side="left", padx=(10, 0))
        
        # Grade color key
        key_frame = tk.LabelFrame(self.root, text="Grade Color Key", 
                                 font=("Arial", 10, "bold"),
                                 bg="#e8f4f8", fg="black", padx=10, pady=8)
        key_frame.pack(fill="x", padx=20, pady=(0, 5))
        
        key_inner = tk.Frame(key_frame, bg="#e8f4f8")
        key_inner.pack()
        
        grades = [
            ('A', '#C8E6C9', 'Excellent'),
            ('B', '#BBDEFB', 'Good'),
            ('C', '#FFF9C4', 'Average'),
            ('D', '#FFCCBC', 'Below Average'),
            ('F', '#FFCDD2', 'Needs Improvement'),
            ('N/A', '#E0E0E0', 'Insufficient Hours')
        ]
        
        for grade, color, desc in grades:
            grade_box = tk.Frame(key_inner, bg="#e8f4f8")
            grade_box.pack(side="left", padx=8)
            
            color_label = tk.Label(grade_box, text=f"  {grade}  ", 
                                  bg=color, fg="black",
                                  font=("Arial", 10, "bold"),
                                  relief="solid", bd=1, padx=8, pady=3)
            color_label.pack(side="left", padx=(0, 5))
            
            desc_label = tk.Label(grade_box, text=desc, 
                                 bg="#e8f4f8", fg="black",
                                 font=("Arial", 9))
            desc_label.pack(side="left")
        
        # Results frame with scrollbar
        results_frame = tk.LabelFrame(self.root, text="Grading Results Preview", 
                                     font=("Arial", 11, "bold"),
                                     bg="#e8f4f8", fg="black", padx=10, pady=10)
        results_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Create Treeview for results
        tree_scroll = ttk.Scrollbar(results_frame)
        tree_scroll.pack(side="right", fill="y")
        
        # Configure style for Treeview
        style = ttk.Style()
        style.configure("Treeview", foreground="black", fieldbackground="white")
        style.configure("Treeview.Heading", foreground="black", background="#d0d0d0")
        
        self.results_tree = ttk.Treeview(results_frame, 
                                        yscrollcommand=tree_scroll.set,
                                        height=15)
        self.results_tree.pack(fill="both", expand=True)
        tree_scroll.config(command=self.results_tree.yview)
        
        # Configure treeview columns
        self.results_tree['columns'] = ('Employee', 'Sales/Hour', 'Turn Time', 'Void %', 'Void $',
                        'Tips %', 'Hours', 'Score', 'Grade')
        self.results_tree.column('#0', width=0, stretch=tk.NO)
        
        # Define column widths and headings
        column_widths = {'Employee': 140, 'Sales/Hour': 90, 'Turn Time': 80, 
                'Void %': 70, 'Void $': 80, 'Tips %': 70, 'Hours': 70, 
                'Score': 80, 'Grade': 60}
        
        for col in self.results_tree['columns']:
            self.results_tree.column(col, width=column_widths.get(col, 100), anchor='center')
            self.results_tree.heading(col, text=col, anchor='center')
        
        # Add right-click menu for editing
        self.context_menu = tk.Menu(self.results_tree, tearoff=0)
        self.context_menu.add_command(label="🗑️ Delete Selected Row", command=self._delete_selected_row)
        self.results_tree.bind("<Button-2>", self._show_context_menu)  # Right-click on Mac
        self.results_tree.bind("<Button-3>", self._show_context_menu)  # Right-click on Windows/Linux
        
        # Statistics frame
        stats_frame = tk.Frame(self.root, bg="#e8f4f8")
        stats_frame.pack(fill="x", padx=20, pady=(0, 10))
        
        self.stats_label = tk.Label(stats_frame, text="", 
                                    font=("Arial", 10), bg="#f0f8ff", fg="black",
                                    relief="solid", bd=1, padx=10, pady=10,
                                    justify="left", anchor="w")
        self.stats_label.pack(fill="x")
        
        # Action buttons frame
        action_frame = tk.Frame(self.root, bg="#e8f4f8")
        action_frame.pack(fill="x", padx=20, pady=(0, 20))
        
        save_btn = tk.Button(action_frame, text="💾 Save to Excel", 
                            command=self._save_excel,
                            bg="#F57C00", fg="black", font=("Arial", 11, "bold"),
                            relief="flat", padx=30, pady=10, cursor="hand2",
                            state="disabled")
        save_btn.pack(side="left", padx=(0, 10))
        self.save_btn = save_btn
        
        clear_btn = tk.Button(action_frame, text="🗑️ Clear", 
                             command=self._clear_results,
                             bg="#616161", fg="black", font=("Arial", 11, "bold"),
                             relief="flat", padx=30, pady=10, cursor="hand2")
        clear_btn.pack(side="left")
        
        delete_row_btn = tk.Button(action_frame, text="❌ Delete Selected", 
                                   command=self._delete_selected_row,
                                   bg="#D32F2F", fg="black", font=("Arial", 11, "bold"),
                                   relief="flat", padx=30, pady=10, cursor="hand2")
        delete_row_btn.pack(side="left", padx=(10, 0))
    
    def _show_context_menu(self, event):
        """Show context menu on right-click"""
        # Select the row under cursor
        item = self.results_tree.identify_row(event.y)
        if item:
            self.results_tree.selection_set(item)
            self.context_menu.post(event.x_root, event.y_root)
    
    def _delete_selected_row(self):
        """Delete selected row from results"""
        selected_items = self.results_tree.selection()
        
        if not selected_items:
            messagebox.showwarning("No Selection", "Please select a row to delete.")
            return
        
        # Get employee name for confirmation
        item = selected_items[0]
        values = self.results_tree.item(item, 'values')
        employee_name = values[0] if values else "this employee"
        
        # Confirm deletion
        confirm = messagebox.askyesno(
            "Confirm Deletion",
            f"Are you sure you want to remove '{employee_name}' from the grading results?\n\n"
            "This will not affect the saved Excel file until you save again."
        )
        
        if confirm:
            # Remove from treeview
            self.results_tree.delete(item)
            
            # Remove from dataframe
            if self.df is not None:
                self.df = self.df[self.df['Employee'] != employee_name].reset_index(drop=True)
                
                # Update statistics
                if len(self.df) > 0:
                    self._update_statistics()
                else:
                    self.stats_label.config(text="No employees remaining in results.")
                    self.save_btn.config(state="disabled")
    
    def _update_statistics(self):
        """Update statistics display"""
        if self.df is None or len(self.df) == 0:
            self.stats_label.config(text="")
            return
        
        grade_counts = self.df['Grade'].value_counts().sort_index()
        stats_text = f"Total Employees: {len(self.df)}\n\n"
        stats_text += "Grade Distribution:  "
        for grade in ['A', 'B', 'C', 'D', 'F', 'N/A']:
            count = grade_counts.get(grade, 0)
            stats_text += f"{grade}: {count}   "
        
        stats_text += f"\n\nAverage Score: {self.df['Weighted Score'].mean():.3f}  |  "
        stats_text += f"Highest: {self.df['Weighted Score'].max():.3f}  |  "
        stats_text += f"Lowest: {self.df['Weighted Score'].min():.3f}"
        
        self.stats_label.config(text=stats_text)
    
    def _select_pdf(self):
        """Open file dialog to select PDF"""
        pdf_path = filedialog.askopenfilename(
            title="Select Employee Productivity PDF",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
            initialdir=os.path.expanduser("~/Downloads")
        )
        
        if pdf_path:
            self.pdf_path = pdf_path
            self.file_label.config(text=os.path.basename(pdf_path), fg="black")
    
    def _process_pdf(self):
        """Process the PDF and generate grades"""
        if not self.pdf_path:
            messagebox.showwarning("No File", "Please select a PDF file first.")
            return
        
        if not os.path.exists(self.pdf_path):
            messagebox.showerror("Error", f"File not found: {self.pdf_path}")
            return
        
        try:
            # Show processing message
            self.file_label.config(text=f"Processing: {os.path.basename(self.pdf_path)}...", fg="black")
            self.root.update()
            
            # Extract and process data
            df = self._extract_and_grade(self.pdf_path)
            
            if df is None or len(df) == 0:
                messagebox.showerror("Error", "No employee data found in PDF")
                return
            
            self.df = df
            self._display_results()
            
            self.file_label.config(text=f"Processed: {os.path.basename(self.pdf_path)}", fg="black")
            self.save_btn.config(state="normal")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to process PDF:\n{str(e)}")
            self.file_label.config(text=f"Error: {os.path.basename(self.pdf_path)}", fg="black")
    
    def _extract_and_grade(self, pdf_path):
        """Extract data from PDF and apply grading"""
        # Extract tables from PDF
        all_tables = []
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    if table:
                        all_tables.append(table)
        
        if not all_tables:
            raise ValueError("No tables found in PDF")
        
        # Find the table with employee data
        df = None
        for table in all_tables:
            if len(table) > 1:
                headers = table[0]
                if any('Employee' in str(cell) or 'Sales' in str(cell) for cell in headers if cell):
                    df = pd.DataFrame(table[1:], columns=headers)
                    break
        
        if df is None:
            df = pd.DataFrame(all_tables[0][1:], columns=all_tables[0][0])
        
        # Clean column names
        df.columns = df.columns.str.replace('\n', ' ').str.strip()
        
        # Map columns
        def find_column(df, keywords):
            for col in df.columns:
                col_lower = str(col).lower()
                for keyword in keywords:
                    if keyword.lower() in col_lower:
                        return col
            return None
        
        employee_col = find_column(df, ['employee', 'name', 'server'])
        sales_col = find_column(df, ['sales'])
        void_col = find_column(df, ['void total', 'void count', 'Void Total'])
        # If not found, try exact match for 'Void Total' (case-insensitive)
        if not void_col:
            for col in df.columns:
                if col.strip().lower() == 'void total':
                    void_col = col
                    break
        hours_col = find_column(df, ['hours worked', 'hours', 'hrs'])
        turn_time_col = find_column(df, ['average turn time', 'turn time', 'avg turn'])
        tips_col = find_column(df, ['non-cash tips', 'tips %', 'non cash tips'])
        
        # Find void total specifically (case-insensitive)
        for col in df.columns:
            if 'void' in col.lower() and 'total' in col.lower():
                void_col = col
                break
        
        # Rename columns
        column_mapping = {}
        if employee_col:
            column_mapping[employee_col] = 'Employee'
        if sales_col:
            column_mapping[sales_col] = 'Sales $'
        if void_col:
            column_mapping[void_col] = 'Void total'
        if hours_col:
            column_mapping[hours_col] = 'Hours worked'
        if turn_time_col:
            column_mapping[turn_time_col] = 'Average turn time'
        if tips_col:
            column_mapping[tips_col] = 'Non-cash tips %'
        
        df = df.rename(columns=column_mapping)
        
        # Clean data
        df = df.dropna(how='all')
        df = df[df['Employee'].notna()]
        df = df[~df['Employee'].astype(str).str.lower().str.contains('employee|total|average|summary', na=False)]
        
        for col in df.columns:
            if df[col].dtype == 'object':
                df[col] = df[col].astype(str).str.replace('\n', ' ').str.strip()
        
        # Clean currency columns
        def clean_currency(col):
            return df[col].astype(str).str.replace(r'[$,]', '', regex=True).replace('--', '0').replace('None', '0').replace('nan', '0').astype(float)
        
        required_cols = ['Sales $', 'Void total', 'Hours worked']
        for col in required_cols:
            if col not in df.columns:
                df[col] = '0'
        
        df['Sales'] = clean_currency('Sales $')
        df['Void total'] = clean_currency('Void total')
        df['Hours worked'] = df['Hours worked'].astype(str).replace('--', '0').replace('None', '0').replace('nan', '0').astype(float)
        
        # Handle non-cash tips
        if 'Non-cash tips %' in df.columns:
            df['Tips %'] = df['Non-cash tips %'].astype(str).str.replace('%', '').str.replace('--', '0').replace('None', '0').replace('nan', '0').astype(float)
        else:
            df['Tips %'] = 0.0
        
        # Handle turn time
        def time_to_seconds(time_str):
            if pd.isna(time_str) or time_str == '--' or str(time_str).lower() in ['none', 'nan', '']:
                return 0
            try:
                time_str = str(time_str).strip()
                parts = time_str.split(':')
                if len(parts) >= 2:
                    minutes = int(parts[0])
                    seconds = int(parts[1])
                    return minutes * 60 + seconds
            except:
                pass
            return 0
        
        if 'Average turn time' in df.columns:
            df['Turn Time (sec)'] = df['Average turn time'].apply(time_to_seconds)
        else:
            df['Turn Time (sec)'] = 0
        
        # Calculate KPIs
        df['Sales/Hour'] = df['Sales'] / df['Hours worked'].replace(0, 1)
        df['Void %'] = (df['Void total'] / df['Sales']).replace([float('inf'), -float('inf')], 0).fillna(0)
        
        # Weighted scoring
        weights = {
            'Sales/Hour': 0.35,
            'Turn Time (sec)': 0.05,
            'Void %': 0.20,
            'Hours worked': 0.20,
            'Tips %': 0.20
        }
        
        min_max = {col: {'min': df[col].min(), 'max': df[col].max()} for col in weights.keys()}
        
        def normalize_score(value, col_name, direction):
            min_val = min_max[col_name]['min']
            max_val = min_max[col_name]['max']
            if max_val == min_val:
                return 1.0
            normalized = (value - min_val) / (max_val - min_val)
            if direction == 'higher_is_better':
                return normalized
            elif direction == 'lower_is_better':
                return 1.0 - normalized
            return 0.0
        
        df['Normalized Sales/Hour'] = df.apply(lambda row: normalize_score(row['Sales/Hour'], 'Sales/Hour', 'higher_is_better'), axis=1)
        df['Normalized Turn Time'] = df.apply(lambda row: normalize_score(row['Turn Time (sec)'], 'Turn Time (sec)', 'lower_is_better'), axis=1)
        df['Normalized Void %'] = df.apply(lambda row: normalize_score(row['Void %'], 'Void %', 'lower_is_better'), axis=1)
        df['Normalized Hours worked'] = df.apply(lambda row: normalize_score(row['Hours worked'], 'Hours worked', 'higher_is_better'), axis=1)
        df['Normalized Tips %'] = df.apply(lambda row: normalize_score(row['Tips %'], 'Tips %', 'higher_is_better'), axis=1)
        
        df['Weighted Score'] = (
            df['Normalized Sales/Hour'] * weights['Sales/Hour'] +
            df['Normalized Turn Time'] * weights['Turn Time (sec)'] +
            df['Normalized Void %'] * weights['Void %'] +
            df['Normalized Hours worked'] * weights['Hours worked'] +
            df['Normalized Tips %'] * weights['Tips %']
        )
        
        # Grade assignment
        def assign_grade(row):
            score = row['Weighted Score']
            tips_pct = row['Tips %']
            void_pct = row['Void %']
            hours = row['Hours worked']
            # N/A grade for employees who didn't work enough hours
            if hours < 12:
                return 'N/A'
            if score >= 0.90:
                base_grade = 'A'
            elif score >= 0.80:
                base_grade = 'B'
            elif score >= 0.70:
                base_grade = 'C'
            elif score >= 0.60:
                base_grade = 'D'
            else:
                base_grade = 'F'
            # Tips adjustments
            if tips_pct >= 15:
                if base_grade == 'B':
                    base_grade = 'A'
                elif base_grade == 'C':
                    base_grade = 'B'
                elif base_grade == 'D':
                    base_grade = 'C'
                elif base_grade == 'F':
                    base_grade = 'D'
            elif tips_pct < 6:
                if base_grade == 'A':
                    base_grade = 'B'
                elif base_grade == 'B':
                    base_grade = 'C'
                elif base_grade == 'C':
                    base_grade = 'D'
                elif base_grade == 'D':
                    base_grade = 'F'
            # Void adjustments
            if void_pct == 0:
                if base_grade == 'B':
                    base_grade = 'A'
                elif base_grade == 'C':
                    base_grade = 'B'
                elif base_grade == 'D':
                    base_grade = 'C'
                elif base_grade == 'F':
                    base_grade = 'D'
            elif void_pct >= 0.08:
                if base_grade == 'A':
                    base_grade = 'B'
                elif base_grade == 'B':
                    base_grade = 'C'
                elif base_grade == 'C':
                    base_grade = 'D'
                elif base_grade == 'D':
                    base_grade = 'F'
            elif void_pct >= 0.06:
                if base_grade == 'A':
                    base_grade = 'B'
                elif base_grade == 'B':
                    base_grade = 'C'
                elif base_grade == 'C':
                    base_grade = 'D'
            elif void_pct >= 0.04:
                if base_grade == 'A':
                    base_grade = 'B'
                elif base_grade == 'B':
                    base_grade = 'C'
            elif void_pct >= 0.02:
                if base_grade == 'A':
                    base_grade = 'B'
            return base_grade
        
        df['Grade'] = df.apply(assign_grade, axis=1)
        
        # Create results sorted by grade then score
        grade_order = {'A': 1, 'B': 2, 'C': 3, 'D': 4, 'F': 5, 'N/A': 6}
        df['Grade_Order'] = df['Grade'].map(grade_order)
        grading_results = df[['Employee', 'Sales/Hour', 'Turn Time (sec)', 'Void %', 'Void total', 'Tips %', 
                              'Hours worked', 'Weighted Score', 'Grade', 'Grade_Order']].sort_values(
                                  by=['Grade_Order', 'Weighted Score'], ascending=[True, False])
        grading_results = grading_results.drop('Grade_Order', axis=1)
        
        return grading_results
    
    def _display_results(self):
        """Display results in the treeview"""
        # Clear existing items
        for item in self.results_tree.get_children():
            self.results_tree.delete(item)
        
        if self.df is None or len(self.df) == 0:
            return
        
        # Add color tags for grades
        self.results_tree.tag_configure('A', background='#C8E6C9', foreground='black')
        self.results_tree.tag_configure('B', background='#BBDEFB', foreground='black')
        self.results_tree.tag_configure('C', background='#FFF9C4', foreground='black')
        self.results_tree.tag_configure('D', background='#FFCCBC', foreground='black')
        self.results_tree.tag_configure('F', background='#FFCDD2', foreground='black')
        self.results_tree.tag_configure('N/A', background='#E0E0E0', foreground='black')
        
        # Insert data
        for idx, row in self.df.iterrows():
            void_dollar = row.get('Void total', row.get('Void Total', 0.0))
            values = (
                row['Employee'],
                f"{row['Sales/Hour']:.2f}",
                f"{row['Turn Time (sec)']:.0f}",
                f"{row['Void %']:.3f}",
                f"${void_dollar:.2f}",
                f"{row['Tips %']:.2f}",
                f"{row['Hours worked']:.2f}",
                f"{row['Weighted Score']:.3f}",
                row['Grade']
            )
            self.results_tree.insert('', 'end', values=values, tags=(row['Grade'],))
        
        # Update statistics
        self._update_statistics()
    
    def _save_excel(self):
        """Save results to Excel with formatting"""
        if self.df is None or len(self.df) == 0:
            messagebox.showwarning("No Data", "No grading results to save.")
            return
        
        # Suggest filename
        default_filename = f"employee_grades_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        filename = filedialog.asksaveasfilename(
            title="Save Grading Results",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=default_filename,
            initialdir=os.path.expanduser("~/Documents/AIO Python")
        )
        
        if filename:
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils.dataframe import dataframe_to_rows
                
                # Create workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Employee Grades"
                
                # Define colors for grades
                grade_colors = {
                    'A': 'C8E6C9',
                    'B': 'BBDEFB',
                    'C': 'FFF9C4',
                    'D': 'FFCCBC',
                    'F': 'FFCDD2',
                    'N/A': 'E0E0E0'
                }
                
                # Define border style
                thin_border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                
                # Write data to worksheet
                for r_idx, row in enumerate(dataframe_to_rows(self.df, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Header row formatting
                        if r_idx == 1:
                            cell.font = Font(bold=True, size=11, color='FFFFFF')
                            cell.fill = PatternFill(start_color='1976D2', end_color='1976D2', fill_type='solid')
                        else:
                            # Data row formatting
                            cell.font = Font(size=10)
                            
                            # Get grade for this row
                            grade_col_idx = list(self.df.columns).index('Grade') + 1
                            grade = ws.cell(row=r_idx, column=grade_col_idx).value
                            
                            # Apply grade color to entire row
                            if grade in grade_colors:
                                cell.fill = PatternFill(start_color=grade_colors[grade], 
                                                       end_color=grade_colors[grade], 
                                                       fill_type='solid')
                            
                            # Make grade column bold
                            if c_idx == grade_col_idx:
                                cell.font = Font(bold=True, size=11)
                
                # Adjust column widths
                column_widths = {
                    'Employee': 20,
                    'Sales/Hour': 12,
                    'Turn Time (sec)': 15,
                    'Void %': 10,
                    'Tips %': 10,
                    'Hours worked': 13,
                    'Weighted Score': 15,
                    'Grade': 8
                }
                
                for idx, col in enumerate(self.df.columns, 1):
                    ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = column_widths.get(col, 12)
                
                # Add summary section
                summary_start_row = len(self.df) + 3
                
                # Summary title
                summary_cell = ws.cell(row=summary_start_row, column=1, value="GRADING FORMULA SUMMARY")
                summary_cell.font = Font(bold=True, size=14, color='FFFFFF')
                summary_cell.fill = PatternFill(start_color='1976D2', end_color='1976D2', fill_type='solid')
                summary_cell.alignment = Alignment(horizontal='left', vertical='center')
                ws.merge_cells(start_row=summary_start_row, start_column=1, 
                              end_row=summary_start_row, end_column=len(self.df.columns))
                
                summary_data = [
                    "",
                    "BASE WEIGHTED SCORE CALCULATION:",
                    "• Sales per Hour: 35% weight (higher is better)",
                    "• Void %: 20% weight (lower is better)",
                    "• Non-Cash Tips %: 20% weight (higher is better)",
                    "• Hours Worked: 20% weight (higher is better)",
                    "• Turn Time: 5% weight (lower is better)",
                    "",
                    "BASE GRADE THRESHOLDS (from weighted score):",
                    "• N/A: Less than 12 hours worked (insufficient hours for grading)",
                    "• A: 0.90 and above",
                    "• B: 0.80 to 0.89",
                    "• C: 0.70 to 0.79",
                    "• D: 0.60 to 0.69",
                    "• F: Below 0.60",
                    "",
                    "GRADE ADJUSTMENTS:",
                    "",
                    "Non-Cash Tips % Adjustments:",
                    "• 15% or higher tips: Grade BOOSTED up one level (B→A, C→B, D→C, F→D)",
                    "• Below 6% tips: Grade REDUCED down one level (A→B, B→C, C→D, D→F)",
                    "",
                    "Void % Adjustments:",
                    "• 0% voids: Grade BOOSTED up one level (perfect performance)",
                    "• 0-2% voids: Minimal or no penalty (A range)",
                    "• 2-4% voids: Slight penalty (A→B only)",
                    "• 4-6% voids: Moderate penalty (A→B, B→C)",
                    "• 6-8% voids: High penalty (A→B, B→C, C→D)",
                    "• 8%+ voids: Maximum penalty - Grade REDUCED down one level",
                    "",
                    "HOW TO IMPROVE YOUR GRADE:",
                    "1. Increase sales per hour (35% weight - HIGHEST PRIORITY!)",
                    "2. Work more consistent hours (20% weight - commitment!)",
                    "3. Minimize voids - aim for 0% (20% weight + grade boost)",
                    "4. Keep tips above 15% (20% weight + grade boost)",
                    "5. Reduce turn time - serve guests faster (5% weight)"
                ]
                
                for idx, text in enumerate(summary_data, start=summary_start_row + 1):
                    cell = ws.cell(row=idx, column=1, value=text)
                    if text.startswith(("BASE", "GRADE", "Non-Cash", "Void %", "HOW TO")):
                        cell.font = Font(bold=True, size=11)
                    else:
                        cell.font = Font(size=10)
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    ws.merge_cells(start_row=idx, start_column=1, 
                                  end_row=idx, end_column=len(self.df.columns))
                
                # Save workbook
                wb.save(filename)
                
                messagebox.showinfo("Success", f"Grading results saved to:\n{filename}")
            
            except Exception as e:
                messagebox.showerror("Error", f"Failed to save file:\n{str(e)}")
    
    def _clear_results(self):
        """Clear all results"""
        for item in self.results_tree.get_children():
            self.results_tree.delete(item)
        
        self.stats_label.config(text="")
        self.file_label.config(text="No file selected", fg="black")
        self.pdf_path = None
        self.df = None
        self.save_btn.config(state="disabled")


if __name__ == "__main__":
    root = tk.Tk()
    app = EmployeeGradingGUI(root)
    root.mainloop()
