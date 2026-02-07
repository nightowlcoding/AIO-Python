import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter


def process_payroll(csv_path):
    df = pd.read_csv(csv_path)
    columns_to_extract = [
        'Employee', 'Job Title', 'Regular Hours', 'Overtime Hours', 
        'Declared Tips', 'Non-Cash Tips', 'Total Tips'
    ]
    extracted_data = df[columns_to_extract]

    def reformat_employee_name(name):
        parts = [part.strip() for part in str(name).split(',')]
        if len(parts) == 2:
            return f"{parts[1]} {parts[0]}"
        return name
    extracted_data.loc[:, 'Employee'] = extracted_data['Employee'].apply(reformat_employee_name)

    # Add custom employees
    new_row = {
        'Employee': 'Arnold Ramirez',
        'Job Title': 'General Manager',
        'Regular Hours': 0,
        'Overtime Hours': 0,
        'Declared Tips': 0,
        'Non-Cash Tips': 0,
        'Total Tips': 0
    }
    new_row2 = {
        'Employee': 'Victor Moreno',
        'Job Title': 'Director of Catering',
        'Regular Hours': 0,
        'Overtime Hours': 0,
        'Declared Tips': 0,
        'Non-Cash Tips': 0,
        'Total Tips': 0
    }
    extracted_data = pd.concat([extracted_data, pd.DataFrame([new_row, new_row2])], ignore_index=True)

    # Calculate totals
    totals = extracted_data.iloc[:, 2:].sum()
    total_combined_hours = float(totals['Regular Hours']) + float(totals['Overtime Hours'])
    totals_row = pd.DataFrame([{
        'Employee': 'Total',
        'Job Title': total_combined_hours,
        'Regular Hours': totals['Regular Hours'],
        'Overtime Hours': totals['Overtime Hours'],
        'Declared Tips': totals['Declared Tips'],
        'Non-Cash Tips': totals['Non-Cash Tips'],
        'Total Tips': totals['Total Tips']
    }])


    sorted_data = extracted_data.sort_values(by='Employee')
    sorted_data = pd.concat([
        sorted_data, totals_row
    ], ignore_index=True)

    reg_hours = totals['Regular Hours']
    ot_hours = totals['Overtime Hours']
    total_hours = total_combined_hours
    ot_pct = (ot_hours / total_hours) if total_hours > 0 else 0
    # Replace zeros with empty string for display
    sorted_data = sorted_data.replace(0, "")
    return reg_hours, ot_hours, total_hours, ot_pct, sorted_data


def get_unique_excel_path(base_name):
    desktop = os.path.join(os.path.expanduser('~'), 'Desktop')
    name, ext = os.path.splitext(base_name)
    out_path = os.path.join(desktop, base_name)
    counter = 1
    while os.path.exists(out_path):
        out_path = os.path.join(desktop, f"{name} ({counter}){ext}")
        counter += 1
    return out_path

def save_to_excel(df, out_path):
    df.to_excel(out_path, index=False)
    # Apply borders and fit to one page
    wb = load_workbook(out_path)
    ws = wb.active
    max_row = ws.max_row
    max_col = ws.max_column
    thick = Side(border_style="thick", color="000000")
    thin = Side(border_style="thin", color="000000")
    # Add thick border around all data
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            # Only add outside border for cells with text
            if cell.value not in (None, ""):
                border = Border(
                    left=thick if col == 1 else thin,
                    right=thick if col == max_col else thin,
                    top=thick if row == 1 else thin,
                    bottom=thick if row == max_row else thin
                )
                cell.border = border
    # Fit to one page when printing
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    wb.save(out_path)


def run_processing():
    file1 = file1_var.get()
    file2 = file2_var.get()
    if not file1 or not file2:
        messagebox.showerror("Missing File", "Please select both CSV files.")
        return
    # Process BHB Kingsville
    reg1, ot1, total1, pct1, df1 = process_payroll(file1)
    summary1 = f"Total hours for the BHB Kingsville for the week is {total1:.2f} of which {ot1:.2f} or {pct1:.2%} is considered overtime."
    # Add summary row to Excel
    summary_row1 = pd.DataFrame([{col: "" for col in df1.columns}])
    summary_row1.iloc[0, 0] = summary1
    df1_with_summary = pd.concat([df1, summary_row1], ignore_index=True)
    out1 = get_unique_excel_path("BHB_Kingsville_sorted.xlsx")
    save_to_excel(df1_with_summary, out1)

    # Process BHB Alice
    reg2, ot2, total2, pct2, df2 = process_payroll(file2)
    summary2 = f"Total hours for the BHB Alice for the week is {total2:.2f} of which {ot2:.2f} or {pct2:.2%} is considered overtime."
    summary_row2 = pd.DataFrame([{col: "" for col in df2.columns}])
    summary_row2.iloc[0, 0] = summary2
    df2_with_summary = pd.concat([df2, summary_row2], ignore_index=True)
    out2 = get_unique_excel_path("BHB_Alice_sorted.xlsx")
    save_to_excel(df2_with_summary, out2)

    # Print summary
    summary = summary1 + "\n" + summary2
    messagebox.showinfo("Payroll Summary", summary)

# GUI setup
root = tk.Tk()
root.title("Payroll Processor")
root.geometry("500x250")

file1_var = tk.StringVar()
file2_var = tk.StringVar()


def select_file(var):
    path = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv")])
    if path:
        var.set(path)

def upload_file(var, label):
    src = var.get()
    if not src:
        messagebox.showerror("No file selected", f"Please select a CSV file for {label}.")
        return
    upload_dir = os.path.join(os.path.dirname(__file__), "UploadedCSVs")
    os.makedirs(upload_dir, exist_ok=True)
    dest = os.path.join(upload_dir, os.path.basename(src))
    try:
        import shutil
        shutil.copy2(src, dest)
        messagebox.showinfo("Upload Complete", f"{label} file uploaded to {upload_dir}.")
    except Exception as e:
        messagebox.showerror("Upload Failed", f"Error uploading file: {e}")

frame = tk.Frame(root)
frame.pack(pady=20)


lbl1 = tk.Label(frame, text="BHB Kingsville CSV:")
lbl1.grid(row=0, column=0, sticky="e")
entry1 = tk.Entry(frame, textvariable=file1_var, width=40)
entry1.grid(row=0, column=1)
btn1 = tk.Button(frame, text="Browse", command=lambda: select_file(file1_var))
btn1.grid(row=0, column=2)
upload1 = tk.Button(frame, text="Upload", command=lambda: upload_file(file1_var, "BHB Kingsville"))
upload1.grid(row=0, column=3)


lbl2 = tk.Label(frame, text="BHB Alice CSV:")
lbl2.grid(row=1, column=0, sticky="e")
entry2 = tk.Entry(frame, textvariable=file2_var, width=40)
entry2.grid(row=1, column=1)
btn2 = tk.Button(frame, text="Browse", command=lambda: select_file(file2_var))
btn2.grid(row=1, column=2)
upload2 = tk.Button(frame, text="Upload", command=lambda: upload_file(file2_var, "BHB Alice"))
upload2.grid(row=1, column=3)

run_btn = tk.Button(root, text="Process Payroll", command=run_processing, height=2, width=20)
run_btn.pack(pady=20)

root.mainloop()
